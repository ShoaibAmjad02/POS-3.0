import io
import json
import secrets
import string
import uuid
import subprocess
import traceback
import qrcode
from django.http import HttpResponse, JsonResponse, FileResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.db import transaction
from django.db.models import QuerySet, Sum, Q, Count, F
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import DetailView, UpdateView, RedirectView
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.conf import settings
from django.utils.dateformat import DateFormat
from datetime import datetime, timedelta, timezone as dt_timezone
from decimal import Decimal


from menu.models import Food, Category
from .models import User, Invoice, InvoiceItem, LoyaltyCard, LoyaltyTransaction, QRTableOffer, TimeBasedOffer, TodayDeal, HeldCart, OperatorPermission, PendingApproval, ReturnInvoice, ReturnInvoiceItem, AccountingEntry, WholesaleCustomer, WholesaleInvoice, WholesaleInvoiceItem, WholesaleAccountTransaction, WholesaleDeposit, WholesaleCreditSettlement, ExpenseCategory, BusinessExpense, SystemSetting, AuditLog, InventoryBatch, SaleItemCost
from .loyalty_utils import generate_qr_code_image, generate_loyalty_card_pdf, generate_loyalty_card_image
from .permissions import operator_permission_required, has_permission, has_admin_permission, admin_permission_required, software_owner_required, module_access_required, has_module_access
from .inventory_service import InventoryValuationService
from .format_utils import format_currency as _fmt_currency, get_currency_config
from cash_handling.services import CashDrawerService

from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.lib.utils import ImageReader

User = get_user_model()


def _cs():
    return get_currency_config()['symbol']


# =========================
# BASE
# =========================
class UserDetailView(LoginRequiredMixin, DetailView):
    model = User
    pk_url_kwarg = "pk"

user_detail_view = UserDetailView.as_view()


class UserUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = User
    fields = ["name"]
    success_message = _("Updated successfully")

    def get_success_url(self) -> str:
        return self.request.user.get_absolute_url()

    def get_object(self, queryset: QuerySet | None = None) -> User:
        return self.request.user

user_update_view = UserUpdateView.as_view()


class UserRedirectView(LoginRequiredMixin, RedirectView):
    def get_redirect_url(self) -> str:
        return reverse("users:detail", kwargs={"pk": self.request.user.pk})

user_redirect_view = UserRedirectView.as_view()


# =========================
# AUTH
# =========================
def register_view(request):
    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        password1 = request.POST.get("password1")
        password2 = request.POST.get("password2")

        if password1 != password2:
            messages.error(request, "Passwords do not match.")
            return redirect("users:login")
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists.")
            return redirect("users:login")

        User.objects.create_user(email=email, password=password1, name=name)
        messages.success(request, "Registration successful.")
        return redirect("users:login")

    return redirect("users:login")


def food_delivery_login(request):
    if request.user.is_authenticated:
        if request.user.is_software_owner:
            return redirect("users:system_settings")
        if request.user.is_staff:
            return redirect("users:admin_dashboard")
        elif getattr(request.user, "is_operator", False):
            return redirect("users:operator_dashboard")
        logout(request)
        return redirect("users:login")
    next_url = request.GET.get("next") or request.POST.get("next") or ""
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")
        user = authenticate(request, username=email, password=password)

        if user is not None:
            login(request, user)
            if next_url:
                return redirect(next_url)
            # Auto-create loyalty card for online registered customers only
            if not user.is_staff and not getattr(user, "is_operator", False) and not user.is_superuser:
                from .loyalty_utils import generate_qr_code_image, generate_loyalty_card_pdf, generate_loyalty_card_image
                card, created = LoyaltyCard.objects.get_or_create(
                    user=user,
                    defaults={'status': 'ACTIVE'}
                )
                if created or not card.qr_code_image or not card.card_pdf:
                    try:
                        generate_qr_code_image(card, request)
                        generate_loyalty_card_pdf(card, request)
                        generate_loyalty_card_image(card, request)
                    except Exception:
                        pass
                # First-time redirect to loyalty card page (using DB flag)
                if not card.first_card_popup_shown:
                    return redirect("users:loyalty_card_view")
            if user.is_software_owner:
                return redirect("users:system_settings")
            if user.is_staff:
                return redirect("users:admin_dashboard")
            elif getattr(user, "is_operator", False):
                return redirect("users:operator_dashboard")
            return redirect("/")
        else:
            messages.error(request, "Invalid credentials")
            return redirect("users:login")

    return render(request, "food-delivery/signin.html")


def logout_view(request):
    logout(request)
    return redirect("users:login")


# =========================
# HOME PAGE
# =========================
def home_view(request):
    return redirect("users:login")


# =========================
# QR MENU ACCESS
# =========================
@login_required(login_url='users:login')
@csrf_exempt
def qr_menu_view(request):
    if not request.session.get("customer_session_id"):
        request.session["customer_session_id"] = str(uuid.uuid4())

    products = Food.objects.filter(available=1)
    active_offer = _get_active_time_offer()
    active_deal = _get_active_deal()

    return render(request, "food-delivery/restaurant-detail.html", {
        "products": products,
        "is_qr_customer": True,
        "active_offer": active_offer,
        "active_deal": active_deal,
    })


@login_required(login_url='users:login')
def food_delivery_restaurant_detail(request):
    products = Food.objects.filter(available=1)
    active_offer = _get_active_time_offer()
    active_deal = _get_active_deal()
    return render(request, 'food-delivery/restaurant-detail.html', {
        "products": products,
        "active_offer": active_offer,
        "active_deal": active_deal,
    })


# =========================
# OFFER HELPERS
# =========================

def _get_active_time_offer():
    """Get active TimeBasedOffer by directly checking time range.
    Does NOT rely on the is_active field (avoids stale-data issues).
    Also updates is_active for future queries."""
    now = timezone.now()
    for offer in TimeBasedOffer.objects.all():
        try:
            start = timezone.make_aware(datetime.combine(offer.start_date, offer.start_time))
            end = timezone.make_aware(datetime.combine(offer.end_date, offer.end_time))
        except Exception:
            continue
        should_be_active = start <= now <= end
        if should_be_active != offer.is_active:
            offer.is_active = should_be_active
            offer.save(update_fields=["is_active"])
        if should_be_active:
            return offer
    return None


def _get_active_deal():
    """Get active TodayDeal by directly checking time range.
    Does NOT rely on the is_active field."""
    now = timezone.now()
    for deal in TodayDeal.objects.all():
        try:
            start = timezone.make_aware(datetime.combine(deal.start_date, deal.start_time))
            end = timezone.make_aware(datetime.combine(deal.end_date, deal.end_time))
        except Exception:
            continue
        should_be_active = start <= now <= end
        if should_be_active != deal.is_active:
            deal.is_active = should_be_active
            deal.save(update_fields=["is_active"])
        if should_be_active:
            return deal
    return None


# =========================
# CHECKOUT
# =========================
def _create_order_from_cart(cart, request, user=None, payment_method="card",
                            customer_timezone="UTC", use_loyalty_points=None,
                            secondary_payment_method=None):
    customer_session = request.session.get("customer_session_id") or str(uuid.uuid4())

    subtotal_amount = 0
    for item in cart:
        qty = int(item["qty"])
        price = float(item["price"])
        subtotal_amount += qty * price

    # ---------- Today's Deal ----------
    deal_obj = None
    deal_discount_amt = 0
    deal_id = request.session.pop("deal_checkout_id", None) if request else None
    if deal_id:
        try:
            deal_obj = TodayDeal.objects.get(id=deal_id)
            if not deal_obj.is_active:
                _get_active_deal()
            if not deal_obj.is_active:
                deal_obj = None
        except TodayDeal.DoesNotExist:
            pass
        request.session.pop("deal_checkout_cart", None)

    # ---------- Offer Discount (all payment methods, not just non-loyalty) ----------
    qr_offer_discount_pct = 0
    qr_offer_discount_amt = 0

    if not deal_obj:
        time_offer = _get_active_time_offer()
        if time_offer:
            qr_offer_discount_pct = float(time_offer.discount_percentage)
            qr_offer_discount_amt = round(subtotal_amount * qr_offer_discount_pct / 100, 2)
        else:
            if user is None:
                qr_offer = QRTableOffer.objects.first()
                if qr_offer:
                    qr_offer.check_and_update_status()
                    if qr_offer.is_active:
                        qr_offer_discount_pct = float(qr_offer.discount_percentage)
                        qr_offer_discount_amt = round(subtotal_amount * qr_offer_discount_pct / 100, 2)

    # ---------- Apply Deal Pricing ----------
    effective_subtotal = subtotal_amount
    if deal_obj:
        deal_products = list(deal_obj.products.all())
        original_total = sum(float(p.price) for p in deal_products)
        if deal_obj.free_product:
            original_total += float(deal_obj.free_product.price)

        if deal_obj.deal_type == 'combo_price' and deal_obj.combo_price:
            combo = float(deal_obj.combo_price)
            deal_discount_amt = original_total - combo
            if deal_discount_amt < 0:
                deal_discount_amt = 0
            effective_subtotal = combo
            qr_offer_discount_amt = deal_discount_amt
        elif deal_obj.deal_type == 'free_product' and deal_obj.free_product:
            free_price = float(deal_obj.free_product.price)
            deal_discount_amt = free_price
            effective_subtotal = original_total - free_price
            qr_offer_discount_amt = deal_discount_amt
        elif deal_obj.deal_type == 'percentage' and deal_obj.discount_percentage:
            pct = float(deal_obj.discount_percentage)
            deal_discount_amt = round(original_total * pct / 100, 2)
            effective_subtotal = original_total - deal_discount_amt
            qr_offer_discount_amt = deal_discount_amt
        else:
            effective_subtotal = original_total
        qr_offer_discount_pct = 0

    # ---------- Loyalty ----------
    is_loyalty = payment_method == "loyalty"
    loyalty_points_used = 0
    loyalty_card = None
    if deal_obj:
        remaining_amount = effective_subtotal
    else:
        remaining_amount = effective_subtotal - qr_offer_discount_amt
    if remaining_amount < 0:
        remaining_amount = 0
    if is_loyalty and user:
        try:
            loyalty_card = LoyaltyCard.objects.get(user=user, status='ACTIVE')
            available = int(loyalty_card.remaining_points)
            requested = int(use_loyalty_points) if use_loyalty_points else available
            loyalty_points_used = max(0, min(requested, available, int(remaining_amount)))
            remaining_amount = remaining_amount - loyalty_points_used
            if remaining_amount < 0:
                remaining_amount = 0
        except LoyaltyCard.DoesNotExist:
            is_loyalty = False
            payment_method = secondary_payment_method or "card"

    if is_loyalty and remaining_amount > 0 and secondary_payment_method:
        store_method = f"loyalty+{secondary_payment_method}"
    elif is_loyalty and remaining_amount <= 0:
        store_method = "loyalty"
    else:
        store_method = payment_method

    tax_pct = 0
    tax_amount = round(remaining_amount * tax_pct / 100, 2)
    grand_total = remaining_amount + tax_amount
    if grand_total < 0:
        grand_total = 0

    invoice = Invoice.objects.create(
        user=user,
        created_by=request.user if request and request.user.is_authenticated else None,
        invoice_number=f"INV-{uuid.uuid4().hex[:8].upper()}",
        payment_method=store_method,
        customer_timezone=customer_timezone,
        tax_percentage=tax_pct,
        tax_amount=tax_amount,
        subtotal_amount=effective_subtotal,
        total_amount=grand_total,
        customer_session_id=customer_session,
        is_loyalty_payment=is_loyalty,
        loyalty_points_used=loyalty_points_used,
        qr_offer_discount_percentage=qr_offer_discount_pct,
        qr_offer_discount_amount=qr_offer_discount_amt,
        deal=deal_obj,
        deal_discount_amount=deal_discount_amt,
    )

    stock_movements = []
    svc = InventoryValuationService()
    for item in cart:
        qty = int(item["qty"])
        price = float(item["price"])
        is_free = item.get("is_free", False)
        display_price = 0 if is_free else price
        subtotal = qty * display_price
        invoice_item = InvoiceItem.objects.create(
            invoice=invoice, product_name=item["name"],
            price=display_price, quantity=qty, subtotal=subtotal,
        )
        food_id = item.get("food_id") or item.get("id")
        if food_id:
            try:
                food = Food.objects.get(pk=food_id)
                if not is_free:
                    old_stock = food.stock
                    food.stock -= qty
                    if food.stock < 0:
                        food.stock = 0
                    food.save()
                    stock_movements.append({
                        "food": food,
                        "qty": qty,
                        "stock_before": old_stock,
                        "stock_after": food.stock,
                    })
                    total_cogs = Decimal('0')
                    for batch, take, unit_cost in svc.consume(food, qty):
                        SaleItemCost.objects.create(
                            inventory_batch=batch,
                            invoice_item=invoice_item,
                            quantity=take,
                            unit_cost=unit_cost,
                        )
                        total_cogs += Decimal(str(take)) * unit_cost
                    invoice_item.unit_cost_at_sale = float(total_cogs / Decimal(str(qty)))
                    invoice_item.save(update_fields=['unit_cost_at_sale'])
            except Food.DoesNotExist:
                pass

    for sm in stock_movements:
        _record_stock_movement(
            food=sm["food"],
            transaction_type='retail_sale',
            quantity_change=-sm["qty"],
            stock_before=sm["stock_before"],
            stock_after=sm["stock_after"],
            reference_number=invoice.invoice_number,
            created_by=invoice.user or None,
        )

    if loyalty_card and loyalty_points_used > 0:
        try:
            loyalty_card.redeem_points(loyalty_points_used, order_number=invoice.invoice_number)
        except ValueError:
            pass

    # Earn loyalty points for registered users
    if user:
        earn_card = LoyaltyCard.objects.filter(user=user, status='ACTIVE').first()
        if earn_card:
            pts_earned = _calculate_reward_points(cart)
            if pts_earned > 0:
                earn_card.add_points(pts_earned, order_number=invoice.invoice_number)
                invoice.loyalty_points_earned = pts_earned
                invoice.save(update_fields=['loyalty_points_earned'])

    invoice.generate_qr_code(request)
    invoice.save()

    return invoice


@csrf_exempt
def guest_checkout(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "POST required"}, status=400)

    try:
        data = json.loads(request.body)
        cart = data.get("cart", [])
        payment_method = data.get("payment_method", "card")
        customer_timezone = data.get("timezone", "UTC")
        use_loyalty_points = data.get("use_loyalty_points")
        secondary_payment_method = data.get("secondary_payment_method")
        if not cart:
            return JsonResponse({"success": False, "message": "Cart is empty"}, status=400)

        invoice = _create_order_from_cart(
            cart, request,
            payment_method=payment_method,
            customer_timezone=customer_timezone,
            use_loyalty_points=use_loyalty_points,
            secondary_payment_method=secondary_payment_method,
        )
        return JsonResponse({
            "success": True,
            "invoice_no": invoice.invoice_number,
            "uuid_token": invoice.uuid_token,
        })
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)


@login_required
@require_POST
def checkout_invoice(request):
    try:
        data = json.loads(request.body)
        cart = data.get("cart", [])
        payment_method = data.get("payment_method", "card")
        customer_timezone = data.get("timezone", "UTC")
        use_loyalty_points = data.get("use_loyalty_points")
        secondary_payment_method = data.get("secondary_payment_method")
        if not cart:
            return JsonResponse({"success": False, "message": "Cart is empty"}, status=400)

        invoice = _create_order_from_cart(
            cart, request,
            user=request.user,
            payment_method=payment_method,
            customer_timezone=customer_timezone,
            use_loyalty_points=use_loyalty_points,
            secondary_payment_method=secondary_payment_method,
        )
        return JsonResponse({
            "success": True,
            "invoice_no": invoice.invoice_number,
            "uuid_token": invoice.uuid_token,
        })
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)


# =========================
# SECURE INVOICE VIEW -> Redirect to PDF
# =========================
def secure_invoice_view(request, uuid_token):
    return redirect("users:invoice_pdf", uuid_token=uuid_token)


# =========================
# SET TIMEZONE
# =========================
@csrf_exempt
def set_timezone(request):
    if request.method == "POST":
        data = json.loads(request.body)
        tz = data.get("timezone", "")
        if tz:
            request.session["user_timezone"] = tz
            if request.user.is_authenticated:
                request.user.timezone = tz
                request.user.save(update_fields=["timezone"])
            return JsonResponse({"status": "ok"})
    return JsonResponse({"status": "error"}, status=400)


# =========================
# INVOICE DETAIL (legacy redirect)
# =========================
def invoice_detail(request, invoice_no):
    try:
        invoice = Invoice.objects.get(invoice_number=invoice_no)
        return redirect("users:invoice_pdf", uuid_token=invoice.uuid_token)
    except Invoice.DoesNotExist:
        return HttpResponse("Invoice not found", status=404)


# =========================
# INVOICE PDF
# =========================
def _calculate_reward_points(cart):
    """Calculate total reward points from cart items based on product reward_points.
    Never uses price/subtotal/total — only the reward_points field on Food."""
    total = 0
    for item in cart:
        qty = int(item.get("qty", 1)) if "qty" in item else int(item.get("quantity", 1))
        food_id = item.get("food_id") or item.get("id")
        if food_id:
            try:
                food = Food.objects.get(pk=food_id)
                total += food.reward_points * qty
            except Food.DoesNotExist:
                pass
    return total


def _generate_invoice_qr_image(invoice, request=None, prefix="invoice"):
    domain = settings.QR_CODE_BASE_URL
    if request:
        domain = f"{request.scheme}://{request.get_host()}"
    if prefix == "wholesale":
        secure_url = f"{domain}/users/wholesale/invoice/{invoice.uuid_token}/verify/"
    else:
        secure_url = f"{domain}/users/invoice/{invoice.uuid_token}/verify/"
    qr = qrcode.make(secure_url)
    buf = io.BytesIO()
    qr.save(buf, format="PNG")
    buf.seek(0)
    return ImageReader(buf)


def invoice_pdf(request, uuid_token):
    invoice = Invoice.objects.filter(uuid_token=uuid_token).first()

    if not invoice:
        return HttpResponse("Invoice not found", status=404)

    items = list(invoice.items.all())

    # ==========================
    # DYNAMIC THERMAL HEIGHT (60mm)
    # ==========================

    width = 60 * mm

    loyalty_height = 0
    if invoice.user:
        if LoyaltyCard.objects.filter(user=invoice.user).exists():
            loyalty_height = 45 * mm

    payment_height = 28 * mm

    header_height = 65 * mm
    item_per_height = 9 * mm
    summary_height = 50 * mm
    qr_height = 40 * mm
    footer_height = 15 * mm

    height = (
        header_height
        + (len(items) * item_per_height)
        + summary_height
        + payment_height
        + qr_height
        + footer_height
        + loyalty_height
    )

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(width, height))
    pdf.setTitle(invoice.invoice_number)

    w, h = width, height

    DARK = HexColor("#111827")
    GRAY = HexColor("#6b7280")
    GREEN = HexColor("#16a34a")

    MARGIN = 4 * mm
    right = w - MARGIN

    # ==========================
    # HEADER
    # ==========================
    y = h - 8 * mm

    settings_obj = SystemSetting.objects.filter(pk=1).first()
    company_name = settings_obj.company_name if settings_obj else 'POS'
    company_phone = settings_obj.company_phone if settings_obj else ''
    company_address = settings_obj.company_address if settings_obj else ''
    company_email = settings_obj.company_email if settings_obj else ''

    logo_y = y
    if settings_obj and settings_obj.company_logo:
        try:
            logo_path = settings_obj.company_logo.path
            logo = ImageReader(logo_path)
            logo_size = 10 * mm
            pdf.drawImage(logo, w / 2 - logo_size / 2, logo_y - logo_size, width=logo_size, height=logo_size, preserveAspectRatio=True)
            y = logo_y - logo_size - 3 * mm
        except Exception:
            pass

    pdf.setFillColor(DARK)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawCentredString(w / 2, y, company_name)

    y -= 5 * mm
    pdf.setFont("Helvetica", 7)
    if company_address:
        pdf.drawCentredString(w / 2, y, company_address)

    y -= 4 * mm
    if company_email:
        pdf.drawCentredString(w / 2, y, company_email)
        y -= 4 * mm
    pdf.drawCentredString(w / 2, y, company_phone)

    y -= 5 * mm
    pdf.setStrokeColor(DARK)
    pdf.line(MARGIN, y, right, y)

    # ==========================
    # CUSTOMER INFO
    # ==========================
    y -= 6 * mm
    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(DARK)

    if invoice.customer_name:
        customer_name = invoice.customer_name
        customer_phone = invoice.customer_phone or "--------"
    elif invoice.user:
        customer_name = invoice.user.name or invoice.user.email
        customer_phone = invoice.user.phone or "--------"
    else:
        customer_name = "Walk-in Customer"
        customer_phone = "--------"

    customer_lines = [
        ("Customer Name", customer_name),
        ("Customer Phone", customer_phone),
    ]
    for label, val in customer_lines:
        pdf.drawString(MARGIN, y, f"{label}")
        pdf.drawRightString(right, y, val)
        y -= 4 * mm

    pdf.line(MARGIN, y, right, y)
    y -= 5 * mm

    # ==========================
    # INVOICE INFO
    # ==========================
    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(DARK)

    info_lines = [
        ("Invoice #", invoice.invoice_number),
        ("Date", timezone.localtime(invoice.created_at).strftime("%d-%m-%Y %I:%M:%S %p")),
    ]
    for label, val in info_lines:
        pdf.drawString(MARGIN, y, f"{label}")
        pdf.drawRightString(right, y, val)
        y -= 4 * mm

    pdf.line(MARGIN, y, right, y)

    # ==========================
    # PAYMENT INFO
    # ==========================
    y -= 5 * mm
    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(DARK)
    pay_method = invoice.payment_method or "N/A"
    pay_lines = [
        ("Payment Method", pay_method.upper()),
    ]
    for label, val in pay_lines:
        pdf.drawString(MARGIN, y, f"{label}")
        pdf.drawRightString(right, y, val)
        y -= 4 * mm

    pdf.line(MARGIN, y, right, y)

    # ==========================
    # ITEMS
    # ==========================
    y -= 5 * mm
    subtotal = 0

    pdf.setFont("Helvetica-Bold", 7)
    pdf.drawString(MARGIN, y, "ITEM")
    pdf.drawRightString(right, y, "TOTAL")
    y -= 4 * mm

    pdf.setFont("Helvetica", 7)
    for item in items:
        item_tax = float(item.tax_amount) if item.tax_amount else 0
        item_tax_pct = float(item.tax_percentage) if item.tax_percentage else 0
        subtotal += float(item.subtotal)
        name = item.product_name[:22]
        pdf.drawString(MARGIN, y, name)
        pdf.drawRightString(right, y, f"{_cs()}{float(item.subtotal):.0f}")
        y -= 4 * mm
        pdf.setFont("Helvetica", 6)
        pdf.setFillColor(GRAY)
        qty_line = f"{item.quantity} x {_cs()}{float(item.price):.0f}"
        if item_tax > 0:
            qty_line += f"  |  Tax ({item_tax_pct:.0f}%): {_cs()}{item_tax:.0f}"
        pdf.drawString(MARGIN, y, qty_line)
        pdf.setFillColor(DARK)
        pdf.setFont("Helvetica", 7)
        y -= 5 * mm

    pdf.line(MARGIN, y, right, y)

    # ==========================
    # SUMMARY
    # ==========================
    y -= 6 * mm
    tax_amount = float(invoice.tax_amount) if invoice.tax_amount else 0
    loyalty_used = float(invoice.loyalty_points_used) if invoice.loyalty_points_used else 0
    stored_total = float(invoice.total_amount)
    original_grand_total = stored_total + loyalty_used
    sub_amt = float(invoice.subtotal_amount) if invoice.subtotal_amount else subtotal
    qr_disc_pct = float(invoice.qr_offer_discount_percentage) if invoice.qr_offer_discount_percentage else 0
    qr_disc_amt = float(invoice.qr_offer_discount_amount) if invoice.qr_offer_discount_amount else 0
    deal_disc_amt = float(invoice.deal_discount_amount) if invoice.deal_discount_amount else 0

    total_discount = qr_disc_amt + deal_disc_amt

    if invoice.deal:
        pdf.setFont("Helvetica", 7)
        pdf.setFillColor(HexColor("#8b5cf6"))
        pdf.drawString(MARGIN, y, invoice.deal.title[:25])
        y -= 4 * mm
        pdf.setFillColor(DARK)

    pdf.setFont("Helvetica", 8)
    pdf.drawString(MARGIN, y, "Subtotal")
    pdf.drawRightString(right, y, f"{_cs()}{sub_amt:.0f}")
    y -= 5 * mm

    if total_discount > 0:
        pdf.setFont("Helvetica", 7)
        pdf.setFillColor(HexColor("#f59e0b"))
        pdf.drawString(MARGIN, y, "Discount")
        pdf.drawRightString(right, y, f"{_cs()}{total_discount:.0f}")
        y -= 5 * mm

    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(GRAY)
    pdf.drawString(MARGIN, y, "Total Tax")
    pdf.drawRightString(right, y, f"{_cs()}{tax_amount:.0f}")
    y -= 5 * mm
    pdf.setFillColor(DARK)

    pdf.line(MARGIN, y, right, y)
    y -= 6 * mm

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(MARGIN, y, "GRAND TOTAL")
    pdf.drawRightString(right, y, f"{_cs()}{original_grand_total:.0f}")

    y -= 8 * mm

    # ==========================
    # PAYMENT SECTION
    # ==========================
    y -= 4 * mm
    pdf.setFont("Helvetica-Bold", 8)
    pdf.setFillColor(HexColor("#16a34a"))
    pdf.drawString(MARGIN, y, "PAYMENT")
    y -= 4 * mm

    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(DARK)

    cash_val = float(invoice.cash_received) if invoice.cash_received else stored_total
    change_val = float(invoice.change_due) if invoice.change_due else 0

    pay_lines = [
        ("Total", f"{_cs()}{original_grand_total:.0f}"),
    ]
    if loyalty_used > 0:
        pay_lines.append(("Loyalty", f"-{_cs()}{loyalty_used:.0f}"))
    pay_lines.append(("Cash", f"{_cs()}{cash_val:.0f}"))
    pay_lines.append(("Change Due", f"{_cs()}{change_val:.0f}"))

    for label, val in pay_lines:
        pdf.drawString(MARGIN, y, label)
        pdf.drawRightString(right, y, val)
        y -= 4 * mm

    pdf.line(MARGIN, y, right, y)
    y -= 4 * mm

    # ==========================
    # LOYALTY CARD INFO (Order Total → Loyalty Used → Paid By Cash → Remaining Balance)
    # ==========================
    # LOYALTY INFO (Customer Name, Points Earned, Points Redeemed, Remaining Balance)
    # ==========================
    if invoice.user:
        lcard = LoyaltyCard.objects.filter(user=invoice.user).first()
        if lcard:
            y -= 5 * mm
            pdf.setFont("Helvetica-Bold", 7)
            pdf.setFillColor(HexColor("#f59e0b"))
            pdf.drawCentredString(w / 2, y, "LOYALTY")
            pdf.setFillColor(DARK)
            y -= 4 * mm
            pdf.setFont("Helvetica", 6)

            display_lines = [
                ("Customer Name", invoice.user.name or invoice.user.email),
            ]
            if invoice.loyalty_points_earned and invoice.loyalty_points_earned > 0:
                display_lines.append(("Points Earned", f"+{invoice.loyalty_points_earned} pts"))
            if invoice.loyalty_points_used and invoice.loyalty_points_used > 0:
                display_lines.append(("Points Redeemed", f"-{invoice.loyalty_points_used} pts"))
            display_lines.append(("Remaining Balance", f"{lcard.remaining_points} pts"))

            for label, val in display_lines:
                pdf.drawString(MARGIN, y, label)
                pdf.drawRightString(right, y, val)
                y -= 3.5 * mm

            y -= 2 * mm
            pdf.line(MARGIN, y, right, y)
            y -= 4 * mm

    # ==========================
    # QR CODE (Bottom Center)
    # ==========================
    y -= 10 * mm

    qr = _generate_invoice_qr_image(invoice, request)
    qr_size = 25 * mm
    qr_x = (w - qr_size) / 2
    pdf.drawImage(qr, qr_x, y - qr_size, width=qr_size, height=qr_size)

    y -= qr_size + 4 * mm

    pdf.setFont("Helvetica", 6)
    pdf.setFillColor(GRAY)
    pdf.drawCentredString(w / 2, y, "Scan To Verify Invoice")
    pdf.setFillColor(DARK)

    # ==========================
    # FOOTER
    # ==========================
    y -= 6 * mm
    pdf.setFont("Helvetica-Bold", 7)
    pdf.setFillColor(DARK)
    pdf.drawCentredString(w / 2, y, "Thank you for purchasing with us")

    pdf.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{invoice.invoice_number}.pdf"'
    return response

# =========================
# INVOICE VERIFY (via QR)
# =========================
def invoice_verify(request, uuid_token):
    invoice = Invoice.objects.filter(uuid_token=uuid_token).first()

    if not invoice:
        return render(request, "food-delivery/invoice_verify.html", {
            "valid": False,
            "error": "Invoice not found. The link may be invalid or the invoice has been removed.",
        })

    items = invoice.items.all()
    subtotal = float(invoice.subtotal_amount) if invoice.subtotal_amount else sum(float(item.subtotal) for item in items)
    tax_pct = float(invoice.tax_percentage) if invoice.tax_percentage else 0
    tax = float(invoice.tax_amount) if invoice.tax_amount else 0

    customer_name = invoice.customer_name or (invoice.user.name if invoice.user else "Walk-in Customer")
    customer_phone = invoice.customer_phone or (invoice.user.phone if invoice.user else "N/A")

    invoice_date = timezone.localtime(invoice.created_at).strftime("%d-%m-%Y %I:%M:%S %p")

    loyalty_used = invoice.loyalty_points_used or 0
    original_grand_total = float(invoice.total_amount) + loyalty_used

    context = {
        "valid": True,
        "invoice": invoice,
        "items": items,
        "customer_name": customer_name,
        "customer_phone": customer_phone,
        "subtotal": subtotal,
        "tax": tax,
        "tax_pct": tax_pct,
        "payment_method": invoice.payment_method or "N/A",
        "invoice_date": invoice_date,
        "original_grand_total": original_grand_total,
        "loyalty_used": loyalty_used,
    }
    return render(request, "food-delivery/invoice_verify.html", context)


# =========================
# ADMIN POS (Point of Sale for Admin)
# =========================
@login_required
@module_access_required('pos')
def admin_pos(request):
    if not request.user.is_staff:
        return redirect('users:admin_dashboard')
    categories = Category.objects.filter(is_active=True)
    can_create_invoice = True
    can_returns = has_permission(request.user, 'can_returns')
    return render(request, "operator/dashboard.html", {
        "categories": categories,
        "can_create_invoice": can_create_invoice,
        "can_returns": can_returns,
    })

# =========================
# ADMIN DASHBOARD
# =========================
@staff_member_required
def admin_dashboard(request):
    foods_count = Food.objects.count()
    invoices_count = Invoice.objects.count()
    users_count = User.objects.count()
    operator_users_count = User.objects.filter(is_operator=True).count()
    customers_count = User.objects.filter(is_staff=False, is_operator=False, is_superuser=False).count()

    # Low stock alerts
    low_stock_products = Food.objects.filter(stock__gt=0, stock__lte=5)[:20]
    out_of_stock_products = Food.objects.filter(stock=0)[:20]

    # Offer & Deal stats
    from django.db.models import Sum
    total_offers = TimeBasedOffer.objects.count()
    active_offers = TimeBasedOffer.objects.filter(is_active=True).count()
    expired_offers = TimeBasedOffer.objects.filter(is_active=False).count()
    total_deals = TodayDeal.objects.count()
    active_deals = TodayDeal.objects.filter(is_active=True).count()
    offer_usage = Invoice.objects.aggregate(total=Sum('qr_offer_discount_amount'))['total'] or 0
    offer_discount_given = Invoice.objects.filter(qr_offer_discount_amount__gt=0).count()

    revenue = Invoice.objects.aggregate(total=Sum("total_amount"))["total"] or 0
    wholesale_revenue = WholesaleInvoice.objects.aggregate(total=Sum("total_amount"))["total"] or 0
    revenue = float(revenue) + float(wholesale_revenue)
    returns_total = ReturnInvoice.objects.aggregate(total=Sum("total_refund_amount"))["total"] or 0
    net_revenue = float(revenue) - float(returns_total)

    card_tax = Invoice.objects.filter(payment_method__iexact="card").aggregate(total=Sum("tax_amount"))["total"] or 0
    cash_tax = Invoice.objects.filter(payment_method__iexact="cash").aggregate(total=Sum("tax_amount"))["total"] or 0
    wholesale_tax = WholesaleInvoice.objects.aggregate(total=Sum("tax_amount"))["total"] or 0
    total_tax = round(float(card_tax) + float(cash_tax) + float(wholesale_tax), 2)

    pending_approvals_count = PendingApproval.objects.filter(is_approved=False, is_rejected=False).count()
    pending_approvals = PendingApproval.objects.filter(
        is_approved=False, is_rejected=False
    ).select_related('operator').order_by('-created_at')[:10]

    # Backup metadata
    from megaone.users.backup_utils import get_backup_metadata, get_backup_file_size, BACKUP_FILE
    backup_meta = get_backup_metadata()
    backup_file_size = get_backup_file_size()
    backup_exists = BACKUP_FILE.exists()
    last_backup_display = backup_meta.get("last_backup_time") or "Never"
    backup_status = backup_meta.get("last_backup_status") or "none"
    auto_backup_enabled = backup_meta.get("auto_backup_enabled", True)

    # Dashboard-specific calculations
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_sales = Invoice.objects.filter(created_at__gte=today_start).aggregate(total=Sum("total_amount"))["total"] or 0
    today_returns = ReturnInvoice.objects.filter(created_at__gte=today_start).aggregate(total=Sum("total_refund_amount"))["total"] or 0
    today_net_sales = float(today_sales) - float(today_returns)

    month_start = today_start.replace(day=1)
    monthly_sales = Invoice.objects.filter(created_at__gte=month_start).aggregate(total=Sum("total_amount"))["total"] or 0
    monthly_returns = ReturnInvoice.objects.filter(created_at__gte=month_start).aggregate(total=Sum("total_refund_amount"))["total"] or 0
    monthly_net_sales = float(monthly_sales) - float(monthly_returns)

    year_start = today_start.replace(month=1, day=1)
    yearly_sales = Invoice.objects.filter(created_at__gte=year_start).aggregate(total=Sum("total_amount"))["total"] or 0
    yearly_returns = ReturnInvoice.objects.filter(created_at__gte=year_start).aggregate(total=Sum("total_refund_amount"))["total"] or 0
    yearly_net_sales = float(yearly_sales) - float(yearly_returns)

    returned_qty = ReturnInvoiceItem.objects.aggregate(total=Sum("quantity"))["total"] or 0

    return render(request, "admin/dashboard.html", {
        "foods_count": foods_count,
        "invoices_count": invoices_count,
        "users_count": users_count,
        "operator_users_count": operator_users_count,
        "customers_count": customers_count,
        "revenue": net_revenue,
        "gross_revenue": float(revenue),
        "returns_total": float(returns_total),
        "total_tax": total_tax,
        "card_tax": card_tax,
        "cash_tax": cash_tax,

        "total_orders": invoices_count,
        "low_stock_products": low_stock_products,
        "out_of_stock_products": out_of_stock_products,
        "total_offers": total_offers,
        "active_offers": active_offers,
        "expired_offers": expired_offers,
        "total_deals": total_deals,
        "active_deals": active_deals,
        "offer_usage": offer_usage,
        "offer_discount_given": offer_discount_given,
        "pending_approvals_count": pending_approvals_count,
        "pending_approvals": pending_approvals,
        "last_backup_time": last_backup_display,
        "backup_status": backup_status,
        "auto_backup_enabled": auto_backup_enabled,
        "backup_exists": backup_exists,
        "backup_file_size": backup_file_size,
        "today_sales": float(today_sales),
        "today_returns": float(today_returns),
        "today_net_sales": today_net_sales,
        "monthly_sales": float(monthly_sales),
        "monthly_returns": float(monthly_returns),
        "monthly_net_sales": monthly_net_sales,
        "yearly_sales": float(yearly_sales),
        "yearly_returns": float(yearly_returns),
        "yearly_net_sales": yearly_net_sales,
        "returned_qty": returned_qty,
    })


@login_required
@module_access_required('reports')
def revenue_filter(request):
    if not has_permission(request.user, 'can_view_reports'):
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        start_str = request.GET.get("start_date")
        end_str = request.GET.get("end_date")
        invoices = Invoice.objects.all()
        if start_str and end_str:
            try:
                start_dt = datetime.strptime(start_str, "%Y-%m-%d")
                end_dt = datetime.strptime(end_str, "%Y-%m-%d") + timedelta(days=1)
                start_utc = timezone.make_aware(start_dt, dt_timezone.utc)
                end_utc = timezone.make_aware(end_dt, dt_timezone.utc)
                invoices = invoices.filter(created_at__range=[start_utc, end_utc])
            except (ValueError, TypeError):
                return JsonResponse({"success": False, "error": "Invalid date format. Use YYYY-MM-DD."}, status=400)
        revenue = invoices.aggregate(total=Sum("total_amount"))["total"] or 0

        returns_qs = ReturnInvoice.objects.all()
        if start_str and end_str:
            returns_qs = returns_qs.filter(created_at__range=[start_utc, end_utc])
        returns_total = returns_qs.aggregate(total=Sum("total_refund_amount"))["total"] or 0
        net_revenue = float(revenue) - float(returns_total)

        return JsonResponse({
            "success": True,
            "revenue": float(revenue),
            "returns_total": float(returns_total),
            "net_revenue": net_revenue,
        })
    except Exception as e:
        traceback.print_exc()
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
@module_access_required('reports')
def tax_analytics(request):
    if not has_permission(request.user, 'can_view_reports'):
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        start_str = request.GET.get("start_date")
        end_str = request.GET.get("end_date")
        invoices = Invoice.objects.all()
        if start_str and end_str:
            try:
                start_dt = datetime.strptime(start_str, "%Y-%m-%d")
                end_dt = datetime.strptime(end_str, "%Y-%m-%d") + timedelta(days=1)
                start_utc = timezone.make_aware(start_dt, dt_timezone.utc)
                end_utc = timezone.make_aware(end_dt, dt_timezone.utc)
                invoices = invoices.filter(created_at__range=[start_utc, end_utc])
            except (ValueError, TypeError):
                return JsonResponse({"success": False, "error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

        revenue = invoices.aggregate(total=Sum("total_amount"))["total"] or 0

        returns_qs = ReturnInvoice.objects.all()
        if start_str and end_str:
            returns_qs = returns_qs.filter(created_at__range=[start_utc, end_utc])
        returns_total = returns_qs.aggregate(total=Sum("total_refund_amount"))["total"] or 0
        net_revenue = float(revenue) - float(returns_total)

        card_tax = invoices.filter(payment_method__iexact="card").aggregate(total=Sum("tax_amount"))["total"] or 0
        cash_tax = invoices.filter(payment_method__iexact="cash").aggregate(total=Sum("tax_amount"))["total"] or 0
        total_tax = round(float(card_tax) + float(cash_tax), 2)

        return JsonResponse({
            "success": True,
            "revenue": float(revenue),
            "returns_total": float(returns_total),
            "net_revenue": net_revenue,
            "total_tax": total_tax,
            "card_tax": float(card_tax),
            "cash_tax": float(cash_tax),
        })
    except Exception as e:
        traceback.print_exc()
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
@module_access_required('profit_loss')
def profit_loss_data(request):
    if not has_permission(request.user, 'can_view_reports'):
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        start_str = request.GET.get("start_date")
        end_str = request.GET.get("end_date")
        start_dt = end_dt = None
        if start_str and end_str:
            start_dt = timezone.make_aware(datetime.strptime(start_str, "%Y-%m-%d"), dt_timezone.utc)
            end_dt = timezone.make_aware(datetime.strptime(end_str, "%Y-%m-%d") + timedelta(days=1), dt_timezone.utc)

        # Retail invoices
        retail_qs = Invoice.objects.all()
        wholesale_qs = WholesaleInvoice.objects.all()
        returns_qs = ReturnInvoice.objects.all()
        if start_dt and end_dt:
            retail_qs = retail_qs.filter(created_at__range=[start_dt, end_dt])
            wholesale_qs = wholesale_qs.filter(created_at__range=[start_dt, end_dt])
            returns_qs = returns_qs.filter(created_at__range=[start_dt, end_dt])

        retail_count = retail_qs.count()
        retail_revenue = float(retail_qs.aggregate(total=Sum("subtotal_amount"))["total"] or 0)
        retail_tax = float(retail_qs.aggregate(total=Sum("tax_amount"))["total"] or 0)
        retail_discounts = float(
            retail_qs.aggregate(
                total=Sum(F("qr_offer_discount_amount") + F("deal_discount_amount"))
            )["total"] or 0
        )

        wholesale_count = wholesale_qs.count()
        wholesale_revenue = float(wholesale_qs.aggregate(total=Sum("subtotal_amount"))["total"] or 0)
        wholesale_tax = float(wholesale_qs.aggregate(total=Sum("tax_amount"))["total"] or 0)
        wholesale_discounts = float(wholesale_qs.aggregate(total=Sum("discount_amount"))["total"] or 0)

        pl_result = _compute_pl_data_core(
            retail_qs, wholesale_qs, returns_qs,
            retail_revenue, retail_tax, retail_discounts,
            wholesale_revenue, wholesale_tax, wholesale_discounts,
            start_dt, end_dt,
        )
        return_revenue = pl_result['return_revenue']
        business_expenses = pl_result['business_expenses']
        total_tax = pl_result['total_tax']
        total_discounts = pl_result['total_discounts']
        retail_cogs_net = pl_result['retail_cogs_net']
        wholesale_cogs_net = pl_result['wholesale_cogs_net']
        total_cogs_net = pl_result['total_cogs_net']
        total_revenue_net = pl_result['total_revenue_net']
        gross_profit = pl_result['gross_profit']
        net_profit = pl_result['net_profit']

        # Inventory Summary (using batch-based valuation)
        inv_svc = InventoryValuationService()
        inv_summary = inv_svc.get_inventory_summary()
        total_products = inv_summary['total_products']
        total_stock_quantity = inv_summary['total_stock_quantity']
        total_inventory_cost = float(inv_summary['total_inventory_value'])

        return JsonResponse({
            "success": True,
            "retail_revenue": round(retail_revenue, 2),
            "retail_tax": round(retail_tax, 2),
            "retail_count": retail_count,
            "retail_discounts": round(retail_discounts, 2),
            "retail_cogs": round(retail_cogs_net, 2),
            "wholesale_revenue": round(wholesale_revenue, 2),
            "wholesale_tax": round(wholesale_tax, 2),
            "wholesale_count": wholesale_count,
            "wholesale_discounts": round(wholesale_discounts, 2),
            "wholesale_cogs": round(wholesale_cogs_net, 2),
            "total_revenue": round(total_revenue_net, 2),
            "total_tax": round(total_tax, 2),
            "total_count": retail_count + wholesale_count,
            "total_discounts": round(total_discounts, 2),
            "return_revenue": round(return_revenue, 2),
            "business_expenses": round(business_expenses, 2),
            "total_cogs": round(total_cogs_net, 2),
            "gross_profit": round(gross_profit, 2),
            "net_profit": round(net_profit, 2),
            "total_products": total_products,
            "total_stock_quantity": total_stock_quantity,
            "total_inventory_cost": round(total_inventory_cost, 2),
        })
    except Exception as e:
        traceback.print_exc()
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
@module_access_required('profit_loss')
def profit_loss_statement(request):
    if not has_permission(request.user, 'can_view_reports'):
        messages.error(request, "Access denied")
        return redirect("users:admin_dashboard")
    return render(request, "admin/profit_loss_statement.html", {"active_page": "profit_loss"})


@login_required
@module_access_required('profit_loss')
def profit_loss_export_pdf(request):
    if not has_permission(request.user, 'can_view_reports'):
        return HttpResponse("Access denied", status=403)
    try:
        start_str = request.GET.get("start_date")
        end_str = request.GET.get("end_date")
        if not start_str or not end_str:
            return HttpResponse("Missing date parameters", status=400)

        start_dt = timezone.make_aware(datetime.strptime(start_str, "%Y-%m-%d"), dt_timezone.utc)
        end_dt = timezone.make_aware(datetime.strptime(end_str, "%Y-%m-%d") + timedelta(days=1), dt_timezone.utc)

        retail_qs = Invoice.objects.filter(created_at__range=[start_dt, end_dt])
        wholesale_qs = WholesaleInvoice.objects.filter(created_at__range=[start_dt, end_dt])
        returns_qs = ReturnInvoice.objects.filter(created_at__range=[start_dt, end_dt])

        retail_revenue = float(retail_qs.aggregate(total=Sum("subtotal_amount"))["total"] or 0)
        retail_tax = float(retail_qs.aggregate(total=Sum("tax_amount"))["total"] or 0)
        retail_discounts = float(retail_qs.aggregate(total=Sum(F("qr_offer_discount_amount") + F("deal_discount_amount")))["total"] or 0)
        wholesale_revenue = float(wholesale_qs.aggregate(total=Sum("subtotal_amount"))["total"] or 0)
        wholesale_tax = float(wholesale_qs.aggregate(total=Sum("tax_amount"))["total"] or 0)
        wholesale_discounts = float(wholesale_qs.aggregate(total=Sum("discount_amount"))["total"] or 0)
        pl_result = _compute_pl_data_core(
            retail_qs, wholesale_qs, returns_qs,
            retail_revenue, retail_tax, retail_discounts,
            wholesale_revenue, wholesale_tax, wholesale_discounts,
            start_dt, end_dt,
        )
        return_revenue = pl_result['return_revenue']
        business_expenses = pl_result['business_expenses']
        total_tax = pl_result['total_tax']
        total_discounts = pl_result['total_discounts']
        retail_cogs_net = pl_result['retail_cogs_net']
        wholesale_cogs_net = pl_result['wholesale_cogs_net']
        total_cogs_net = pl_result['total_cogs_net']
        total_revenue_net = pl_result['total_revenue_net']
        gross_profit = pl_result['gross_profit']
        net_profit = pl_result['net_profit']

        # Inventory Summary (using batch-based valuation)
        inv_svc = InventoryValuationService()
        inv_summary = inv_svc.get_inventory_summary()
        total_products = inv_summary['total_products']
        total_stock_quantity = inv_summary['total_stock_quantity']
        total_inventory_cost = float(inv_summary['total_inventory_value'])

        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=18, spaceAfter=6, textColor=colors.HexColor("#1e293b"))
        subtitle_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#64748b"), spaceAfter=20)

        elements.append(Paragraph("Profit & Loss Statement", title_style))
        elements.append(Paragraph(f"Period: {start_str} to {end_str}", subtitle_style))
        elements.append(Spacer(1, 10))

        def fmt(v):
            try:
                return f"{_cs()}{float(v or 0):,.2f}"
            except (ValueError, TypeError):
                return f"{_cs()}0.00"

        def section_row(label, value, bold=False, color=None):
            return [Paragraph(f"<b>{label}</b>" if bold else label, styles["Normal"]),
                    Paragraph(f"<b>{fmt(value)}</b>" if bold else fmt(value), styles["Normal"])]

        data = [
            section_row("REVENUE", "", bold=True),
            section_row("Retail Sales Revenue", retail_revenue),
            section_row("Wholesale Sales Revenue", wholesale_revenue),
            [Paragraph("Less: Returns", styles["Normal"]),
             Paragraph(f"({fmt(return_revenue)})", styles["Normal"])],
            section_row("Total Revenue", total_revenue_net, bold=True),
            ["", ""],
            section_row("TAXES COLLECTED", "", bold=True),
            section_row("Retail Tax", retail_tax),
            section_row("Wholesale Tax", wholesale_tax),
            section_row("Total Tax Collected", total_tax, bold=True),
            ["", ""],
            section_row("COSTS & EXPENSES", "", bold=True),
            section_row("Product Cost (COGS)", total_cogs_net),
            section_row("Discounts Given", total_discounts),
            section_row("Business Expenses", business_expenses),
            ["", ""],
            section_row("INVENTORY SUMMARY", "", bold=True),
            [Paragraph("Total Products", styles["Normal"]),
             Paragraph(str(total_products), styles["Normal"])],
            [Paragraph("Total Stock Quantity", styles["Normal"]),
             Paragraph(f"{total_stock_quantity:,} Units", styles["Normal"])],
            section_row("Total Inventory Cost", total_inventory_cost),
            ["", ""],
            section_row("Gross Profit", gross_profit, bold=True, color="green" if gross_profit >= 0 else "red"),
            section_row("Net Profit", net_profit, bold=True, color="green" if net_profit >= 0 else "red"),
        ]

        table = Table(data, colWidths=[350, 160])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
            ("BACKGROUND", (0, 6), (-1, 6), colors.HexColor("#f1f5f9")),
            ("BACKGROUND", (0, 10), (-1, 10), colors.HexColor("#f1f5f9")),
            ("BACKGROUND", (0, 16), (-1, 16), colors.HexColor("#f1f5f9")),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#e2e8f0")),
            ("LINEBELOW", (0, 4), (-1, 4), 0.5, colors.HexColor("#94a3b8")),
            ("LINEBELOW", (0, 9), (-1, 9), 0.5, colors.HexColor("#94a3b8")),
            ("LINEBELOW", (0, 13), (-1, 13), 0.5, colors.HexColor("#94a3b8")),
            ("LINEABOVE", (0, 15), (-1, -1), 1, colors.HexColor("#1e293b")),
        ]))
        elements.append(table)

        doc.build(elements)
        buf.seek(0)
        return FileResponse(buf, as_attachment=True, filename=f"profit_loss_{start_str}_{end_str}.pdf")
    except Exception as e:
        traceback.print_exc()
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


@login_required
@module_access_required('profit_loss')
def profit_loss_export_excel(request):
    if not has_permission(request.user, 'can_view_reports'):
        return HttpResponse("Access denied", status=403)
    try:
        start_str = request.GET.get("start_date")
        end_str = request.GET.get("end_date")
        if not start_str or not end_str:
            return HttpResponse("Missing date parameters", status=400)

        start_dt = timezone.make_aware(datetime.strptime(start_str, "%Y-%m-%d"), dt_timezone.utc)
        end_dt = timezone.make_aware(datetime.strptime(end_str, "%Y-%m-%d") + timedelta(days=1), dt_timezone.utc)

        retail_qs = Invoice.objects.filter(created_at__range=[start_dt, end_dt])
        wholesale_qs = WholesaleInvoice.objects.filter(created_at__range=[start_dt, end_dt])
        returns_qs = ReturnInvoice.objects.filter(created_at__range=[start_dt, end_dt])

        retail_revenue = float(retail_qs.aggregate(total=Sum("subtotal_amount"))["total"] or 0)
        retail_tax = float(retail_qs.aggregate(total=Sum("tax_amount"))["total"] or 0)
        retail_discounts = float(retail_qs.aggregate(total=Sum(F("qr_offer_discount_amount") + F("deal_discount_amount")))["total"] or 0)
        wholesale_revenue = float(wholesale_qs.aggregate(total=Sum("subtotal_amount"))["total"] or 0)
        wholesale_tax = float(wholesale_qs.aggregate(total=Sum("tax_amount"))["total"] or 0)
        wholesale_discounts = float(wholesale_qs.aggregate(total=Sum("discount_amount"))["total"] or 0)
        pl_result = _compute_pl_data_core(
            retail_qs, wholesale_qs, returns_qs,
            retail_revenue, retail_tax, retail_discounts,
            wholesale_revenue, wholesale_tax, wholesale_discounts,
            start_dt, end_dt,
        )
        return_revenue = pl_result['return_revenue']
        business_expenses = pl_result['business_expenses']
        total_tax = pl_result['total_tax']
        total_discounts = pl_result['total_discounts']
        retail_cogs_net = pl_result['retail_cogs_net']
        wholesale_cogs_net = pl_result['wholesale_cogs_net']
        total_cogs_net = pl_result['total_cogs_net']
        total_revenue_net = pl_result['total_revenue_net']
        gross_profit = pl_result['gross_profit']
        net_profit = pl_result['net_profit']

        # Inventory Summary (using batch-based valuation)
        inv_svc = InventoryValuationService()
        inv_summary = inv_svc.get_inventory_summary()
        total_products = inv_summary['total_products']
        total_stock_quantity = inv_summary['total_stock_quantity']
        total_inventory_cost = float(inv_summary['total_inventory_value'])

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Profit & Loss"

        header_font = Font(bold=True, size=14, color="1e293b")
        sub_header_font = Font(bold=True, size=11, color="475569")
        section_font = Font(bold=True, size=11, color="1e293b")
        value_font = Font(size=11)
        profit_font = Font(bold=True, size=12, color="22c55e")
        loss_font = Font(bold=True, size=12, color="ef4444")
        header_fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
        thin_border = Border(
            bottom=Side(style="thin", color="e2e8f0"),
        )

        ws.merge_cells("A1:B1")
        ws["A1"] = "Profit & Loss Statement"
        ws["A1"].font = header_font
        ws["A2"] = f"Period: {start_str} to {end_str}"
        ws["A2"].font = Font(size=10, color="64748b")

        row = 4
        def write_section(label, value="", bold=False):
            nonlocal row
            ws.cell(row=row, column=1, value=label).font = section_font if bold else value_font
            if value:
                ws.cell(row=row, column=2, value=value).font = section_font if bold else value_font
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
            if bold:
                ws.cell(row=row, column=1).fill = header_fill
            row += 1

        write_section("REVENUE", bold=True)
        write_section("Retail Sales Revenue", round(retail_revenue, 2))
        write_section("Wholesale Sales Revenue", round(wholesale_revenue, 2))
        write_section("Less: Returns", round(return_revenue, 2))
        write_section("Total Revenue", round(total_revenue_net, 2), bold=True)
        row += 1

        write_section("TAXES COLLECTED", bold=True)
        write_section("Retail Tax", round(retail_tax, 2))
        write_section("Wholesale Tax", round(wholesale_tax, 2))
        write_section("Total Tax Collected", round(total_tax, 2), bold=True)
        row += 1

        write_section("COSTS & EXPENSES", bold=True)
        write_section("Product Cost (COGS)", round(total_cogs_net, 2))
        write_section("Discounts Given", round(total_discounts, 2))
        write_section("Business Expenses", round(business_expenses, 2))
        row += 1

        write_section("INVENTORY SUMMARY", bold=True)
        write_section("Total Products", total_products)
        write_section("Total Stock Quantity", f"{total_stock_quantity} Units")
        write_section("Total Inventory Cost", round(total_inventory_cost, 2))
        row += 1

        ws.cell(row=row, column=1, value="Gross Profit").font = section_font
        ws.cell(row=row, column=2, value=round(gross_profit, 2)).font = profit_font if gross_profit >= 0 else loss_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        row += 1
        ws.cell(row=row, column=1, value="Net Profit").font = Font(bold=True, size=13)
        ws.cell(row=row, column=2, value=round(net_profit, 2)).font = profit_font if net_profit >= 0 else loss_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return FileResponse(buf, as_attachment=True, filename=f"profit_loss_{start_str}_{end_str}.xlsx")
    except Exception as e:
        traceback.print_exc()
        return HttpResponse(f"Error generating Excel: {str(e)}", status=500)


# =========================
# PRODUCT MANAGEMENT
# =========================
@staff_member_required
@module_access_required('products')
def product_list(request):
    products = Food.objects.all()
    return render(request, "admin/products.html", {
        "products": products,
        "today": timezone.now(),
        "active_page": "products",
    })


@staff_member_required
@module_access_required('products')
def add_product(request):
    categories = Category.objects.all()
    if request.method == "POST":
        category = Category.objects.get(id=request.POST.get("category"))
        ws_price = request.POST.get("wholesale_price")
        stock_qty = int(request.POST.get("stock", 0))
        cost_price = request.POST.get("cost_price", 0)
        food = Food.objects.create(
            category=category,
            name=request.POST.get("name"),
            description=request.POST.get("description"),
            cost_price=cost_price,
            default_purchase_cost=cost_price,
            price=request.POST.get("price"),
            wholesale_price=ws_price if ws_price else None,
            wholesale_discount_type=request.POST.get("wholesale_discount_type") or None,
            wholesale_discount_value=request.POST.get("wholesale_discount_value", 0),
            stock=stock_qty,
            sku=request.POST.get("sku", ""),
            reward_points=request.POST.get("reward_points", 0),
            image=request.FILES.get("image"),
            available=request.POST.get("available") == "on",
            discount_type=request.POST.get("discount_type") or None,
            discount_value=request.POST.get("discount_value", 0),
        )
        if stock_qty > 0:
            svc = InventoryValuationService()
            svc.create_batch(
                food=food,
                quantity=stock_qty,
                unit_cost=cost_price,
                notes='Opening stock'
            )
            _record_stock_movement(
                food=food,
                transaction_type='opening_stock',
                quantity_change=stock_qty,
                stock_before=0,
                stock_after=stock_qty,
                created_by=request.user if request.user.is_authenticated else None,
            )
        messages.success(request, "Product added successfully.")
        return redirect("users:product_list")
    return render(request, "admin/add_product.html", {"categories": categories})


@staff_member_required
@module_access_required('products')
def edit_product(request, pk):
    product = get_object_or_404(Food, pk=pk)
    categories = Category.objects.all()
    if request.method == "POST":
        old_stock = product.stock
        new_purchase_cost = Decimal(str(request.POST.get("cost_price", 0)))
        product.name = request.POST.get("name")
        product.description = request.POST.get("description")
        product.price = request.POST.get("price")
        ws_price = request.POST.get("wholesale_price")
        product.wholesale_price = ws_price if ws_price else None
        product.wholesale_discount_type = request.POST.get("wholesale_discount_type") or None
        product.wholesale_discount_value = request.POST.get("wholesale_discount_value", 0)
        product.stock = int(request.POST.get("stock", 0))
        product.sku = request.POST.get("sku", "")
        product.reward_points = request.POST.get("reward_points", 0)
        product.category = get_object_or_404(Category, id=request.POST.get("category_id"))
        product.available = request.POST.get("available") == "on"
        product.discount_type = request.POST.get("discount_type") or None
        product.discount_value = request.POST.get("discount_value", 0)
        if request.FILES.get("image"):
            product.image = request.FILES.get("image")

        # Product Cost Management:
        # If stock is 0, update cost_price as well (no existing inventory to protect)
        # If stock > 0, only update default_purchase_cost (future purchases), don't touch cost_price
        product.default_purchase_cost = new_purchase_cost
        if old_stock == 0:
            product.cost_price = new_purchase_cost

        product.save()
        qty_diff = product.stock - old_stock
        if qty_diff != 0:
            svc = InventoryValuationService()
            if qty_diff > 0:
                svc.create_batch(
                    food=product,
                    quantity=qty_diff,
                    unit_cost=product.default_purchase_cost,
                    notes=request.POST.get("stock_reason", "Stock adjustment"),
                )
            elif qty_diff < 0:
                # Consume from batches using FIFO/AVCO when stock is decreased
                svc.consume(product, abs(qty_diff))
            _record_stock_movement(
                food=product,
                transaction_type='stock_adjustment',
                quantity_change=qty_diff,
                stock_before=old_stock,
                stock_after=product.stock,
                created_by=request.user if request.user.is_authenticated else None,
                notes=request.POST.get("stock_reason", ""),
            )
        messages.success(request, "Product updated successfully")
        return redirect("users:product_list")
    return render(request, "admin/edit_product.html", {"product": product, "categories": categories})


@staff_member_required
@module_access_required('products')
def delete_product(request, pk):
    product = get_object_or_404(Food, pk=pk)
    if request.method == "POST":
        product.delete()
        messages.success(request, "Product deleted successfully")
    return redirect("users:product_list")


# =========================
# SEARCH / DATA
# =========================
@login_required
@csrf_exempt
def search_invoice(request):
    """Unified invoice search across Retail, Retail Returns, Wholesale, and Wholesale Returns."""
    if request.method == "POST":
        data = json.loads(request.body)
        q = data.get("search", "").strip()
        inv_type_filter = data.get("inv_type", "").strip()
        if not q:
            return JsonResponse({"invoices": []})

        results = []

        # 1. Retail Invoices
        if not inv_type_filter or inv_type_filter == "retail":
            retail_qs = Invoice.objects.filter(
                Q(invoice_number__icontains=q) |
                Q(customer_phone__icontains=q) |
                Q(customer_name__icontains=q) |
                Q(user__phone__icontains=q) |
                Q(user__name__icontains=q)
            ).order_by('-created_at')[:20]
            for inv in retail_qs:
                loyalty_used = inv.loyalty_points_used or 0
                results.append({
                    "type": "retail",
                    "id": inv.id,
                    "uuid_token": inv.uuid_token,
                    "invoice_number": inv.invoice_number,
                    "name": inv.customer_name or (inv.user.name if inv.user else ""),
                    "phone": inv.customer_phone or (inv.user.phone if inv.user else ""),
                    "total": float(inv.total_amount) + loyalty_used,
                    "total_stored": float(inv.total_amount),
                    "loyalty_points_used": loyalty_used,
                    "created_at": timezone.localtime(inv.created_at).strftime("%d-%m-%Y %I:%M %p"),
                    "payment_status": "paid",
                    "invoice_status": "completed",
                })

        # 2. Retail Return Invoices
        if not inv_type_filter or inv_type_filter == "retail_return":
            ret_return_qs = ReturnInvoice.objects.filter(
                invoice_type='retail',
                original_invoice__isnull=False
            ).filter(
                Q(return_number__icontains=q) |
                Q(customer_name__icontains=q) |
                Q(original_invoice__invoice_number__icontains=q)
            ).select_related('original_invoice').order_by('-created_at')[:10]
            for ret in ret_return_qs:
                results.append({
                    "type": "retail_return",
                    "id": ret.id,
                    "uuid_token": None,
                    "invoice_number": ret.return_number or f"RI-{ret.id:06d}",
                    "name": ret.customer_name or (ret.original_invoice.customer_name if ret.original_invoice else ""),
                    "email": ret.original_invoice.customer_email if ret.original_invoice else "",
                    "total": float(ret.total_refund_amount),
                    "created_at": timezone.localtime(ret.created_at).strftime("%d-%m-%Y %I:%M %p"),
                    "payment_status": "refunded",
                    "invoice_status": "returned",
                })

        # 3. Wholesale Invoices
        if not inv_type_filter or inv_type_filter == "wholesale":
            ws_qs = WholesaleInvoice.objects.filter(
                Q(invoice_number__icontains=q) |
                Q(wholesale_customer__company_name__icontains=q) |
                Q(wholesale_customer__email__icontains=q)
            ).select_related('wholesale_customer').order_by('-created_at')[:20]
            for inv in ws_qs:
                results.append({
                    "type": "wholesale",
                    "id": inv.id,
                    "uuid_token": inv.uuid_token,
                    "invoice_number": inv.invoice_number,
                    "name": inv.wholesale_customer.company_name if inv.wholesale_customer else "Wholesale",
                    "email": inv.wholesale_customer.email if inv.wholesale_customer else "",
                    "total": float(inv.total_amount),
                    "created_at": timezone.localtime(inv.created_at).strftime("%d-%m-%Y %I:%M %p"),
                    "payment_status": "paid",
                    "invoice_status": "completed",
                })

        # 4. Wholesale Return Invoices
        if not inv_type_filter or inv_type_filter == "wholesale_return":
            ws_return_qs = ReturnInvoice.objects.filter(
                invoice_type='wholesale',
                wholesale_original_invoice__isnull=False
            ).filter(
                Q(return_number__icontains=q) |
                Q(customer_name__icontains=q) |
                Q(wholesale_original_invoice__invoice_number__icontains=q)
            ).select_related('wholesale_original_invoice').order_by('-created_at')[:10]
            for ret in ws_return_qs:
                results.append({
                    "type": "wholesale_return",
                    "id": ret.id,
                    "uuid_token": None,
                    "invoice_number": ret.return_number or f"RI-{ret.id:06d}",
                    "name": ret.customer_name or (ret.wholesale_original_invoice.wholesale_customer.company_name if ret.wholesale_original_invoice and ret.wholesale_original_invoice.wholesale_customer else ""),
                    "email": ret.wholesale_original_invoice.wholesale_customer.email if ret.wholesale_original_invoice and ret.wholesale_original_invoice.wholesale_customer else "",
                    "total": float(ret.total_refund_amount),
                    "created_at": timezone.localtime(ret.created_at).strftime("%d-%m-%Y %I:%M %p"),
                    "payment_status": "refunded",
                    "invoice_status": "returned",
                })

        results.sort(key=lambda x: x["created_at"], reverse=True)
        return JsonResponse({"invoices": results[:30]})
    return JsonResponse({"invoices": []})


@csrf_exempt
def search_products(request):
    q = request.GET.get("q", "")
    category_id = request.GET.get("category", "")
    products = Food.objects.all()
    if q:
        products = products.filter(
            Q(name__icontains=q) |
            Q(barcode__icontains=q) |
            Q(product_code__icontains=q) |
            Q(sku__icontains=q) |
            Q(category__name__icontains=q)
        )
    if category_id:
        products = products.filter(category_id=category_id)
    products = products[:50]
    return JsonResponse({
        "products": [{
            "id": p.id,
            "name": p.name,
            "price": float(p.price),
            "cost_price": float(p.default_purchase_cost),
            "wholesale_price": float(p.wholesale_price) if p.wholesale_price else 0,
            "wholesale_discount_type": p.wholesale_discount_type,
            "wholesale_discount_value": float(p.wholesale_discount_value),
            "stock": p.stock,
            "barcode": p.barcode or "",
            "product_code": p.product_code or "",
            "sku": p.sku or "",
            "available": p.available and p.stock > 0,
            "category": p.category.name,
            "image": p.image.url if p.image else "",
            "discount_type": p.discount_type,
            "discount_value": float(p.discount_value),
            "category_discount_type": p.category.discount_type,
            "category_discount_value": float(p.category.discount_value),
        } for p in products]
    })


@staff_member_required
@module_access_required('products')
def add_stock(request):
    """AJAX endpoint for Add Stock modal. GET returns product details, POST creates stock purchase."""
    if request.method == 'GET':
        product_id = request.GET.get('product_id')
        if not product_id:
            return JsonResponse({'error': 'No product selected'}, status=400)
        try:
            product = Food.objects.get(pk=product_id)
        except Food.DoesNotExist:
            return JsonResponse({'error': 'Product not found'}, status=404)
        svc = InventoryValuationService()
        inventory_value = float(svc.get_product_inventory_value(product))
        batches = InventoryBatch.objects.filter(food=product, remaining_quantity__gt=0).order_by('purchase_date', 'id')
        batch_data = [{
            'id': b.id,
            'purchase_date': timezone.localtime(b.purchase_date).strftime('%Y-%m-%d %H:%M') if b.purchase_date else '',
            'quantity': b.quantity,
            'remaining_quantity': b.remaining_quantity,
            'unit_cost': float(b.unit_cost),
            'total_cost': float(b.total_cost),
            'supplier': b.supplier or '',
            'purchase_reference': b.purchase_reference or '',
        } for b in batches]
        return JsonResponse({
            'id': product.id,
            'name': product.name,
            'stock': product.stock,
            'price': float(product.price),
            'wholesale_price': float(product.wholesale_price or 0),
            'default_purchase_cost': float(product.default_purchase_cost or 0),
            'unit': product.sku or '',
            'inventory_value': inventory_value,
            'batches': batch_data,
        })

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        product_id = data.get('product_id')
        quantity = data.get('quantity')
        purchase_cost = data.get('purchase_cost')
        selling_price = data.get('selling_price')
        wholesale_price = data.get('wholesale_price', 0)
        purchase_date_str = data.get('purchase_date', '')
        supplier = data.get('supplier', '')
        invoice_number = data.get('invoice_number', '')
        remarks = data.get('remarks', '')

        # Validation
        errors = []
        if not product_id:
            errors.append('Product is required.')
        try:
            product = Food.objects.get(pk=product_id) if product_id else None
        except Food.DoesNotExist:
            errors.append('Product not found.')
            product = None

        try:
            qty = int(quantity)
            if qty <= 0:
                errors.append('Quantity must be greater than zero.')
        except (TypeError, ValueError):
            errors.append('Invalid quantity.')
            qty = 0

        try:
            cost = Decimal(str(purchase_cost))
            if cost < 0:
                errors.append('Purchase cost cannot be negative.')
        except (TypeError, ValueError):
            errors.append('Invalid purchase cost.')
            cost = Decimal('0')

        try:
            sp = Decimal(str(selling_price))
            if sp < 0:
                errors.append('Selling price cannot be negative.')
        except (TypeError, ValueError):
            errors.append('Invalid selling price.')
            sp = Decimal('0')

        try:
            wsp = Decimal(str(wholesale_price))
            if wsp < 0:
                errors.append('Wholesale price cannot be negative.')
        except (TypeError, ValueError):
            errors.append('Invalid wholesale price.')
            wsp = Decimal('0')

        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)

        try:
            with transaction.atomic():
                svc = InventoryValuationService()
                purchase_date = None
                if purchase_date_str:
                    try:
                        purchase_date = timezone.make_aware(
                            datetime.strptime(purchase_date_str, '%Y-%m-%d')
                        )
                    except (ValueError, TypeError):
                        pass
                # Create new inventory batch
                svc.create_batch(
                    food=product,
                    quantity=qty,
                    unit_cost=cost,
                    purchase_date=purchase_date,
                    supplier=supplier or '',
                    purchase_reference=invoice_number or '',
                    notes=remarks or '',
                    selling_price=sp,
                    wholesale_price=wsp,
                )
                # Update product stock
                old_stock = product.stock
                product.stock = F('stock') + qty
                product.save(update_fields=['stock'])
                product.refresh_from_db()
                # Update latest prices on product master
                Food.objects.filter(pk=product.pk).update(
                    price=sp,
                    wholesale_price=wsp if wsp > 0 else None,
                    default_purchase_cost=cost,
                )
                # Record stock movement
                _record_stock_movement(
                    food=product,
                    transaction_type='stock_purchase',
                    quantity_change=qty,
                    stock_before=old_stock,
                    stock_after=old_stock + qty,
                    reference_number=invoice_number or '',
                    created_by=request.user if request.user.is_authenticated else None,
                    notes=remarks or '',
                )
            return JsonResponse({
                'success': True,
                'message': f'Successfully added {qty} units of "{product.name}" to stock.',
                'new_stock': old_stock + qty,
            })
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'success': False, 'errors': [str(e)]}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
def barcode_lookup(request):
    if not (request.user.is_operator or request.user.is_staff):
        return JsonResponse({"error": "Access denied"}, status=403)
    barcode = request.GET.get("barcode", "").strip()
    if not barcode:
        return JsonResponse({"error": "No barcode provided"}, status=400)
    try:
        product = Food.objects.get(barcode=barcode)
        return JsonResponse({
            "id": product.id,
            "name": product.name,
            "price": float(product.price),
            "stock": product.stock,
        })
    except Food.DoesNotExist:
        return JsonResponse({"error": "Product not found"}, status=404)


@staff_member_required
@module_access_required('user_management')
def search_users(request):
    users = User.objects.all()
    return JsonResponse({
        "users": [{"id": u.id, "email": u.email, "name": u.name, "phone": u.phone} for u in users]
    })


# =========================
# DASHBOARD DATA API
# =========================
@login_required
@module_access_required('reports')
def dashboard_data(request):
    """Returns real-time dashboard stats as JSON for auto-refresh."""
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = today_start.replace(day=1)
    year_start = today_start.replace(month=1, day=1)

    revenue = Invoice.objects.aggregate(total=Sum("total_amount"))["total"] or 0
    ws_revenue = WholesaleInvoice.objects.aggregate(total=Sum("total_amount"))["total"] or 0
    revenue = float(revenue) + float(ws_revenue)
    returns_total = ReturnInvoice.objects.aggregate(total=Sum("total_refund_amount"))["total"] or 0
    net_revenue = float(revenue) - float(returns_total)

    today_sales = Invoice.objects.filter(created_at__gte=today_start).aggregate(total=Sum("total_amount"))["total"] or 0
    ws_today_sales = WholesaleInvoice.objects.filter(created_at__gte=today_start).aggregate(total=Sum("total_amount"))["total"] or 0
    today_sales = float(today_sales) + float(ws_today_sales)
    today_returns = ReturnInvoice.objects.filter(created_at__gte=today_start).aggregate(total=Sum("total_refund_amount"))["total"] or 0
    today_net_sales = float(today_sales) - float(today_returns)

    monthly_sales = Invoice.objects.filter(created_at__gte=month_start).aggregate(total=Sum("total_amount"))["total"] or 0
    ws_monthly_sales = WholesaleInvoice.objects.filter(created_at__gte=month_start).aggregate(total=Sum("total_amount"))["total"] or 0
    monthly_sales = float(monthly_sales) + float(ws_monthly_sales)
    monthly_returns = ReturnInvoice.objects.filter(created_at__gte=month_start).aggregate(total=Sum("total_refund_amount"))["total"] or 0
    monthly_net_sales = float(monthly_sales) - float(monthly_returns)

    returned_qty = ReturnInvoiceItem.objects.aggregate(total=Sum("quantity"))["total"] or 0

    return JsonResponse({
        "revenue": net_revenue,
        "gross_revenue": float(revenue),
        "returns_total": float(returns_total),
        "today_sales": float(today_sales),
        "today_returns": float(today_returns),
        "today_net_sales": today_net_sales,
        "monthly_sales": float(monthly_sales),
        "monthly_returns": float(monthly_returns),
        "monthly_net_sales": monthly_net_sales,
        "returned_qty": returned_qty,
    })


# =========================
# OPERATOR DASHBOARD
# =========================
@login_required
@module_access_required('pos')
def operator_dashboard(request):
    if not request.user.is_operator:
        return JsonResponse({"message": "Access denied"}, status=403)
    categories = Category.objects.filter(is_active=True)
    can_create_invoice = has_permission(request.user, 'can_create_invoice')
    can_returns = has_permission(request.user, 'can_returns')
    return render(request, "operator/dashboard.html", {
        "categories": categories,
        "can_create_invoice": can_create_invoice,
        "can_returns": can_returns,
    })


@login_required
@module_access_required('pos')
def operator_session_status(request):
    if not request.user.is_operator:
        return JsonResponse({"error": "Access denied"}, status=403)
    from cash_handling.models import CashDrawerSession
    session = CashDrawerSession.objects.filter(user=request.user, status='open').first()
    if not session:
        return JsonResponse({
            'has_active_session': False,
            'session_open_perm': has_permission(request.user, 'cash_session_open'),
            'session_close_perm': has_permission(request.user, 'cash_session_close'),
        })
    balance = float(CashDrawerService.get_session_balance(session))
    return JsonResponse({
        'has_active_session': True,
        'session_id': session.id,
        'opened_at': session.opened_at.isoformat(),
        'opening_balance': float(session.opening_balance),
        'current_balance': balance,
        'operator_name': session.user.name or session.user.email,
        'session_open_perm': has_permission(request.user, 'cash_session_open'),
        'session_close_perm': has_permission(request.user, 'cash_session_close'),
    })


# =========================
# OPERATOR USER MANAGEMENT
# =========================
@staff_member_required
@module_access_required('user_management')
@csrf_exempt
def create_operator(request):
    if request.method == "POST":
        data = json.loads(request.body)
        if User.objects.filter(email=data.get("email")).exists():
            return JsonResponse({"message": "Email already exists"}, status=400)
        user = User.objects.create_user(name=data.get("name"), email=data.get("email"), password=data.get("password"))
        user.is_operator = True
        user.is_staff = False
        user.is_superuser = False
        user.save()
        return JsonResponse({"message": "Operator user created"})
    return JsonResponse({"message": "Invalid request"}, status=400)


@staff_member_required
@module_access_required('user_management')
def operator_list(request):
    operators = User.objects.filter(is_operator=True)
    return JsonResponse({
        "operators": [{"id": k.id, "name": k.name, "email": k.email} for k in operators]
    })


@module_access_required('user_management')
@csrf_exempt
def edit_operator(request, id):
    if request.method != "POST":
        return JsonResponse({"message": "Invalid request"}, status=400)
    try:
        user = User.objects.get(id=id, is_operator=True)
        data = json.loads(request.body)
        user.name = data.get("name")
        user.save()
        return JsonResponse({"message": "Operator user updated"})
    except User.DoesNotExist:
        return JsonResponse({"message": "Operator user not found"}, status=404)


@module_access_required('user_management')
@csrf_exempt
def delete_operator(request, id):
    User.objects.get(id=id).delete()
    return JsonResponse({"message": "Operator user deleted"})


@staff_member_required
@module_access_required('user_management')
@csrf_exempt
def toggle_operator_active(request, id):
    try:
        user = User.objects.get(id=id, is_operator=True)
        user.is_active = not user.is_active
        user.save()
        status = "activated" if user.is_active else "deactivated"
        return JsonResponse({"message": f"Operator {status}"})
    except User.DoesNotExist:
        return JsonResponse({"message": "Operator not found"}, status=404)


@staff_member_required
@module_access_required('user_management')
def reset_operator_password(request, id):
    if request.method == "POST":
        try:
            user = User.objects.get(id=id, is_operator=True)
            data = json.loads(request.body)
            password = data.get("password", "changeme123")
            user.set_password(password)
            user.save()
            return JsonResponse({"message": "Password reset successfully"})
        except User.DoesNotExist:
            return JsonResponse({"message": "Operator not found"}, status=404)
    return JsonResponse({"message": "Invalid request"}, status=400)



# =========================
# INVOICE SEARCH PAGE
# =========================
@staff_member_required
@module_access_required('reports')
def invoice_search_page(request):
    return render(request, "admin/invoices.html")


# =========================
# KITCHEN USERS PAGE
# =========================
@staff_member_required
@module_access_required('user_management')
def operator_users_page(request):
    return render(request, "admin/operator_users.html")


def _get_pos_permission_fields():
    from .context_processors import PERMISSION_MAP, PERMISSION_FLAT_MAP, POS_ONLY_KEYS, _deep_resolve_flat
    flat = {}
    flat.update(_deep_resolve_flat(PERMISSION_MAP))
    flat.update(PERMISSION_FLAT_MAP)
    return sorted(field for key, field in flat.items() if key in POS_ONLY_KEYS and field is not None)


def _get_pos_permission_groups():
    from .context_processors import PERMISSION_MAP, PERMISSION_FLAT_MAP, POS_ONLY_KEYS, _deep_resolve_flat
    flat = {}
    flat.update(_deep_resolve_flat(PERMISSION_MAP))
    flat.update(PERMISSION_FLAT_MAP)
    pos_perms = [(k, v) for k, v in flat.items() if k in POS_ONLY_KEYS and v is not None]

    MODULE_LABELS = {
        'sales': 'Sales & POS',
        'customers': 'Customers',
        'return_invoice': 'Returns',
        'wholesale': 'Wholesale',
        'cash_handling': 'Cash Handling',
    }
    MODULE_ORDER = ['sales', 'customers', 'return_invoice', 'wholesale', 'cash_handling']

    by_module = {}
    for key, field in pos_perms:
        parts = key.split('.')
        mkey = parts[0]
        if mkey not in by_module:
            by_module[mkey] = {
                'module': MODULE_LABELS.get(mkey, mkey.replace('_', ' ').title()),
                'module_key': mkey,
                'permissions': [],
            }
        by_module[mkey]['permissions'].append({
            'field': field,
            'label': field.replace('can_', '').replace('_', ' ').strip().title(),
        })

    result = []
    for mk in MODULE_ORDER:
        if mk in by_module:
            by_module[mk]['permissions'].sort(key=lambda p: p['label'])
            result.append(by_module[mk])
    return result


@staff_member_required
@module_access_required('user_management')
def admin_operator_permissions(request):
    pos_fields = _get_pos_permission_fields()

    users = User.objects.filter(
        is_operator=True
    ).exclude(is_software_owner=True).exclude(is_superuser=True).prefetch_related('operator_permissions').distinct()

    if request.method == "POST":
        # Bulk save support
        import json
        try:
            body = json.loads(request.body) if request.body else {}
        except Exception:
            body = {}
        changes = body.get('changes') if isinstance(body, dict) else None

        if changes is not None:
            if not isinstance(changes, list) or len(changes) == 0:
                return JsonResponse({"success": False, "error": "No changes provided"}, status=400)
            results = []
            errors = []
            for change in changes:
                uid = change.get('user_id')
                pname = change.get('perm_name')
                val = change.get('value', False)
                if not uid or not pname:
                    errors.append("Missing user_id or perm_name")
                    continue
                if pname not in pos_fields:
                    errors.append(f"Invalid permission: {pname}")
                    continue
                try:
                    target = User.objects.get(pk=uid)
                    if target.is_software_owner or target.is_superuser:
                        errors.append("Cannot modify superuser or software owner")
                        continue
                    perm, _ = OperatorPermission.objects.get_or_create(user=target)
                    setattr(perm, pname, val)
                    perm.save()
                    AuditLog.objects.create(
                        user=request.user,
                        action='permission_change',
                        description=f"{'Enabled' if val else 'Disabled'} '{pname}' for {target.email}",
                        ip_address=request.META.get('REMOTE_ADDR'),
                    )
                    results.append(True)
                except Exception as e:
                    errors.append(str(e))
            return JsonResponse({"success": len(errors) == 0, "updated": len(results), "errors": errors})
        else:
            user_id = request.POST.get("user_id")
            perm_name = request.POST.get("perm_name")
            value = request.POST.get("value") == "1"
            if not user_id or not perm_name:
                return JsonResponse({"success": False, "error": "Missing user_id or perm_name"}, status=400)
            try:
                target = get_object_or_404(User, pk=user_id)
                if target.is_software_owner or target.is_superuser:
                    return JsonResponse({"success": False, "error": "Cannot modify superuser or software owner"}, status=400)
                if perm_name not in pos_fields:
                    return JsonResponse({"success": False, "error": "Invalid permission"}, status=400)
                perm, _ = OperatorPermission.objects.get_or_create(user=target)
                setattr(perm, perm_name, value)
                perm.save()
                AuditLog.objects.create(
                    user=request.user,
                    action='permission_change',
                    description=f"{'Enabled' if value else 'Disabled'} '{perm_name}' for {target.email}",
                    ip_address=request.META.get('REMOTE_ADDR'),
                )
                return JsonResponse({"success": True})
            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)}, status=500)

    for u in users:
        try:
            _ = u.operator_permissions
        except Exception:
            OperatorPermission.objects.get_or_create(user=u)

    return render(request, "admin/operator_permissions.html", {
        "operators": users,
        "permission_fields": pos_fields,
        "permission_groups": _get_pos_permission_groups(),
    })


# =========================
# TABLE MANAGEMENT
# =========================
# =========================
# CATEGORY MANAGEMENT
# =========================
@staff_member_required
@module_access_required('products')
def category_list(request):
    categories = Category.objects.all()
    return render(request, "admin/categories.html", {"categories": categories})


@staff_member_required
@module_access_required('products')
def category_add(request):
    if request.method == "POST":
        name = request.POST.get("name")
        if name:
            Category.objects.create(
                name=name,
                description=request.POST.get("description", ""),
                is_active=request.POST.get("is_active") == "on",
                discount_type=request.POST.get("discount_type") or None,
                discount_value=request.POST.get("discount_value", 0),
            )
            messages.success(request, "Category added successfully")
        return redirect("users:category_list")
    return redirect("users:category_list")


@staff_member_required
@module_access_required('products')
def category_edit(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        cat.name = request.POST.get("name", cat.name)
        cat.description = request.POST.get("description", "")
        cat.is_active = request.POST.get("is_active") == "on"
        cat.discount_type = request.POST.get("discount_type") or None
        cat.discount_value = request.POST.get("discount_value", 0)
        cat.save()
        messages.success(request, "Category updated successfully")
        return redirect("users:category_list")
    return redirect("users:category_list")


@staff_member_required
@module_access_required('products')
def category_delete(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        cat.delete()
        messages.success(request, "Category deleted successfully")
    return redirect("users:category_list")


@staff_member_required
@module_access_required('products')
def category_toggle_active(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    cat.is_active = not cat.is_active
    cat.save()
    status = "activated" if cat.is_active else "deactivated"
    return JsonResponse({"message": f"Category {status}"})


# =========================
# MYSQL BACKUP
# =========================


@staff_member_required
@module_access_required('user_management')
def mysql_backup(request):
    if not request.user.is_staff:
        return redirect('users:admin_dashboard')
    from megaone.users.backup_utils import get_backup_metadata, get_backup_file_size, BACKUP_FILE, should_run_auto_backup
    metadata = get_backup_metadata()
    file_size_bytes = get_backup_file_size()
    backup_exists = BACKUP_FILE.exists()

    if file_size_bytes > 0:
        if file_size_bytes < 1024:
            file_size = f"{file_size_bytes} B"
        elif file_size_bytes < 1024 * 1024:
            file_size = f"{file_size_bytes / 1024:.1f} KB"
        else:
            file_size = f"{file_size_bytes / (1024 * 1024):.1f} MB"
    else:
        file_size = "No backup"

    next_backup = "N/A"
    if metadata.get("auto_backup_enabled") and metadata.get("last_backup_time"):
        from datetime import datetime, timedelta
        try:
            last_time = datetime.fromisoformat(metadata["last_backup_time"])
            next_time = last_time + timedelta(hours=24)
            next_backup = next_time.strftime("%Y-%m-%d %H:%M:%S")
        except (ValueError, TypeError):
            next_backup = "After next backup"
    elif metadata.get("auto_backup_enabled") and not metadata.get("last_backup_time"):
        next_backup = "Immediately"

    return render(request, "admin/backup.html", {
        "metadata": metadata,
        "file_size": file_size,
        "backup_exists": backup_exists,
        "next_backup": next_backup,
        "active_page": "backup",
    })


@login_required
@module_access_required('user_management')
@require_POST
def backup_run(request):
    if not request.user.is_staff:
        return JsonResponse({"success": False, "error": "Access denied"}, status=403)
    from megaone.users.backup_utils import run_backup
    result = run_backup()
    return JsonResponse(result)


@login_required
@module_access_required('user_management')
@require_POST
def backup_toggle_auto(request):
    if not request.user.is_staff:
        return JsonResponse({"success": False, "error": "Access denied"}, status=403)
    from megaone.users.backup_utils import get_backup_metadata, save_backup_metadata
    metadata = get_backup_metadata()
    metadata["auto_backup_enabled"] = not metadata.get("auto_backup_enabled", True)
    save_backup_metadata(metadata)
    return JsonResponse({"success": True, "auto_backup_enabled": metadata["auto_backup_enabled"]})


@login_required
def backup_download(request):
    if not request.user.is_staff:
        return redirect('users:admin_dashboard')
    from megaone.users.backup_utils import BACKUP_FILE
    if not BACKUP_FILE.exists():
        messages.error(request, "No backup file available to download.")
        return redirect('users:mysql_backup')
    response = FileResponse(open(BACKUP_FILE, "rb"), content_type="application/sql")
    response["Content-Disposition"] = 'attachment; filename="backup.sql"'
    return response


# =============================================================================
# LOYALTY CARD VIEWS
# =============================================================================

@login_required
def loyalty_card_view(request):
    if not request.user.is_authenticated:
        messages.warning(request, "Please log in to view your loyalty information.")
        return redirect(f"{reverse('users:login')}?next={reverse('users:loyalty_card_view')}")
    if request.user.is_staff or request.user.is_superuser or getattr(request.user, 'is_operator', False):
        messages.error(request, "Loyalty cards are only available for customers.")
        return redirect('/')
    card = LoyaltyCard.objects.filter(user=request.user).first()
    if not card:
        card = LoyaltyCard.objects.create(user=request.user, status='ACTIVE')
    qr_ok = card.qr_code_image and card.qr_code_image.storage.exists(card.qr_code_image.name)
    if not qr_ok:
        try:
            generate_qr_code_image(card, request)
        except Exception:
            pass
    if not card.card_pdf or not card.card_image:
        try:
            generate_loyalty_card_pdf(card, request)
            generate_loyalty_card_image(card, request)
        except Exception:
            pass
    show_welcome = not card.first_card_popup_shown
    if show_welcome:
        card.first_card_popup_shown = True
        card.save(update_fields=['first_card_popup_shown'])
    transactions = LoyaltyTransaction.objects.filter(card=card).order_by('-created_at')
    return render(request, 'users/loyalty_card.html', {
        'card': card,
        'transactions': transactions,
        'show_welcome': show_welcome,
    })


@login_required
def download_loyalty_pdf(request, card_number):
    try:
        card = get_object_or_404(LoyaltyCard, card_number=card_number, user=request.user)
        if not card.qr_code_image or not card.qr_code_image.storage.exists(card.qr_code_image.name):
            generate_qr_code_image(card, request)
            card.refresh_from_db()
        if not card.card_pdf or not card.card_pdf.storage.exists(card.card_pdf.name):
            card = generate_loyalty_card_pdf(card, request)
        card.card_pdf.open('rb')
        response = FileResponse(card.card_pdf, as_attachment=True, filename=f"loyalty_card_{card.card_number}.pdf")
        response['Content-Type'] = 'application/pdf'
        response['Content-Disposition'] = f'attachment; filename="loyalty_card_{card.card_number}.pdf"'
        return response
    except Exception as e:
        messages.error(request, f"Could not generate PDF. Please try again. Error: {str(e)}")
        return redirect('users:loyalty_card_view')


@login_required
def download_loyalty_image(request, card_number):
    try:
        card = get_object_or_404(LoyaltyCard, card_number=card_number, user=request.user)
        if not card.qr_code_image or not card.qr_code_image.storage.exists(card.qr_code_image.name):
            generate_qr_code_image(card, request)
            card.refresh_from_db()
        if not card.card_image or not card.card_image.storage.exists(card.card_image.name):
            card = generate_loyalty_card_image(card, request)
        card.card_image.open('rb')
        response = FileResponse(card.card_image, as_attachment=True, filename=f"loyalty_card_{card.card_number}.png")
        response['Content-Type'] = 'image/png'
        response['Content-Disposition'] = f'attachment; filename="loyalty_card_{card.card_number}.png"'
        return response
    except Exception as e:
        messages.error(request, f"Could not generate card image. Please try again. Error: {str(e)}")
        return redirect('users:loyalty_card_view')


@login_required
def loyalty_card_data(request):
    card = LoyaltyCard.objects.filter(user=request.user).first()
    if not card:
        return JsonResponse({'has_card': False})
    return JsonResponse({
        'has_card': True,
        'card_number': card.card_number,
        'total_points': card.total_points,
        'used_points': card.used_points,
        'remaining_points': card.remaining_points,
        'status': card.status,
    })


@login_required
def loyalty_checkout_info(request):
    card = LoyaltyCard.objects.filter(user=request.user, status='ACTIVE').first()
    if card:
        return JsonResponse({
            'has_card': True,
            'total_points': card.total_points,
            'remaining_points': card.remaining_points,
            'card_number': card.card_number,
        })
    return JsonResponse({'has_card': False})


@csrf_exempt
def verify_loyalty_qr(request, qr_token):
    if request.method == 'GET':
        card = LoyaltyCard.objects.filter(qr_token=qr_token).first()
        if not card:
            return JsonResponse({'valid': False, 'error': 'Invalid Loyalty Card'}, status=404)
        if card.status != 'ACTIVE':
            return JsonResponse({'valid': False, 'error': 'Loyalty Card is blocked'}, status=403)
        user_name = card.user.name if card.user else 'Customer'
        return JsonResponse({
            'valid': True,
            'card_number': card.card_number,
            'customer': user_name,
            'card_status': card.status,
        })
    return JsonResponse({'valid': False}, status=405)


def qr_loyalty_redirect(request, qr_token):
    card = LoyaltyCard.objects.filter(qr_token=qr_token).first()
    if not card:
        messages.error(request, "Invalid Loyalty Card")
        return redirect("users:login")
    if card.status != 'ACTIVE':
        messages.error(request, "Loyalty Card is blocked")
        return redirect("users:login")
    if not request.user.is_authenticated:
        login_url = reverse("users:login")
        next_url = reverse("users:qr_loyalty_redirect", args=[qr_token])
        return redirect(f"{login_url}?next={next_url}")
    if card.user != request.user:
        messages.error(request, "Invalid Loyalty Card")
        return redirect("users:login")
    return redirect("users:loyalty_card_view")


@csrf_exempt
@login_required
def loyalty_checkout_validate(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            total_amount = float(data.get('total_amount', 0))
            use_points = data.get('use_points')
            if use_points is not None:
                use_points = int(use_points)
            card = LoyaltyCard.objects.filter(user=request.user, status='ACTIVE').first()
            if not card:
                return JsonResponse({'can_pay': False, 'error': 'No active loyalty card found'})
            available = card.remaining_points
            if available <= 0:
                return JsonResponse({
                    'can_pay': False,
                    'error': 'No loyalty points available.',
                    'available_points': 0,
                })
            max_allowed = min(available, int(total_amount))
            if use_points is None:
                use_points = max_allowed
            elif use_points < 0:
                return JsonResponse({'can_pay': False, 'error': 'Points cannot be negative', 'available_points': available})
            elif use_points > available:
                return JsonResponse({
                    'can_pay': False,
                    'error': f'You only have {available} loyalty points available.',
                    'available_points': available,
                })
            elif use_points > int(total_amount):
                use_points = int(total_amount)
            remaining_due = int(total_amount) - use_points
            return JsonResponse({
                'can_pay': True,
                'available_points': available,
                'points_to_use': use_points,
                'remaining_due': remaining_due if remaining_due > 0 else 0,
                'needs_secondary': remaining_due > 0,
                'card_number': card.card_number,
            })
        except Exception as e:
            return JsonResponse({'can_pay': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid method'}, status=405)


# Admin loyalty views
@staff_member_required
@module_access_required('customers')
def admin_loyalty_list(request):
    cards = LoyaltyCard.objects.select_related('user').all().order_by('-created_at')
    search = request.GET.get('search', '')
    if search:
        cards = cards.filter(
            Q(card_number__icontains=search) |
            Q(user__name__icontains=search) |
            Q(user__phone__icontains=search)
        )
    status_filter = request.GET.get('status', '')
    if status_filter:
        cards = cards.filter(status=status_filter)
    return render(request, 'users/admin_loyalty_list.html', {
        'cards': cards,
        'search': search,
        'status_filter': status_filter,
    })


@staff_member_required
@module_access_required('customers')
def admin_loyalty_detail(request, card_number):
    card = get_object_or_404(LoyaltyCard, card_number=card_number)
    transactions = LoyaltyTransaction.objects.filter(card=card).order_by('-created_at')
    return render(request, 'users/admin_loyalty_detail.html', {
        'card': card,
        'transactions': transactions,
    })


@staff_member_required
@module_access_required('customers')
def admin_toggle_card_status(request, card_number):
    card = get_object_or_404(LoyaltyCard, card_number=card_number)
    if request.method == 'POST':
        card.status = 'BLOCKED' if card.status == 'ACTIVE' else 'ACTIVE'
        card.save()
        messages.success(request, f"Card {card.card_number} is now {card.get_status_display()}")
    return redirect('users:admin_loyalty_detail', card_number=card.card_number)


@staff_member_required
@module_access_required('customers')
def admin_reset_points(request, card_number):
    card = get_object_or_404(LoyaltyCard, card_number=card_number)
    if request.method == 'POST':
        card.total_points = 0
        card.used_points = 0
        card.remaining_points = 0
        card.save()
        LoyaltyTransaction.objects.filter(card=card).delete()
        messages.success(request, f"Points reset for card {card.card_number}")
    return redirect('users:admin_loyalty_detail', card_number=card.card_number)


@login_required
def loyalty_transactions(request):
    card = LoyaltyCard.objects.filter(user=request.user).first()
    if not card:
        return JsonResponse({'transactions': []})
    transactions = LoyaltyTransaction.objects.filter(card=card).order_by('-created_at')
    return JsonResponse({
        'transactions': [{
            'id': t.id,
            'order_number': t.order_number,
            'earned_points': t.earned_points,
            'redeemed_points': t.redeemed_points,
            'remaining_balance': t.remaining_balance,
            'transaction_type': t.transaction_type,
            'created_at': t.created_at.strftime('%d-%m-%Y %I:%M %p'),
        } for t in transactions],
    })


# =========================
# OFFERS & DEALS API
# =========================
def active_offer_data(request):
    offer = _get_active_time_offer()
    if not offer:
        return JsonResponse({"active": False})
    now = timezone.now()
    end_dt = timezone.make_aware(
        datetime.combine(offer.end_date, offer.end_time)
    )
    return JsonResponse({
        "active": True,
        "id": offer.id,
        "title": offer.title,
        "description": offer.description or "",
        "discount_percentage": float(offer.discount_percentage),
        "banner_image": offer.banner_image.url if offer.banner_image else "",
        "background_color": offer.background_color or "#f59e0b",
        "popup_image": offer.popup_image.url if offer.popup_image else "",
        "start_date": offer.start_date.strftime("%d-%m-%Y"),
        "start_time": offer.start_time.strftime("%I:%M %p"),
        "end_date": offer.end_date.strftime("%d-%m-%Y"),
        "end_time": offer.end_time.strftime("%I:%M %p"),
        "end_timestamp": int(end_dt.timestamp()),
    })


def active_deal_data(request):
    deal = _get_active_deal()
    if not deal:
        return JsonResponse({"active": False})
    now = timezone.now()
    end_dt = timezone.make_aware(
        datetime.combine(deal.end_date, deal.end_time)
    )
    deal_products = deal.products.all()
    products_data = []
    original_total = 0
    for p in deal_products:
        products_data.append({
            "id": p.id,
            "name": p.name,
            "price": float(p.price),
            "image": p.image.url if p.image else "",
        })
        original_total += float(p.price)
    if deal.free_product:
        products_data.append({
            "id": deal.free_product.id,
            "name": deal.free_product.name + " (Free)",
            "price": float(deal.free_product.price),
            "image": deal.free_product.image.url if deal.free_product.image else "",
            "free": True,
        })
        original_total += float(deal.free_product.price)
    savings = 0
    if deal.combo_price and original_total > 0:
        savings = original_total - float(deal.combo_price)
        if savings < 0:
            savings = 0
    return JsonResponse({
        "active": True,
        "id": deal.id,
        "title": deal.title,
        "description": deal.description or "",
        "deal_image": deal.deal_image.url if deal.deal_image else "",
        "deal_banner": deal.deal_banner.url if deal.deal_banner else "",
        "products": products_data,
        "combo_price": float(deal.combo_price) if deal.combo_price else None,
        "original_total": original_total,
        "savings": savings,
        "start_date": deal.start_date.strftime("%d-%m-%Y"),
        "start_time": deal.start_time.strftime("%I:%M %p"),
        "end_date": deal.end_date.strftime("%d-%m-%Y"),
        "end_time": deal.end_time.strftime("%I:%M %p"),
        "end_timestamp": int(end_dt.timestamp()),
    })


def offer_banner_data(request):
    offer = _get_active_time_offer()
    deal = _get_active_deal()
    data = {"offer": None, "deal": None}
    if offer:
        end_dt = timezone.make_aware(
            datetime.combine(offer.end_date, offer.end_time)
        )
        data["offer"] = {
            "title": offer.title,
            "discount_percentage": float(offer.discount_percentage),
            "background_color": offer.background_color or "#f59e0b",
            "end_timestamp": int(end_dt.timestamp()),
        }
    if deal:
        end_dt = timezone.make_aware(
            datetime.combine(deal.end_date, deal.end_time)
        )
        deal_products = deal.products.all()
        original_total = sum(float(p.price) for p in deal_products)
        if deal.free_product:
            original_total += float(deal.free_product.price)
        savings = 0
        if deal.combo_price and original_total > 0:
            savings = original_total - float(deal.combo_price)
            if savings < 0:
                savings = 0
        data["deal"] = {
            "id": deal.id,
            "title": deal.title,
            "description": deal.description or "",
            "deal_banner": deal.deal_banner.url if deal.deal_banner else "",
            "combo_price": float(deal.combo_price) if deal.combo_price else None,
            "original_total": original_total,
            "savings": savings,
            "end_timestamp": int(end_dt.timestamp()),
        }
    return JsonResponse(data)


# =========================
# OFFER CRUD VIEWS
# =========================
@staff_member_required
@module_access_required('products')
def offer_list(request):
    offers = TimeBasedOffer.objects.all().order_by("-created_at")
    total_offers = offers.count()
    active_offers = offers.filter(is_active=True).count()
    expired_offers = total_offers - active_offers
    usage_count = Invoice.objects.filter(qr_offer_discount_amount__gt=0).count()
    return render(request, "admin/offer_list.html", {
        "offers": offers,
        "total_offers": total_offers,
        "active_offers": active_offers,
        "expired_offers": expired_offers,
        "usage_count": usage_count,
        "active_page": "offers",
    })


@staff_member_required
@module_access_required('products')
def offer_add(request):
    if request.method == "POST":
        try:
            offer = TimeBasedOffer(
                title=request.POST.get("title"),
                description=request.POST.get("description", ""),
                discount_percentage=request.POST.get("discount_percentage", 0),
                background_color=request.POST.get("background_color", "#f59e0b"),
                is_active=request.POST.get("is_active") == "1",
                start_date=request.POST.get("start_date"),
                start_time=request.POST.get("start_time"),
                end_date=request.POST.get("end_date"),
                end_time=request.POST.get("end_time"),
            )
            if "banner_image" in request.FILES:
                offer.banner_image = request.FILES["banner_image"]
            if "popup_image" in request.FILES:
                offer.popup_image = request.FILES["popup_image"]
            offer.save()
            messages.success(request, "Offer created successfully.")
            return redirect("users:offer_list")
        except Exception as e:
            messages.error(request, f"Error: {e}")
    return render(request, "admin/offer_form.html", {"active_page": "offers"})


@staff_member_required
@module_access_required('products')
def offer_edit(request, pk):
    offer = get_object_or_404(TimeBasedOffer, pk=pk)
    if request.method == "POST":
        try:
            offer.title = request.POST.get("title")
            offer.description = request.POST.get("description", "")
            offer.discount_percentage = request.POST.get("discount_percentage", 0)
            offer.background_color = request.POST.get("background_color", "#f59e0b")
            offer.is_active = request.POST.get("is_active") == "1"
            offer.start_date = request.POST.get("start_date")
            offer.start_time = request.POST.get("start_time")
            offer.end_date = request.POST.get("end_date")
            offer.end_time = request.POST.get("end_time")
            if "banner_image" in request.FILES:
                offer.banner_image = request.FILES["banner_image"]
            if "popup_image" in request.FILES:
                offer.popup_image = request.FILES["popup_image"]
            offer.save()
            messages.success(request, "Offer updated successfully.")
            return redirect("users:offer_list")
        except Exception as e:
            messages.error(request, f"Error: {e}")
    return render(request, "admin/offer_form.html", {"offer": offer, "active_page": "offers"})


@staff_member_required
@module_access_required('products')
def offer_detail(request, pk):
    offer = get_object_or_404(TimeBasedOffer, pk=pk)
    return render(request, "admin/offer_detail.html", {"offer": offer, "active_page": "offers"})


@staff_member_required
@module_access_required('products')
def offer_delete(request, pk):
    offer = get_object_or_404(TimeBasedOffer, pk=pk)
    if request.method == "POST":
        try:
            offer.delete()
            messages.success(request, "Offer deleted successfully.")
        except Exception as e:
            messages.error(request, f"Error: {e}")
    return redirect("users:offer_list")


# =========================
# DEAL CRUD VIEWS
# =========================
@staff_member_required
@module_access_required('products')
def deal_list(request):
    deals = TodayDeal.objects.all().order_by("-created_at")
    total_deals = deals.count()
    active_deals = deals.filter(is_active=True).count()
    expired_deals = total_deals - active_deals
    return render(request, "admin/deal_list.html", {
        "deals": deals,
        "total_deals": total_deals,
        "active_deals": active_deals,
        "expired_deals": expired_deals,
        "active_page": "deals",
    })


@staff_member_required
@module_access_required('products')
def deal_add(request):
    if request.method == "POST":
        try:
            deal = TodayDeal(
                title=request.POST.get("title"),
                description=request.POST.get("description", ""),
                is_active=request.POST.get("is_active") == "1",
                start_date=request.POST.get("start_date"),
                start_time=request.POST.get("start_time"),
                end_date=request.POST.get("end_date"),
                end_time=request.POST.get("end_time"),
                combo_price=request.POST.get("combo_price") or None,
                discount_percentage=request.POST.get("discount_percentage") or None,
                free_product_id=request.POST.get("free_product") or None,
            )
            if "deal_image" in request.FILES:
                deal.deal_image = request.FILES["deal_image"]
            if "deal_banner" in request.FILES:
                deal.deal_banner = request.FILES["deal_banner"]
            deal.save()
            product_ids = request.POST.getlist("products")
            if product_ids:
                deal.products.set(Food.objects.filter(id__in=product_ids))
            messages.success(request, "Deal created successfully.")
            return redirect("users:deal_list")
        except Exception as e:
            messages.error(request, f"Error: {e}")
    return render(request, "admin/deal_form.html", {"active_page": "deals", "products": products})


@staff_member_required
@module_access_required('products')
def deal_edit(request, pk):
    deal = get_object_or_404(TodayDeal, pk=pk)
    products = Food.objects.all()
    if request.method == "POST":
        try:
            deal.title = request.POST.get("title")
            deal.description = request.POST.get("description", "")
            deal.is_active = request.POST.get("is_active") == "1"
            deal.start_date = request.POST.get("start_date")
            deal.start_time = request.POST.get("start_time")
            deal.end_date = request.POST.get("end_date")
            deal.end_time = request.POST.get("end_time")
            deal.combo_price = request.POST.get("combo_price") or None
            deal.discount_percentage = request.POST.get("discount_percentage") or None
            deal.free_product_id = request.POST.get("free_product") or None
            if "deal_image" in request.FILES:
                deal.deal_image = request.FILES["deal_image"]
            if "deal_banner" in request.FILES:
                deal.deal_banner = request.FILES["deal_banner"]
            deal.save()
            product_ids = request.POST.getlist("products")
            if product_ids:
                deal.products.set(Food.objects.filter(id__in=product_ids))
            else:
                deal.products.clear()
            messages.success(request, "Deal updated successfully.")
            return redirect("users:deal_list")
        except Exception as e:
            messages.error(request, f"Error: {e}")
    return render(request, "admin/deal_form.html", {"deal": deal, "active_page": "deals", "products": products})


@staff_member_required
@module_access_required('products')
def deal_detail(request, pk):
    deal = get_object_or_404(TodayDeal, pk=pk)
    deal_products = deal.products.all()
    original_total = sum(float(p.price) for p in deal_products)
    if deal.free_product:
        original_total += float(deal.free_product.price)
    savings = 0
    deal_price = original_total
    if deal.deal_type == 'combo_price' and deal.combo_price:
        deal_price = float(deal.combo_price)
        savings = original_total - deal_price
        if savings < 0:
            savings = 0
    elif deal.deal_type == 'free_product' and deal.free_product:
        free_price = float(deal.free_product.price)
        deal_price = original_total - free_price
        savings = free_price
    elif deal.deal_type == 'percentage' and deal.discount_percentage:
        pct = float(deal.discount_percentage)
        deal_price = round(original_total - (original_total * pct / 100), 2)
        savings = original_total - deal_price
    return render(request, "admin/deal_detail.html", {
        "deal": deal,
        "deal_products": deal_products,
        "original_total": original_total,
        "savings": savings,
        "deal_price": deal_price,
        "active_page": "deals",
    })


@staff_member_required
@module_access_required('products')
def deal_delete(request, pk):
    deal = get_object_or_404(TodayDeal, pk=pk)
    if request.method == "POST":
        try:
            deal.delete()
            messages.success(request, "Deal deleted successfully.")
        except Exception as e:
            messages.error(request, f"Error: {e}")
    return redirect("users:deal_list")


# =========================
# PUBLIC DEAL VIEWS
# =========================
@login_required(login_url='users:login')
def public_deal_detail(request, pk):
    deal = get_object_or_404(TodayDeal, pk=pk)
    if not deal.is_active:
        _get_active_deal()
    if not deal.is_active:
        return render(request, "food-delivery/deal_detail.html", {"deal": None, "error": "This deal is no longer active."})
    deal_products = deal.products.all()
    original_total = sum(float(p.price) for p in deal_products)
    if deal.free_product:
        original_total += float(deal.free_product.price)
    savings = 0
    deal_price = original_total
    if deal.deal_type == 'combo_price' and deal.combo_price:
        deal_price = float(deal.combo_price)
        savings = original_total - deal_price
        if savings < 0:
            savings = 0
    elif deal.deal_type == 'free_product' and deal.free_product:
        free_price = float(deal.free_product.price)
        deal_price = original_total - free_price
        savings = free_price
    elif deal.deal_type == 'percentage' and deal.discount_percentage:
        pct = float(deal.discount_percentage)
        deal_price = round(original_total - (original_total * pct / 100), 2)
        savings = original_total - deal_price
    active_offer = _get_active_time_offer()
    active_deal = _get_active_deal()
    return render(request, "food-delivery/deal_detail.html", {
        "deal": deal,
        "deal_products": deal_products,
        "original_total": original_total,
        "savings": savings,
        "deal_price": deal_price,
        "active_offer": active_offer,
        "active_deal": active_deal,
    })


@login_required(login_url='users:login')
def deal_checkout(request, pk):
    deal = get_object_or_404(TodayDeal, pk=pk)
    if not deal.is_active:
        _get_active_deal()
    if not deal.is_active:
        messages.error(request, "This deal is no longer active.")
        return redirect("/")
    if request.method == "POST":
        deal_products = list(deal.products.all())
        original_total = sum(float(p.price) for p in deal_products)
        free_product = deal.free_product
        if free_product:
            deal_products.append(free_product)
            original_total += float(free_product.price)

        cart_data = []
        effective_total = original_total
        for p in deal_products:
            is_free = free_product and p.id == free_product.id
            price = 0 if is_free else float(p.price)
            cart_data.append({
                "id": p.id,
                "name": p.name,
                "price": price,
                "image": p.image.url if p.image else "/static/food-delivery/img/item1.png",
                "qty": 1,
                "is_free": is_free,
            })

        if deal.deal_type == 'combo_price' and deal.combo_price:
            effective_total = float(deal.combo_price)
        elif deal.deal_type == 'free_product' and free_product:
            effective_total = original_total - float(free_product.price)
        elif deal.deal_type == 'percentage' and deal.discount_percentage:
            pct = float(deal.discount_percentage)
            effective_total = round(original_total - (original_total * pct / 100), 2)

        request.session["deal_checkout_cart"] = json.dumps(cart_data)
        request.session["deal_checkout_id"] = deal.id
        request.session["deal_effective_total"] = effective_total
        return redirect("users:food_delivery_restaurant_detail")
    return redirect("users:public_deal_detail", pk=pk)


@login_required
@csrf_exempt
def clear_deal_cart(request):
    if request.method == "POST":
        request.session.pop("deal_checkout_cart", None)
        request.session.pop("deal_checkout_id", None)
        request.session.pop("deal_effective_total", None)
        return JsonResponse({"success": True})
    return JsonResponse({"success": False}, status=400)


def _get_unit_cost(item, food):
    """Return the cost per unit for an invoice item.
    Uses SaleItemCost records if available, otherwise falls back to unit_cost_at_sale or food.default_purchase_cost."""
    from .models import InvoiceItem, WholesaleInvoiceItem, SaleItemCost

    if isinstance(item, InvoiceItem):
        costs = SaleItemCost.objects.filter(invoice_item=item)
    elif isinstance(item, WholesaleInvoiceItem):
        costs = SaleItemCost.objects.filter(wholesale_invoice_item=item)
    else:
        costs = SaleItemCost.objects.none()

    if costs.exists():
        total_cost = sum(Decimal(str(c.quantity)) * c.unit_cost for c in costs)
        total_qty = sum(c.quantity for c in costs)
        if total_qty > 0:
            return float(total_cost / Decimal(str(total_qty)))
    if hasattr(item, 'unit_cost_at_sale') and item.unit_cost_at_sale and float(item.unit_cost_at_sale) > 0:
        return float(item.unit_cost_at_sale)
    if food:
        return float(food.default_purchase_cost or 0)
    return 0


def _record_stock_movement(food, transaction_type, quantity_change, stock_before, stock_after, reference_number='', created_by=None, notes=''):
    from .models import StockMovement
    StockMovement.objects.create(
        food=food,
        transaction_type=transaction_type,
        reference_number=reference_number or '',
        quantity_change=quantity_change,
        stock_before=stock_before,
        stock_after=stock_after,
        notes=notes,
        created_by=created_by,
    )


def _compute_pl_data_core(retail_qs, wholesale_qs, returns_qs,
                           retail_revenue, retail_tax, retail_discounts,
                           wholesale_revenue, wholesale_tax, wholesale_discounts,
                           start_dt, end_dt):
    """Shared P&L calculation core – computes return COGS, COGS, expenses, profit."""
    # Return revenue and COGS reversal
    retail_return_revenue = 0
    wholesale_return_revenue = 0
    retail_return_cogs = Decimal("0")
    wholesale_return_cogs = Decimal("0")

    retail_orig_ids = set()
    ws_orig_ids = set()
    return_items_by_type = {'retail': [], 'wholesale': []}
    for ret in returns_qs.prefetch_related("items").iterator(chunk_size=1000):
        for item in ret.items.all():
            item_subtotal = float(item.subtotal)
            if ret.invoice_type == 'retail':
                retail_return_revenue += item_subtotal
                if item.original_item_id:
                    retail_orig_ids.add(item.original_item_id)
                return_items_by_type['retail'].append(item)
            else:
                wholesale_return_revenue += item_subtotal
                if item.original_item_id:
                    ws_orig_ids.add(item.original_item_id)
                return_items_by_type['wholesale'].append(item)

    retail_orig_items_map = {o.id: o for o in InvoiceItem.objects.filter(pk__in=retail_orig_ids)} if retail_orig_ids else {}
    ws_orig_items_map = {o.id: o for o in WholesaleInvoiceItem.objects.filter(pk__in=ws_orig_ids)} if ws_orig_ids else {}

    retail_orig_cost_data = SaleItemCost.objects.filter(invoice_item_id__in=retail_orig_ids).values('invoice_item_id').annotate(
        item_total=Sum(F('quantity') * F('unit_cost'))
    ) if retail_orig_ids else []
    retail_orig_cost_map = {c['invoice_item_id']: c['item_total'] for c in retail_orig_cost_data}
    ws_orig_cost_data = SaleItemCost.objects.filter(wholesale_invoice_item_id__in=ws_orig_ids).values('wholesale_invoice_item_id').annotate(
        item_total=Sum(F('quantity') * F('unit_cost'))
    ) if ws_orig_ids else []
    ws_orig_cost_map = {c['wholesale_invoice_item_id']: c['item_total'] for c in ws_orig_cost_data}

    for item in return_items_by_type['retail']:
        orig_item = retail_orig_items_map.get(item.original_item_id)
        unit_cost = Decimal('0')
        if orig_item:
            cost_val = retail_orig_cost_map.get(orig_item.id)
            if cost_val is not None:
                unit_cost = cost_val / orig_item.quantity if orig_item.quantity > 0 else Decimal('0')
            elif orig_item.unit_cost_at_sale and float(orig_item.unit_cost_at_sale) > 0:
                unit_cost = Decimal(str(orig_item.unit_cost_at_sale))
        if unit_cost == 0:
            food = Food.objects.filter(name=item.product_name).first()
            unit_cost = Decimal(str(food.default_purchase_cost or 0)) if food else Decimal('0')
        if unit_cost > 0:
            retail_return_cogs += unit_cost * item.quantity

    for item in return_items_by_type['wholesale']:
        orig_item = ws_orig_items_map.get(item.original_item_id)
        unit_cost = Decimal('0')
        if orig_item:
            cost_val = ws_orig_cost_map.get(orig_item.id)
            if cost_val is not None:
                unit_cost = cost_val / orig_item.quantity if orig_item.quantity > 0 else Decimal('0')
            elif orig_item.unit_cost_at_sale and float(orig_item.unit_cost_at_sale) > 0:
                unit_cost = Decimal(str(orig_item.unit_cost_at_sale))
        if unit_cost == 0:
            food = Food.objects.filter(name=item.product_name).first()
            unit_cost = Decimal(str(food.default_purchase_cost or 0)) if food else Decimal('0')
        if unit_cost > 0:
            wholesale_return_cogs += unit_cost * item.quantity

    return_revenue = retail_return_revenue + wholesale_return_revenue
    total_return_cogs = float(retail_return_cogs + wholesale_return_cogs)

    expense_qs = BusinessExpense.objects.filter(is_deleted=False)
    if start_dt and end_dt:
        expense_qs = expense_qs.filter(expense_date__range=[start_dt, end_dt])
    business_expenses = float(expense_qs.aggregate(total=Sum("amount"))["total"] or 0)

    total_revenue_sum = retail_revenue + wholesale_revenue
    total_tax_sum = retail_tax + wholesale_tax
    total_discounts_sum = retail_discounts + wholesale_discounts

    # COGS – retail
    retail_cogs = Decimal("0")
    retail_item_ids = InvoiceItem.objects.filter(invoice__in=retail_qs).values_list('id', flat=True)
    retail_costs = SaleItemCost.objects.filter(invoice_item_id__in=retail_item_ids).values('invoice_item_id').annotate(
        item_total=Sum(F('quantity') * F('unit_cost'))
    )
    retail_cost_map = {c['invoice_item_id']: c['item_total'] for c in retail_costs}
    for inv in retail_qs.prefetch_related("items").iterator(chunk_size=1000):
        for item in inv.items.all():
            cost_val = retail_cost_map.get(item.id)
            if cost_val is not None:
                retail_cogs += Decimal(str(cost_val))
            elif item.unit_cost_at_sale and float(item.unit_cost_at_sale) > 0:
                retail_cogs += Decimal(str(item.unit_cost_at_sale)) * item.quantity
            else:
                food = Food.objects.filter(name=item.product_name).first()
                unit_cost = float(food.default_purchase_cost or 0) if food else 0
                if unit_cost > 0:
                    retail_cogs += Decimal(str(unit_cost)) * item.quantity

    # COGS – wholesale
    wholesale_cogs = Decimal("0")
    ws_item_ids = WholesaleInvoiceItem.objects.filter(wholesale_invoice__in=wholesale_qs).values_list('id', flat=True)
    ws_costs = SaleItemCost.objects.filter(wholesale_invoice_item_id__in=ws_item_ids).values('wholesale_invoice_item_id').annotate(
        item_total=Sum(F('quantity') * F('unit_cost'))
    )
    ws_cost_map = {c['wholesale_invoice_item_id']: c['item_total'] for c in ws_costs}
    for inv in wholesale_qs.prefetch_related("items").iterator(chunk_size=1000):
        for item in inv.items.all():
            cost_val = ws_cost_map.get(item.id)
            if cost_val is not None:
                wholesale_cogs += Decimal(str(cost_val))
            elif item.unit_cost_at_sale and float(item.unit_cost_at_sale) > 0:
                wholesale_cogs += Decimal(str(item.unit_cost_at_sale)) * item.quantity
            else:
                food = Food.objects.filter(name=item.product_name).first()
                unit_cost = float(food.default_purchase_cost or 0) if food else 0
                if unit_cost > 0:
                    wholesale_cogs += Decimal(str(unit_cost)) * item.quantity

    # Apply return reversals
    retail_cogs_net = float(retail_cogs - retail_return_cogs)
    wholesale_cogs_net = float(wholesale_cogs - wholesale_return_cogs)
    total_cogs_net = retail_cogs_net + wholesale_cogs_net
    retail_revenue_net = retail_revenue - retail_return_revenue
    wholesale_revenue_net = wholesale_revenue - wholesale_return_revenue
    total_revenue_net = (retail_revenue_net + wholesale_revenue_net) - total_discounts_sum

    gross_profit = total_revenue_net - total_cogs_net
    net_profit = gross_profit - business_expenses

    return {
        'retail_return_revenue': retail_return_revenue,
        'wholesale_return_revenue': wholesale_return_revenue,
        'retail_return_cogs': retail_return_cogs,
        'wholesale_return_cogs': wholesale_return_cogs,
        'return_revenue': return_revenue,
        'total_return_cogs': total_return_cogs,
        'business_expenses': business_expenses,
        'total_revenue': total_revenue_sum,
        'total_tax': total_tax_sum,
        'total_discounts': total_discounts_sum,
        'retail_cogs': float(retail_cogs),
        'wholesale_cogs': float(wholesale_cogs),
        'retail_cogs_net': retail_cogs_net,
        'wholesale_cogs_net': wholesale_cogs_net,
        'total_cogs_net': total_cogs_net,
        'retail_revenue_net': retail_revenue_net,
        'wholesale_revenue_net': wholesale_revenue_net,
        'total_revenue_net': total_revenue_net,
        'gross_profit': gross_profit,
        'net_profit': net_profit,
    }


# =========================
# OPERATOR CART (Session-based, AJAX)
# =========================
CART_SESSION_KEY = "operator_cart"


def _get_op_cart(request):
    return request.session.get(CART_SESSION_KEY, [])


def _save_op_cart(request, cart):
    request.session[CART_SESSION_KEY] = cart
    request.session.modified = True


def _compute_cart_totals(cart):
    subtotal = 0
    total_tax = 0
    for item in cart:
        qty = int(item.get("quantity", 1))
        price = float(item.get("price", 0))
        disc = float(item.get("discount", 0))
        tax_rate = float(item.get("tax_rate", 0))
        line_total = qty * price - disc
        if line_total < 0:
            line_total = 0
        item_tax = round(line_total * tax_rate / 100, 2)
        item["line_total"] = round(line_total, 2)
        item["tax_rate"] = tax_rate
        item["tax_amount"] = item_tax
        item["line_total_with_tax"] = round(line_total + item_tax, 2)
        subtotal += line_total
        total_tax += item_tax
    return round(subtotal, 2), round(total_tax, 2)


def _get_effective_price(food):
    """Return (effective_price, discount_label) after applying product or category discount."""
    price = float(food.price)
    disc_label = ""
    if food.discount_type == "percentage" and food.discount_value > 0:
        price = round(price - (price * float(food.discount_value) / 100), 2)
        disc_label = f"{float(food.discount_value):.0f}% off"
    elif food.discount_type == "fixed" and food.discount_value > 0:
        price = round(price - float(food.discount_value), 2)
        if price < 0:
            price = 0
        disc_label = f"{_cs()}{float(food.discount_value):.0f} off"
    else:
        cat = food.category
        if cat.discount_type == "percentage" and cat.discount_value > 0:
            price = round(price - (price * float(cat.discount_value) / 100), 2)
            disc_label = f"Cat {float(cat.discount_value):.0f}% off"
        elif cat.discount_type == "fixed" and cat.discount_value > 0:
            price = round(price - float(cat.discount_value), 2)
            if price < 0:
                price = 0
            disc_label = f"Cat {_cs()}{float(cat.discount_value):.0f} off"
    return price, disc_label


@login_required
@module_access_required('pos')
def operator_cart_data(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    cart = _get_op_cart(request)
    subtotal, total_tax = _compute_cart_totals(cart)
    inv_disc = float(request.session.get("operator_invoice_discount", 0))
    after_inv_disc = round(subtotal - inv_disc, 2)
    if after_inv_disc < 0:
        after_inv_disc = 0
    grand_total = round(after_inv_disc + total_tax, 2)
    return JsonResponse({
        "cart": cart,
        "subtotal": subtotal,
        "invoice_discount": inv_disc,
        "total_tax": total_tax,
        "grand_total": grand_total,
    })


@login_required
@module_access_required('pos')
@require_POST
def operator_cart_add(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        data = json.loads(request.body)
        food_id = data.get("food_id")
        qty = int(data.get("quantity", 1))
        food = get_object_or_404(Food, pk=food_id)
        if food.stock <= 0:
            return JsonResponse({"error": "Out of stock"}, status=400)
        if qty > food.stock:
            return JsonResponse({"error": f"Insufficient stock. Only {food.stock} available."}, status=400)
        effective_price, disc_label = _get_effective_price(food)
        cart = _get_op_cart(request)
        existing = None
        for item in cart:
            if item["food_id"] == food_id:
                existing = item
                break
        if existing:
            new_qty = int(existing["quantity"]) + qty
            if new_qty > food.stock:
                return JsonResponse({"error": f"Insufficient stock. Only {food.stock} available."}, status=400)
            existing["quantity"] = new_qty
            existing["line_total"] = round(new_qty * float(existing["price"]), 2)
        else:
            cart.append({
                "food_id": food_id,
                "name": food.name,
                "price": effective_price,
                "quantity": qty,
                "discount": 0,
                "stock": food.stock,
                "image": food.image.url if food.image else "",
                "discount_label": disc_label,
                "tax_rate": 0,
            })
        _save_op_cart(request, cart)
        subtotal, _ = _compute_cart_totals(cart)
        return JsonResponse({"success": True, "cart": cart, "subtotal": subtotal})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@login_required
@module_access_required('pos')
@require_POST
def operator_cart_update(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        data = json.loads(request.body)
        food_id = data.get("food_id")
        qty = int(data.get("quantity", 0))
        tax_rate = data.get("tax_rate")
        food = get_object_or_404(Food, pk=food_id)
        cart = _get_op_cart(request)
        for item in cart:
            if item["food_id"] == food_id:
                if qty <= 0:
                    cart.remove(item)
                else:
                    if qty > food.stock:
                        return JsonResponse({"error": f"Insufficient stock. Only {food.stock} available."}, status=400)
                    item["quantity"] = qty
                    if tax_rate is not None:
                        item["tax_rate"] = float(tax_rate)
                break
        _save_op_cart(request, cart)
        subtotal, _ = _compute_cart_totals(cart)
        return JsonResponse({"success": True, "cart": cart, "subtotal": subtotal})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@login_required
@module_access_required('pos')
@require_POST
def operator_cart_remove(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        data = json.loads(request.body)
        food_id = data.get("food_id")
        cart = _get_op_cart(request)
        cart[:] = [item for item in cart if item["food_id"] != food_id]
        _save_op_cart(request, cart)
        subtotal, _ = _compute_cart_totals(cart)
        return JsonResponse({"success": True, "cart": cart, "subtotal": subtotal})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@login_required
@module_access_required('pos')
@require_POST
def operator_cart_clear(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    request.session.pop(CART_SESSION_KEY, None)
    request.session.pop("operator_invoice_discount", None)
    return JsonResponse({"success": True, "cart": []})


@login_required
@module_access_required('pos')
@require_POST
def operator_set_discount(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    if not has_permission(request.user, 'can_apply_discount'):
        return JsonResponse({"error": "Permission denied"}, status=403)
    try:
        data = json.loads(request.body)
        inv_disc = float(data.get("discount", 0))
        if inv_disc < 0:
            inv_disc = 0
        request.session["operator_invoice_discount"] = inv_disc
        return JsonResponse({"success": True, "invoice_discount": inv_disc})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@login_required
@module_access_required('pos')
@require_POST
def operator_hold_invoice(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        data = json.loads(request.body)
        cart = _get_op_cart(request)
        if not cart:
            return JsonResponse({"error": "Cart is empty"}, status=400)
        held = HeldCart.objects.create(
            operator=request.user,
            cart_data=cart,
            customer_name=data.get("customer_name", ""),
            customer_email=data.get("customer_email", ""),
            customer_phone=data.get("customer_phone", ""),
            invoice_discount=float(request.session.get("operator_invoice_discount", 0)),
            notes=data.get("notes", ""),
        )
        request.session.pop(CART_SESSION_KEY, None)
        request.session.pop("operator_invoice_discount", None)
        request.session.pop("operator_tax_percentage", None)
        return JsonResponse({"success": True, "held_id": held.id})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@login_required
@module_access_required('pos')
def operator_held_invoices(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    held = HeldCart.objects.filter(operator=request.user)
    return JsonResponse({
        "held": [{
            "id": h.id,
            "customer_name": h.customer_name or "",
            "customer_email": h.customer_email or "",
            "customer_phone": h.customer_phone or "",
            "items_count": len(h.cart_data),
            "invoice_discount": float(h.invoice_discount),
            "notes": h.notes or "",
            "created_at": h.created_at.strftime("%d-%m-%Y %I:%M %p"),
        } for h in held]
    })


@login_required
@module_access_required('pos')
@require_POST
def operator_resume_held(request, held_id):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    held = get_object_or_404(HeldCart, pk=held_id, operator=request.user)
    request.session[CART_SESSION_KEY] = held.cart_data
    request.session["operator_invoice_discount"] = float(held.invoice_discount)
    held.delete()
    return JsonResponse({"success": True, "cart": held.cart_data})


@login_required
@module_access_required('pos')
@require_POST
def operator_delete_held(request, held_id):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    held = get_object_or_404(HeldCart, pk=held_id, operator=request.user)
    held.delete()
    return JsonResponse({"success": True})


@login_required
@module_access_required('pos')
def operator_customer_search(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    q = request.GET.get("q", "").strip()
    if len(q) < 1:
        return JsonResponse({"results": []})
    users = User.objects.filter(name__icontains=q, is_staff=False, is_operator=False, is_superuser=False)
    results = []
    for u in users[:20]:
        card = LoyaltyCard.objects.filter(user=u, status='ACTIVE').first()
        results.append({
            "id": u.id,
            "name": u.name,
            "phone": u.phone or "",
            "has_card": card is not None,
            "card_status": card.status if card else "NONE",
            "points_available": card.remaining_points if card else 0,
        })
    return JsonResponse({"results": results})


@login_required
@module_access_required('pos')
def operator_loyalty_lookup(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    phone = request.GET.get("phone", "").strip()
    if not phone:
        return JsonResponse({"exists": False})
    user = User.objects.filter(phone=phone).first()
    if not user or user.is_staff or getattr(user, 'is_operator', False) or user.is_superuser:
        return JsonResponse({"exists": False})
    card = LoyaltyCard.objects.filter(user=user, status='ACTIVE').first()
    return JsonResponse({
        "exists": True,
        "name": user.name,
        "points_available": card.remaining_points if card else 0,
        "card_status": card.status if card else "NONE",
        "has_card": card is not None,
    })


@login_required
@module_access_required('pos')
@require_POST
def operator_checkout(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        data = json.loads(request.body)
        cart = _get_op_cart(request)
        if not cart:
            return JsonResponse({"error": "Cart is empty"}, status=400)

        customer_name = data.get("customer_name", "").strip()
        customer_phone = data.get("customer_phone", "").strip()
        payment_method = data.get("payment_method", "cash")
        loyalty_points_to_redeem = data.get("loyalty_points_to_redeem")
        if loyalty_points_to_redeem is not None:
            loyalty_points_to_redeem = int(loyalty_points_to_redeem)
        else:
            loyalty_points_to_redeem = 0
        cash_received = float(data.get("cash_received", 0))

        # Feature 2: Auto-create/find user for loyalty
        invoice_user = None
        if customer_name and customer_phone:
            existing_user = User.objects.filter(phone=customer_phone).first()
            if existing_user:
                if not existing_user.is_staff and not getattr(existing_user, 'is_operator', False) and not existing_user.is_superuser:
                    invoice_user = existing_user
            else:
                placeholder_email = f"cust_{customer_phone.replace('+', '').replace(' ', '')}@pos.local"
                invoice_user = User.objects.create_user(
                    email=placeholder_email,
                    name=customer_name,
                    phone=customer_phone,
                    password=''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(10)),
                )

        # Auto-create loyalty card for registered customer
        lcard = None
        if invoice_user:
            lcard, lcard_created = LoyaltyCard.objects.get_or_create(
                user=invoice_user,
                defaults={'status': 'ACTIVE'}
            )
            if lcard_created:
                try:
                    generate_qr_code_image(lcard, request)
                    generate_loyalty_card_pdf(lcard, request)
                    generate_loyalty_card_image(lcard, request)
                except Exception:
                    pass

        if not has_permission(request.user, 'can_create_invoice'):
            if loyalty_points_to_redeem > 0:
                if not invoice_user:
                    return JsonResponse({"error": "Customer name and phone required for loyalty redemption."}, status=400)
                lcard = LoyaltyCard.objects.filter(user=invoice_user, status='ACTIVE').first()
                if not lcard:
                    return JsonResponse({"error": "Customer does not have an active loyalty card."}, status=400)
                if lcard.remaining_points < loyalty_points_to_redeem:
                    return JsonResponse({
                        "error": f"Insufficient loyalty points. Available: {lcard.remaining_points}, Requested: {loyalty_points_to_redeem}"
                    }, status=400)
            subtotal, _ = _compute_cart_totals(cart)
            inv_disc = float(request.session.get("operator_invoice_discount", 0))
            PendingApproval.objects.create(
                operator=request.user,
                cart_data=cart,
                customer_name=customer_name,
                customer_email="",
                customer_phone=customer_phone,
                payment_method=payment_method,
                invoice_discount=inv_disc,
                tax_percentage=0,
                subtotal=subtotal,
                loyalty_points_to_redeem=loyalty_points_to_redeem,
                notes=data.get("notes", ""),
            )
            request.session.pop(CART_SESSION_KEY, None)
            request.session.pop("operator_invoice_discount", None)
            return JsonResponse({
                "success": True,
                "sent_for_approval": True,
                "message": "Invoice sent to admin for approval",
            })

        # Build the cart format expected by _create_order_from_cart
        checkout_cart = []
        for item in cart:
            food = Food.objects.get(pk=item["food_id"])
            qty = int(item["quantity"])
            if qty > food.stock:
                return JsonResponse({"error": f"Insufficient stock for {food.name}: only {food.stock} available"}, status=400)
            checkout_cart.append({
                "id": food.id,
                "name": item["name"],
                "price": float(item["price"]),
                "qty": qty,
                "image": item.get("image", ""),
            })

        subtotal, total_tax = _compute_cart_totals(cart)
        inv_disc = float(request.session.get("operator_invoice_discount", 0))
        after_disc = round(subtotal - inv_disc, 2)
        if after_disc < 0:
            after_disc = 0
        grand_total = round(after_disc + total_tax, 2)

        # Deduct stock and record batch consumption
        stock_movements = []
        svc = InventoryValuationService()
        for item in cart:
            food = Food.objects.get(pk=item["food_id"])
            qty = int(item["quantity"])
            if qty > food.stock:
                return JsonResponse({"error": f"Insufficient stock for {food.name}: only {food.stock} available"}, status=400)
            stock_before = food.stock
            food.stock -= qty
            if food.stock < 0:
                food.stock = 0
            food.save()
            stock_movements.append({
                "food": food,
                "qty": qty,
                "stock_before": stock_before,
                "stock_after": food.stock,
            })

            # Record batch costs for this item (stored after invoice creation)
            item['_batch_consumptions'] = svc.consume(food, qty)

        # Create invoice
        highest_tax_pct = max([float(item.get("tax_rate", 0)) for item in cart] or [0])
        invoice = Invoice.objects.create(
            user=invoice_user,
            created_by=request.user,
            customer_name=customer_name or "Walk-in Customer",
            customer_email="",
            customer_phone=customer_phone or "",
            invoice_number=f"INV-{uuid.uuid4().hex[:8].upper()}",
            payment_method=payment_method,
            tax_percentage=highest_tax_pct,
            tax_amount=total_tax,
            subtotal_amount=subtotal,
            total_amount=grand_total,
            qr_offer_discount_amount=inv_disc,
        )

        for item in cart:
            invoice_item = InvoiceItem.objects.create(
                invoice=invoice,
                product_name=item["name"],
                price=float(item["price"]),
                quantity=int(item["quantity"]),
                subtotal=float(item["line_total"]),
                tax_percentage=float(item.get("tax_rate", 0)),
                tax_amount=float(item.get("tax_amount", 0)),
            )
            # Record SaleItemCost for each batch consumed
            total_cogs = Decimal('0')
            for batch, take, unit_cost in item.get('_batch_consumptions', []):
                SaleItemCost.objects.create(
                    inventory_batch=batch,
                    invoice_item=invoice_item,
                    quantity=take,
                    unit_cost=unit_cost,
                )
                total_cogs += Decimal(str(take)) * unit_cost
            # Update unit_cost_at_sale on the invoice item
            if item.get('_batch_consumptions'):
                qty = int(item['quantity'])
                invoice_item.unit_cost_at_sale = float(total_cogs / Decimal(str(qty)))
                invoice_item.save(update_fields=['unit_cost_at_sale'])

        invoice.generate_qr_code(request)
        invoice.save()

        # Earn loyalty points for the purchase (only reward_points, never price)
        if invoice_user:
            earn_lcard = LoyaltyCard.objects.filter(user=invoice_user, status='ACTIVE').first()
            if earn_lcard:
                points_earned = _calculate_reward_points(cart)
                if points_earned > 0:
                    earn_lcard.add_points(points_earned, order_number=invoice.invoice_number)
                    invoice.loyalty_points_earned = points_earned
                    invoice.save(update_fields=['loyalty_points_earned'])

        # Feature 3: Redeem loyalty points
        if invoice_user and loyalty_points_to_redeem > 0:
            lcard = LoyaltyCard.objects.filter(user=invoice_user, status='ACTIVE').first()
            if not lcard:
                return JsonResponse({"error": "Customer does not have an active loyalty card."}, status=400)
            if lcard.remaining_points < loyalty_points_to_redeem:
                return JsonResponse({
                    "error": f"Customer does not have enough loyalty points. Available: {lcard.remaining_points}, Requested: {loyalty_points_to_redeem}"
                }, status=400)
            try:
                lcard.redeem_points(loyalty_points_to_redeem, order_number=invoice.invoice_number)
                invoice.is_loyalty_payment = True
                invoice.loyalty_points_used = loyalty_points_to_redeem
                invoice.total_amount = max(0, grand_total - loyalty_points_to_redeem)
                invoice.save(update_fields=['is_loyalty_payment', 'loyalty_points_used', 'total_amount'])
            except Exception:
                pass

        # Save cash received and change due
        invoice.refresh_from_db()
        invoice.cash_received = cash_received
        invoice.change_due = max(0, cash_received - float(invoice.total_amount))
        invoice.save(update_fields=['cash_received', 'change_due'])

        # Record stock movements with invoice reference
        for sm in stock_movements:
            _record_stock_movement(
                food=sm["food"],
                transaction_type='retail_sale',
                quantity_change=-sm["qty"],
                stock_before=sm["stock_before"],
                stock_after=sm["stock_after"],
                reference_number=invoice.invoice_number,
                created_by=invoice_user or request.user if request.user.is_authenticated else None,
            )

        # Clear session cart
        request.session.pop(CART_SESSION_KEY, None)
        request.session.pop("operator_invoice_discount", None)

        return JsonResponse({
            "success": True,
            "invoice_no": invoice.invoice_number,
            "uuid_token": invoice.uuid_token,
        })
    except Exception as e:
        traceback.print_exc()
        return JsonResponse({"error": str(e)}, status=400)


# =========================
# BARCODE PRINTING
# =========================
@login_required
@module_access_required('products')
def barcode_print(request, pk):
    food = get_object_or_404(Food, pk=pk)
    return render(request, "operator/barcode_label.html", {"product": food})


@staff_member_required
@module_access_required('products')
def barcode_print_multiple(request):
    ids = request.GET.get("ids", "")
    if ids:
        id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        products = Food.objects.filter(pk__in=id_list)
    else:
        products = Food.objects.filter(barcode__isnull=False)[:50]
    return render(request, "operator/barcode_label.html", {"products": products, "multiple": True})


# =========================
# STOCK POSITION
# =========================
@login_required
@module_access_required('inventory')
def stock_position(request):
    if not has_permission(request.user, 'can_stock_position'):
        if request.user.is_operator:
            return redirect('users:operator_dashboard')
        return redirect('users:login')
    from django.utils.timezone import make_aware, now as tz_now
    from datetime import datetime, timedelta

    from .models import StockMovement
    from django.db.models import Sum as DbSum

    foods = Food.objects.all().order_by("name")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    has_date_filter = bool(start_date and end_date)
    if has_date_filter:
        try:
            start_dt = make_aware(datetime.strptime(start_date, "%Y-%m-%d"))
            end_dt = make_aware(datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1))
        except (ValueError, TypeError):
            has_date_filter = False

    stock_data = []
    svc = InventoryValuationService()
    for food in foods:
        closing = food.stock
        inventory_value = float(svc.get_product_inventory_value(food))

        last_sale = food.stock_movements.filter(transaction_type__in=['retail_sale', 'wholesale_sale']).first()
        last_update = food.stock_movements.exclude(transaction_type__in=['retail_sale', 'wholesale_sale']).first()

        movements = food.stock_movements
        if has_date_filter:
            movements = movements.filter(created_at__range=[start_dt, end_dt])

        # Purchased = opening_stock + stock_purchase + manual stock additions (positive adjustments)
        purchased_opening = movements.filter(transaction_type='opening_stock').aggregate(total=DbSum("quantity_change"))["total"] or 0
        purchased_stock_purchase = movements.filter(transaction_type='stock_purchase').aggregate(total=DbSum("quantity_change"))["total"] or 0
        purchased_adjustments = movements.filter(transaction_type='stock_adjustment', quantity_change__gt=0).aggregate(total=DbSum("quantity_change"))["total"] or 0
        purchased = purchased_opening + purchased_stock_purchase + purchased_adjustments

        # Gross sales (absolute, stored as negative in StockMovement)
        gross_sold = movements.filter(transaction_type__in=['retail_sale', 'wholesale_sale']).aggregate(total=DbSum("quantity_change"))["total"] or 0
        gross_sold = abs(gross_sold)

        # Returns (stored as positive in StockMovement)
        returned = movements.filter(transaction_type__in=['retail_return', 'wholesale_return']).aggregate(total=DbSum("quantity_change"))["total"] or 0

        # Net sold = gross sold - returns
        sold_net = max(0, gross_sold - returned)

        # Opening stock at beginning of period
        if has_date_filter:
            before_movements = food.stock_movements.filter(created_at__lt=start_dt)
            last_before = before_movements.first()
            opening = last_before.stock_after if last_before else 0
        else:
            opening = 0

        # Calculate batch-weighted average cost per unit
        batches = InventoryBatch.objects.filter(food=food, remaining_quantity__gt=0)
        total_batch_qty = sum(b.remaining_quantity for b in batches)
        if total_batch_qty > 0:
            avg_batch_cost = float(sum(Decimal(str(b.remaining_quantity)) * b.unit_cost for b in batches)) / total_batch_qty
        else:
            avg_batch_cost = float(food.default_purchase_cost or 0)

        stock_data.append({
            "name": food.name,
            "opening": opening,
            "purchased": purchased,
            "sold": sold_net,
            "returned": returned,
            "closing": closing,
            "cost_price": round(avg_batch_cost, 2),
            "inventory_value": round(inventory_value, 2),
            "last_sale_date": last_sale.created_at.isoformat() if last_sale else '',
            "last_update_date": last_update.created_at.isoformat() if last_update else '',
            "food_id": food.id,
        })

    return render(request, "admin/stock_position.html", {
        "stock_data": stock_data,
        "start_date": start_date,
        "end_date": end_date,
    })


# =========================
# STOCK HISTORY
# =========================
@login_required
@module_access_required('inventory')
def stock_history(request):
    if not has_permission(request.user, 'can_manage_inventory'):
        if request.user.is_operator:
            return redirect('users:operator_dashboard')
        return redirect('users:login')

    from .models import StockMovement

    product_id = request.GET.get("product", "")
    transaction_type = request.GET.get("type", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    user_id = request.GET.get("user", "")

    movements = StockMovement.objects.all().select_related('food', 'created_by')

    if product_id:
        movements = movements.filter(food_id=product_id)
    if transaction_type:
        movements = movements.filter(transaction_type=transaction_type)
    if start_date:
        from django.utils.timezone import make_aware
        try:
            sd = make_aware(datetime.strptime(start_date, "%Y-%m-%d"))
            movements = movements.filter(created_at__gte=sd)
        except (ValueError, TypeError):
            pass
    if end_date:
        from django.utils.timezone import make_aware
        try:
            ed = make_aware(datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1))
            movements = movements.filter(created_at__lt=ed)
        except (ValueError, TypeError):
            pass
    if user_id:
        movements = movements.filter(created_by_id=user_id)

    products = Food.objects.all().order_by("name")
    users = User.objects.filter(is_staff=True).order_by("email")

    return render(request, "admin/stock_history.html", {
        "movements": movements,
        "products": products,
        "users": users,
        "selected_product": product_id,
        "selected_type": transaction_type,
        "selected_user": user_id,
        "start_date": start_date,
        "end_date": end_date,
    })


# =========================
# STOCK MOVEMENT REPORT
# =========================
@login_required
@module_access_required('inventory')
def stock_movement_report(request):
    if not has_permission(request.user, 'can_stock_position'):
        if request.user.is_operator:
            return redirect('users:operator_dashboard')
        return redirect('users:login')

    from .models import StockMovement
    from django.utils.timezone import make_aware, now as tz_now
    from datetime import datetime, timedelta

    start_str = request.GET.get("start_date", "")
    end_str = request.GET.get("end_date", "")
    export_format = request.GET.get("format", "")

    if not start_str or not end_str:
        today = tz_now()
        start_str = today.strftime("%Y-%m-%d")
        end_str = today.strftime("%Y-%m-%d")

    try:
        start_dt = make_aware(datetime.strptime(start_str, "%Y-%m-%d"))
        end_dt = make_aware(datetime.strptime(end_str, "%Y-%m-%d") + timedelta(days=1))
    except (ValueError, TypeError):
        start_dt = make_aware(datetime.now().replace(hour=0, minute=0, second=0))
        end_dt = make_aware(datetime.now().replace(hour=23, minute=59, second=59))

    products = Food.objects.all().order_by("name")
    report_data = []
    svc = InventoryValuationService()

    for food in products:
        movements = food.stock_movements.filter(created_at__range=[start_dt, end_dt])
        stock_added = sum(m.quantity_change for m in movements if m.quantity_change > 0 and m.transaction_type in ['opening_stock', 'stock_purchase', 'stock_adjustment', 'retail_return', 'wholesale_return'])
        stock_sold = abs(sum(m.quantity_change for m in movements if m.quantity_change < 0 and m.transaction_type in ['retail_sale', 'wholesale_sale']))
        stock_returned = sum(m.quantity_change for m in movements if m.quantity_change > 0 and m.transaction_type in ['retail_return', 'wholesale_return'])
        stock_adjusted = sum(m.quantity_change for m in movements if m.transaction_type == 'stock_adjustment')

        opening_movement = food.stock_movements.filter(created_at__lt=start_dt).first()
        opening_stock = opening_movement.stock_after if opening_movement else max(0, food.stock - stock_added + stock_sold)

        closing_stock = food.stock
        inventory_value = float(svc.get_product_inventory_value(food))

        report_data.append({
            "food": food,
            "opening_stock": max(0, closing_stock - stock_added + stock_sold),
            "stock_added": stock_added,
            "stock_sold": stock_sold,
            "stock_returned": stock_returned,
            "stock_adjusted": stock_adjusted,
            "closing_stock": closing_stock,
            "inventory_value": round(inventory_value, 2),
        })

    if export_format == 'pdf':
        return _export_stock_movement_pdf(start_str, end_str, report_data)
    elif export_format == 'excel':
        return _export_stock_movement_excel(start_str, end_str, report_data)

    return render(request, "admin/stock_movement_report.html", {
        "report_data": report_data,
        "start_date": start_str,
        "end_date": end_str,
    })


def _export_stock_movement_pdf(start_str, end_str, report_data):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Stock Movement Report", styles["Title"]))
    elements.append(Paragraph(f"Period: {start_str} to {end_str}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    data = [["Product", "Opening", "Added", "Sold", "Returned", "Adjusted", "Closing", "Inventory Value"]]
    for rd in report_data:
        data.append([
            rd["food"].name,
            str(rd["opening_stock"]),
            str(rd["stock_added"]),
            str(rd["stock_sold"]),
            str(rd["stock_returned"]),
            str(rd["stock_adjusted"]),
            str(rd["closing_stock"]),
            f"{_cs()}{rd['inventory_value']:,.2f}",
        ])

    table = Table(data, colWidths=[120, 60, 50, 50, 60, 60, 60, 80])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f"stock_movement_{start_str}_{end_str}.pdf")


def _export_stock_movement_excel(start_str, end_str, report_data):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Movement"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Stock Movement Report ({start_str} to {end_str})"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Product", "Opening", "Added", "Sold", "Returned", "Adjusted", "Closing", "Inventory Value"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, rd in enumerate(report_data, 4):
        ws.cell(row=i, column=1, value=rd["food"].name)
        ws.cell(row=i, column=2, value=rd["opening_stock"])
        ws.cell(row=i, column=3, value=rd["stock_added"])
        ws.cell(row=i, column=4, value=rd["stock_sold"])
        ws.cell(row=i, column=5, value=rd["stock_returned"])
        ws.cell(row=i, column=6, value=rd["stock_adjusted"])
        ws.cell(row=i, column=7, value=rd["closing_stock"])
        ws.cell(row=i, column=8, value=round(rd["inventory_value"], 2))

    ws.column_dimensions["A"].width = 25
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f"stock_movement_{start_str}_{end_str}.xlsx")


# =========================
# PENDING APPROVALS
# =========================
def _normalize_cart_data(cart_data):
    if isinstance(cart_data, dict):
        items = cart_data.get("_items", [])
    elif isinstance(cart_data, str):
        try:
            items = json.loads(cart_data)
        except (json.JSONDecodeError, TypeError):
            items = []
    elif isinstance(cart_data, list):
        items = cart_data
    else:
        items = []
    result = []
    for item in items:
        if isinstance(item, dict):
            result.append({
                "name": item.get("name", ""),
                "price": item.get("price", 0),
                "quantity": item.get("quantity", 1),
            })
        else:
            result.append({"name": str(item), "price": 0, "quantity": 1})
    return result


@staff_member_required
@module_access_required('purchases')
def admin_pending_approvals(request):
    status_filter = request.GET.get("status", "")
    qs = PendingApproval.objects.select_related('operator', 'invoice').all()
    if status_filter == "approved":
        qs = qs.filter(is_approved=True)
    elif status_filter == "rejected":
        qs = qs.filter(is_rejected=True)
    elif status_filter == "pending":
        qs = qs.filter(is_approved=False, is_rejected=False)
    for pa in qs:
        pa.normalized_cart = _normalize_cart_data(pa.cart_data)
    return render(request, "admin/pending_approvals.html", {
        "pending_list": qs,
    })


@staff_member_required
@module_access_required('purchases')
@require_POST
def admin_approve_request(request, pk):
    pending = get_object_or_404(PendingApproval, pk=pk, is_approved=False, is_rejected=False)
    action = request.POST.get("action", "approve")
    admin_notes = request.POST.get("admin_notes", "")

    if action == "reject":
        pending.is_rejected = True
        pending.admin_notes = admin_notes
        pending.save()
        messages.success(request, f"Request #{pending.id} has been rejected.")
        return redirect("users:admin_pending_approvals")

    cart = pending.cart_data
    if isinstance(cart, dict):
        cart = cart.get("_items", [])
    if not cart:
        messages.error(request, "Cart data is empty.")
        return redirect("users:admin_pending_approvals")

    try:
        with transaction.atomic():
            subtotal = 0
            total_tax = 0
            for item in cart:
                qty = int(item.get("quantity", 1))
                price = float(item.get("price", 0))
                tax_rate = float(item.get("tax_rate", 0))
                line_total = qty * price
                item_tax = round(line_total * tax_rate / 100, 2)
                item["line_total"] = round(line_total, 2)
                item["tax_rate"] = tax_rate
                item["tax_amount"] = item_tax
                subtotal += line_total
                total_tax += item_tax
            subtotal = round(subtotal, 2)
            total_tax = round(total_tax, 2)

            inv_disc = float(pending.invoice_discount)
            after_disc = round(subtotal - inv_disc, 2)
            if after_disc < 0:
                after_disc = 0
            grand_total = round(after_disc + total_tax, 2)

            invoice_user = None
            customer_phone = getattr(pending, 'customer_phone', '')
            if pending.customer_name and (pending.customer_email or customer_phone):
                lookup_val = customer_phone or pending.customer_email
                existing_user = User.objects.filter(phone=lookup_val).first() if customer_phone else User.objects.filter(email=pending.customer_email).first()
                if existing_user:
                    if not existing_user.is_staff and not getattr(existing_user, 'is_operator', False) and not existing_user.is_superuser:
                        invoice_user = existing_user
                else:
                    placeholder_email = pending.customer_email or f"cust_{customer_phone.replace('+', '').replace(' ', '')}@pos.local"
                    invoice_user = User.objects.create_user(
                        email=placeholder_email,
                        name=pending.customer_name,
                        phone=customer_phone,
                        password=''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(10)),
                    )

            # Validate loyalty BEFORE any state change
            pts_redeem = getattr(pending, 'loyalty_points_to_redeem', 0)
            rcard = None
            if invoice_user and pts_redeem > 0:
                rcard = LoyaltyCard.objects.filter(user=invoice_user, status='ACTIVE').first()
                if not rcard:
                    messages.error(request, "Customer does not have an active loyalty card.")
                    return redirect("users:admin_pending_approvals")
                if rcard.remaining_points < pts_redeem:
                    messages.error(
                        request,
                        f"Insufficient loyalty points. Available: {rcard.remaining_points}, Requested: {pts_redeem}"
                    )
                    return redirect("users:admin_pending_approvals")

            lcard = None
            if invoice_user:
                lcard, lcard_created = LoyaltyCard.objects.get_or_create(
                    user=invoice_user,
                    defaults={'status': 'ACTIVE'}
                )
                if lcard_created:
                    try:
                        generate_qr_code_image(lcard, request)
                        generate_loyalty_card_pdf(lcard, request)
                        generate_loyalty_card_image(lcard, request)
                    except Exception:
                        pass

            highest_tax_pct = max([float(item.get("tax_rate", 0)) for item in cart] or [0])
            invoice = Invoice.objects.create(
                user=invoice_user,
                created_by=request.user,
                customer_name=pending.customer_name or "Walk-in Customer",
                customer_email=pending.customer_email or "",
                customer_phone=customer_phone or "",
                invoice_number=f"INV-{uuid.uuid4().hex[:8].upper()}",
                payment_method=pending.payment_method,
                tax_percentage=highest_tax_pct,
                tax_amount=total_tax,
                subtotal_amount=subtotal,
                total_amount=grand_total,
                qr_offer_discount_amount=inv_disc,
            )

            for item in cart:
                food = Food.objects.filter(pk=item["food_id"]).first()
                invoice_item = InvoiceItem.objects.create(
                    invoice=invoice,
                    product_name=item["name"],
                    price=float(item["price"]),
                    quantity=int(item["quantity"]),
                    subtotal=float(item["line_total"]),
                    tax_percentage=float(item.get("tax_rate", 0)),
                    tax_amount=float(item.get("tax_amount", 0)),
                )

            invoice.generate_qr_code(request)
            invoice.save()

            # Earn loyalty points
            if invoice_user:
                earn_lcard = LoyaltyCard.objects.filter(user=invoice_user, status='ACTIVE').first()
                if earn_lcard:
                    points_earned = _calculate_reward_points(cart)
                    if points_earned > 0:
                        earn_lcard.add_points(points_earned, order_number=invoice.invoice_number)
                        invoice.loyalty_points_earned = points_earned
                        invoice.save(update_fields=['loyalty_points_earned'])

            # Redeem loyalty points (already validated above)
            if invoice_user and pts_redeem > 0 and rcard:
                rcard.redeem_points(pts_redeem, order_number=invoice.invoice_number)
                invoice.is_loyalty_payment = True
                invoice.loyalty_points_used = pts_redeem
                invoice.total_amount = max(0, grand_total - pts_redeem)
                invoice.save(update_fields=['is_loyalty_payment', 'loyalty_points_used', 'total_amount'])

            # Stock deduction and SaleItemCost creation
            svc = InventoryValuationService()
            for item in cart:
                food_id = item.get("food_id")
                if not food_id:
                    continue
                try:
                    food = Food.objects.get(pk=food_id)
                    qty = int(item["quantity"])
                    food.stock -= qty
                    if food.stock < 0:
                        food.stock = 0
                    food.save()
                    invoice_item = InvoiceItem.objects.filter(invoice=invoice, product_name=item["name"]).last()
                    total_cogs = Decimal('0')
                    for batch, take, unit_cost in svc.consume(food, qty):
                        SaleItemCost.objects.create(
                            inventory_batch=batch,
                            invoice_item=invoice_item,
                            quantity=take,
                            unit_cost=unit_cost,
                        )
                        total_cogs += Decimal(str(take)) * unit_cost
                    if invoice_item:
                        invoice_item.unit_cost_at_sale = float(total_cogs / Decimal(str(qty)))
                        invoice_item.save(update_fields=['unit_cost_at_sale'])
                except Food.DoesNotExist:
                    pass

            pending.invoice = invoice
            pending.is_approved = True
            pending.admin_notes = admin_notes
            pending.save()

        messages.success(request, f"Request #{pending.id} approved. Invoice {invoice.invoice_number} created.")
        return redirect("users:admin_pending_approvals")
    except Exception as e:
        traceback.print_exc()
        messages.error(request, f"Error approving request: {str(e)}")
        return redirect("users:admin_pending_approvals")


# =========================
# RETURN INVOICE
# =========================
@login_required
@module_access_required('returns')
def return_invoice_page(request):
    if not has_permission(request.user, 'can_returns'):
        messages.error(request, "You do not have permission to process returns.")
        if request.user.is_operator:
            return redirect('users:operator_dashboard')
        return redirect('users:admin_dashboard')
    categories = Category.objects.filter(is_active=True)
    return render(request, "admin/return_invoice.html", {
        'active_page': 'return_invoice',
        'categories': categories,
    })


@csrf_exempt
@module_access_required('returns')
def return_invoice_search(request):
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Authentication required"}, status=403)
    if not has_permission(request.user, 'can_returns'):
        return JsonResponse({"error": "Permission denied"}, status=403)
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=400)
    try:
        data = json.loads(request.body)
        q = data.get("search", "").strip()
        if not q:
            return JsonResponse({"invoices": []})
        # Retail invoices
        retail_qs = Invoice.objects.filter(
            Q(invoice_number__icontains=q) |
            Q(customer_name__icontains=q) |
            Q(customer_phone__icontains=q) |
            Q(user__phone__icontains=q) |
            Q(user__name__icontains=q)
        ).order_by('-created_at')[:20]
        # Wholesale invoices
        wholesale_qs = WholesaleInvoice.objects.filter(
            Q(invoice_number__icontains=q) |
            Q(wholesale_customer__company_name__icontains=q) |
            Q(wholesale_customer__email__icontains=q)
        ).order_by('-created_at')[:20]
        results = []
        for inv in retail_qs:
            results.append({
                "id": inv.id,
                "invoice_number": inv.invoice_number,
                "customer_name": inv.customer_name or (inv.user.name if inv.user else "Walk-in Customer"),
                "customer_phone": inv.customer_phone or (inv.user.phone if inv.user else ""),
                "total": float(inv.total_amount),
                "date": inv.created_at.strftime("%d-%m-%Y %I:%M %p"),
                "type": "retail",
            })
        winv_ids = set()
        for inv in wholesale_qs:
            if inv.id in winv_ids:
                continue
            winv_ids.add(inv.id)
            results.append({
                "id": inv.id,
                "invoice_number": inv.invoice_number,
                "customer_name": inv.wholesale_customer.company_name if inv.wholesale_customer else "Wholesale Customer",
                "customer_email": inv.wholesale_customer.email if inv.wholesale_customer else "",
                "total": float(inv.total_amount),
                "date": inv.created_at.strftime("%d-%m-%Y %I:%M %p"),
                "type": "wholesale",
            })
        # Sort by date descending
        results.sort(key=lambda x: x["date"], reverse=True)
        return JsonResponse({"invoices": results[:20]})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@module_access_required('returns')
def return_invoice_get_items(request, invoice_id):
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Authentication required"}, status=403)
    if not has_permission(request.user, 'can_returns'):
        return JsonResponse({"error": "Permission denied"}, status=403)
    try:
        invoice_type = request.GET.get("invoice_type", "retail")
        if invoice_type == "wholesale":
            invoice = get_object_or_404(WholesaleInvoice, pk=invoice_id)
            items = WholesaleInvoiceItem.objects.filter(wholesale_invoice=invoice)
            total_returned_qty = {}
            for ret in ReturnInvoice.objects.filter(wholesale_original_invoice=invoice):
                for ret_item in ret.items.all():
                    oid = ret_item.original_item_id
                    total_returned_qty[oid] = total_returned_qty.get(oid, 0) + ret_item.quantity

            items_data = []
            for item in items:
                already_returned = total_returned_qty.get(item.id, 0)
                returnable = item.quantity - already_returned
                if returnable < 0:
                    returnable = 0
                food = Food.objects.filter(name=item.product_name).first()
                items_data.append({
                    "id": item.id,
                    "product_name": item.product_name,
                    "price": float(item.wholesale_price),
                    "wholesale_price": float(item.wholesale_price),
                    "wholesale_cost_price": float(item.unit_cost_at_sale) if item.unit_cost_at_sale and float(item.unit_cost_at_sale) > 0 else (float(food.default_purchase_cost) if food else 0),
                    "quantity": item.quantity,
                    "subtotal": float(item.subtotal),
                    "already_returned": already_returned,
                    "returnable": returnable,
                    "barcode": food.barcode if food else "",
                    "product_code": food.product_code if food else "",
                    "sku": food.sku if food else "",
                    "image": food.image.url if food and food.image else "",
                    "stock": food.stock if food else 0,
                })

            customer_name = invoice.wholesale_customer.company_name if invoice.wholesale_customer else "Wholesale Customer"
            customer_phone = invoice.wholesale_customer.phone if invoice.wholesale_customer else ""
            return JsonResponse({
                "success": True,
                "invoice": {
                    "id": invoice.id,
                    "invoice_number": invoice.invoice_number,
                    "customer_name": customer_name,
                    "customer_phone": customer_phone,
                    "date": invoice.created_at.strftime("%d-%m-%Y %I:%M %p"),
                    "type": "wholesale",
                },
                "items": items_data,
            })
        else:
            invoice = get_object_or_404(Invoice, pk=invoice_id)
            items = InvoiceItem.objects.filter(invoice=invoice)
            total_returned_qty = {}
            for ret in ReturnInvoice.objects.filter(original_invoice=invoice):
                for ret_item in ret.items.all():
                    oid = ret_item.original_item_id
                    total_returned_qty[oid] = total_returned_qty.get(oid, 0) + ret_item.quantity

            items_data = []
            for item in items:
                already_returned = total_returned_qty.get(item.id, 0)
                returnable = item.quantity - already_returned
                if returnable < 0:
                    returnable = 0
                food = Food.objects.filter(name=item.product_name).first()
                items_data.append({
                    "id": item.id,
                    "product_name": item.product_name,
                    "price": float(item.price),
                    "quantity": item.quantity,
                    "subtotal": float(item.subtotal),
                    "already_returned": already_returned,
                    "returnable": returnable,
                    "barcode": food.barcode if food else "",
                    "product_code": food.product_code if food else "",
                    "sku": food.sku if food else "",
                    "image": food.image.url if food and food.image else "",
                    "stock": food.stock if food else 0,
                })
            customer_name = invoice.customer_name or (invoice.user.name if invoice.user else "Walk-in Customer")
            customer_phone = invoice.customer_phone or (invoice.user.phone if invoice.user else "")
            return JsonResponse({
                "success": True,
                "invoice": {
                    "id": invoice.id,
                    "invoice_number": invoice.invoice_number,
                    "customer_name": customer_name,
                    "customer_phone": customer_phone,
                    "date": invoice.created_at.strftime("%d-%m-%Y %I:%M %p"),
                    "type": "retail",
                },
                "items": items_data,
            })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@csrf_exempt
@module_access_required('returns')
def return_invoice_save(request):
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Authentication required"}, status=403)
    if not has_permission(request.user, 'can_returns'):
        return JsonResponse({"error": "Permission denied"}, status=403)
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=400)
    try:
        data = json.loads(request.body)
        invoice_id = data.get("invoice_id")
        return_items = data.get("items", [])
        notes = data.get("notes", "")
        invoice_type = data.get("invoice_type", "retail")

        if not invoice_id or not return_items:
            return JsonResponse({"error": "Invoice ID and items are required"}, status=400)

        from django.db import transaction

        if invoice_type == "wholesale":
            w_invoice = get_object_or_404(WholesaleInvoice, pk=invoice_id)
            with transaction.atomic():
                return_invoice = ReturnInvoice.objects.create(
                    wholesale_original_invoice=w_invoice,
                    invoice_type='wholesale',
                    returned_by=request.user,
                    notes=notes,
                    payment_method=w_invoice.payment_method,
                    tax_percentage=w_invoice.tax_percentage,
                    customer_name=w_invoice.wholesale_customer.company_name if w_invoice.wholesale_customer else "Wholesale Customer",
                    customer_phone=w_invoice.wholesale_customer.phone if w_invoice.wholesale_customer else "",
                )

                total_refund = 0
                total_return_qty = 0
                created_items = []

                for ret_item in return_items:
                    item_id = ret_item.get("item_id")
                    return_qty = int(ret_item.get("quantity", 0))
                    if return_qty <= 0:
                        continue
                    original_item = get_object_or_404(WholesaleInvoiceItem, pk=item_id, wholesale_invoice=w_invoice)
                    already_returned = ReturnInvoiceItem.objects.filter(
                        return_invoice__wholesale_original_invoice=w_invoice,
                        original_item_id=item_id,
                    ).aggregate(total=Sum("quantity"))["total"] or 0
                    if return_qty > original_item.quantity - already_returned:
                        raise ValueError(
                            f"Cannot return more than {original_item.quantity - already_returned} of '{original_item.product_name}'"
                        )
                    item_subtotal = round(return_qty * float(original_item.wholesale_price), 2)
                    food = Food.objects.filter(name=original_item.product_name).first()
                    wholesale_cost_price = float(original_item.unit_cost_at_sale) if original_item.unit_cost_at_sale and float(original_item.unit_cost_at_sale) > 0 else (float(food.default_purchase_cost) if food else 0)
                    ReturnInvoiceItem.objects.create(
                        return_invoice=return_invoice,
                        original_item_id=original_item.id,
                        product_name=original_item.product_name,
                        barcode=food.barcode if food else "",
                        sku=food.sku if food else "",
                        price=float(original_item.wholesale_price),
                        quantity=return_qty,
                        subtotal=item_subtotal,
                    )
                    total_refund += item_subtotal
                    total_return_qty += return_qty
                    created_items.append({
                        "product_name": original_item.product_name,
                        "quantity": return_qty,
                        "price": float(original_item.wholesale_price),
                        "subtotal": item_subtotal,
                        "barcode": food.barcode if food else "",
                        "sku": food.sku if food else "",
                    })

                    # Restock inventory via batch system
                    food_item = Food.objects.filter(name=original_item.product_name).first()
                    svc = InventoryValuationService()
                    if food_item:
                        old_stock = food_item.stock
                        food_item.stock = F("stock") + return_qty
                        food_item.save(update_fields=["stock"])
                        food_item.refresh_from_db()
                        svc.return_to_stock(
                            product=food_item,
                            quantity=return_qty,
                            unit_cost=wholesale_cost_price if wholesale_cost_price > 0 else None,
                        )
                        _record_stock_movement(
                            food=food_item,
                            transaction_type='wholesale_return',
                            quantity_change=return_qty,
                            stock_before=old_stock,
                            stock_after=food_item.stock,
                            reference_number=return_invoice.return_number or f"RI-{return_invoice.id:06d}",
                            created_by=request.user if request.user.is_authenticated else None,
                        )

                    # COGS reversal accounting entry using wholesale cost price
                    if wholesale_cost_price > 0:
                        cogs_reversal = round(return_qty * wholesale_cost_price, 2)
                        AccountingEntry.objects.create(
                            entry_type='return',
                            return_invoice=return_invoice,
                            invoice=None,
                            description=f"Inventory increase and COGS decrease for {original_item.product_name} (wholesale return)",
                            debit_account='Inventory',
                            credit_account='COGS',
                            amount=cogs_reversal,
                        )

                return_invoice.total_refund_amount = total_refund
                return_invoice.subtotal_amount = total_refund
                return_invoice.save(update_fields=["total_refund_amount", "subtotal_amount"])

                # Update wholesale invoice returned amount tracking
                w_invoice.total_returned_amount = F("total_returned_amount") + total_refund
                w_invoice.total_returned_qty = F("total_returned_qty") + total_return_qty
                w_invoice.save(update_fields=["total_returned_amount", "total_returned_qty"])

                # Create accounting entries for the refund
                AccountingEntry.objects.create(
                    entry_type='return',
                    return_invoice=return_invoice,
                    invoice=None,
                    description=f"Wholesale Sales Revenue reduction for return {return_invoice.return_number}",
                    debit_account='Wholesale Sales Revenue',
                    credit_account='Sales Returns',
                    amount=total_refund,
                )

                # If there's a wholesale customer, update their account balance (refund)
                if w_invoice.wholesale_customer:
                    cust = w_invoice.wholesale_customer
                    balance_before = float(cust.balance)
                    cust.balance = F("balance") + total_refund
                    cust.save()
                    cust.refresh_from_db()
                    WholesaleAccountTransaction.objects.create(
                        customer=cust,
                        transaction_type='refund',
                        amount=total_refund,
                        balance_before=balance_before,
                        balance_after=float(cust.balance),
                        performed_by=request.user,
                        notes=f"Refund from wholesale return {return_invoice.return_number}",
                    )

            return_invoice.refresh_from_db()
            return JsonResponse({
                "success": True,
                "return_invoice_id": return_invoice.id,
                "return_number": return_invoice.return_number or f"RI-{return_invoice.id:06d}",
                "total_refund_amount": total_refund,
                "net_refund": total_refund,
                "items": created_items,
                "invoice_number": w_invoice.invoice_number,
                "message": f"Wholesale Return {return_invoice.return_number} processed. Refund: {_cs()}{total_refund:.0f}",
            })

        # --- RETAIL RETURN (existing logic) ---
        invoice = get_object_or_404(Invoice, pk=invoice_id)
        if getattr(invoice, '_return_processed', False):
            return JsonResponse({"error": "This return is already being processed"}, status=400)

        with transaction.atomic():
            return_invoice = ReturnInvoice.objects.create(
                original_invoice=invoice,
                invoice_type='retail',
                returned_by=request.user,
                notes=notes,
                payment_method=invoice.payment_method,
                tax_percentage=invoice.tax_percentage,
                customer_name=invoice.customer_name or (invoice.user.name if invoice.user else "Walk-in Customer"),
                customer_phone=invoice.user.phone if invoice.user else "",
            )

            total_refund = 0
            total_discount = float(invoice.qr_offer_discount_amount) + float(invoice.deal_discount_amount)
            total_return_qty = 0
            created_items = []

            for ret_item in return_items:
                item_id = ret_item.get("item_id")
                return_qty = int(ret_item.get("quantity", 0))
                if return_qty <= 0:
                    continue
                original_item = get_object_or_404(InvoiceItem, pk=item_id, invoice=invoice)
                already_returned = ReturnInvoiceItem.objects.filter(
                    return_invoice__original_invoice=invoice,
                    original_item_id=item_id,
                ).aggregate(total=Sum("quantity"))["total"] or 0
                if return_qty > original_item.quantity - already_returned:
                    raise ValueError(
                        f"Cannot return more than {original_item.quantity - already_returned} of '{original_item.product_name}'"
                    )
                item_subtotal = round(return_qty * float(original_item.price), 2)
                food = Food.objects.filter(name=original_item.product_name).first()
                ReturnInvoiceItem.objects.create(
                    return_invoice=return_invoice,
                    original_item_id=original_item.id,
                    product_name=original_item.product_name,
                    barcode=food.barcode if food else "",
                    sku=food.sku if food else "",
                    price=float(original_item.price),
                    quantity=return_qty,
                    subtotal=item_subtotal,
                )
                total_refund += item_subtotal
                total_return_qty += return_qty
                created_items.append({
                    "product_name": original_item.product_name,
                    "quantity": return_qty,
                    "price": float(original_item.price),
                    "subtotal": item_subtotal,
                    "barcode": food.barcode if food else "",
                    "sku": food.sku if food else "",
                })
                food_item = Food.objects.filter(name=original_item.product_name).first()
                svc = InventoryValuationService()
                orig_item_ref = InvoiceItem.objects.filter(pk=item_id, invoice=invoice).first()
                cost_for_return = float(orig_item_ref.unit_cost_at_sale) if orig_item_ref and orig_item_ref.unit_cost_at_sale else None
                if food_item:
                    old_stock = food_item.stock
                    food_item.stock = F("stock") + return_qty
                    food_item.save(update_fields=["stock"])
                    food_item.refresh_from_db()
                    svc.return_to_stock(
                        product=food_item,
                        quantity=return_qty,
                        unit_cost=cost_for_return,
                    )
                    _record_stock_movement(
                        food=food_item,
                        transaction_type='retail_return',
                        quantity_change=return_qty,
                        stock_before=old_stock,
                        stock_after=food_item.stock,
                        reference_number=return_invoice.return_number or f"RI-{return_invoice.id:06d}",
                        created_by=request.user if request.user.is_authenticated else None,
                    )

            invoice_total = float(invoice.total_amount) or 1
            return_ratio = total_refund / invoice_total if invoice_total > 0 else 0
            proportional_discount = round(total_discount * return_ratio, 2)
            proportional_tax = round(float(invoice.tax_amount) * return_ratio, 2)
            net_refund = round(total_refund - proportional_discount + proportional_tax, 2)

            return_invoice.total_refund_amount = total_refund
            return_invoice.subtotal_amount = total_refund
            return_invoice.discount_amount = proportional_discount
            return_invoice.tax_amount = proportional_tax
            return_invoice.save(update_fields=[
                "total_refund_amount", "subtotal_amount",
                "discount_amount", "tax_amount",
            ])

            invoice.total_returned_amount = F("total_returned_amount") + total_refund
            invoice.total_returned_qty = F("total_returned_qty") + total_return_qty
            invoice.save(update_fields=["total_returned_amount", "total_returned_qty"])

            AccountingEntry.objects.create(
                entry_type='return',
                return_invoice=return_invoice,
                invoice=invoice,
                description=f"Sales Revenue reduction for return {return_invoice.return_number}",
                debit_account='Sales Revenue',
                credit_account='Sales Returns',
                amount=total_refund,
            )
            AccountingEntry.objects.create(
                entry_type='return',
                return_invoice=return_invoice,
                invoice=invoice,
                description=f"Cash/Bank reduction for return {return_invoice.return_number}",
                debit_account='Sales Returns',
                credit_account='Cash/Bank',
                amount=net_refund,
            )
            for ret_item in return_items:
                item_id = ret_item.get("item_id")
                return_qty = int(ret_item.get("quantity", 0))
                if return_qty <= 0:
                    continue
                original_item = InvoiceItem.objects.get(pk=item_id, invoice=invoice)
                food = Food.objects.filter(name=original_item.product_name).first()
                unit_cost = float(original_item.unit_cost_at_sale) if original_item.unit_cost_at_sale and float(original_item.unit_cost_at_sale) > 0 else (float(food.default_purchase_cost) if food else 0)
                cogs_reversal = round(return_qty * unit_cost, 2)
                AccountingEntry.objects.create(
                    entry_type='return',
                    return_invoice=return_invoice,
                    invoice=invoice,
                    description=f"Inventory increase and COGS decrease for {original_item.product_name}",
                    debit_account='Inventory',
                    credit_account='COGS',
                    amount=cogs_reversal,
                )

            if invoice.user and invoice.loyalty_points_earned > 0:
                earn_card = LoyaltyCard.objects.filter(user=invoice.user, status='ACTIVE').first()
                if earn_card:
                    pts_to_deduct = round(invoice.loyalty_points_earned * return_ratio)
                    if pts_to_deduct > 0:
                        try:
                            earn_card.total_points = F("total_points") - pts_to_deduct
                            earn_card.remaining_points = F("remaining_points") - pts_to_deduct
                            earn_card.save(update_fields=["total_points", "remaining_points"])
                            LoyaltyTransaction.objects.create(
                                card=earn_card,
                                order_number=return_invoice.return_number,
                                earned_points=0,
                                redeemed_points=pts_to_deduct,
                                remaining_balance=max(0, earn_card.remaining_points - pts_to_deduct),
                                transaction_type='REDEEM',
                            )
                        except Exception:
                            pass

        return_invoice.refresh_from_db()
        return JsonResponse({
            "success": True,
            "return_invoice_id": return_invoice.id,
            "return_number": return_invoice.return_number or f"RI-{return_invoice.id:06d}",
            "total_refund_amount": total_refund,
            "net_refund": net_refund,
            "items": created_items,
            "invoice_number": invoice.invoice_number,
            "message": f"Return {return_invoice.return_number} processed. Refund: {_cs()}{total_refund:.0f}",
        })

    except ValueError as e:
        return JsonResponse({"error": str(e)}, status=400)
    except Exception as e:
        traceback.print_exc()
        return JsonResponse({"error": str(e)}, status=400)


@login_required
@module_access_required('returns')
def return_invoice_history(request):
    if not has_permission(request.user, 'can_returns'):
        messages.error(request, "You do not have permission to view return history.")
        if request.user.is_operator:
            return redirect('users:operator_dashboard')
        return redirect('users:admin_dashboard')
    returns = ReturnInvoice.objects.select_related(
        'original_invoice', 'wholesale_original_invoice', 'returned_by'
    ).prefetch_related('items').all().order_by('-created_at')[:50]
    # Attach a convenience display_invoice attribute for templates
    for r in returns:
        if r.invoice_type == 'wholesale' and r.wholesale_original_invoice:
            r.display_invoice = r.wholesale_original_invoice
        else:
            r.display_invoice = r.original_invoice
    return render(request, "admin/return_history.html", {
        'returns': returns,
        'active_page': 'return_history',
    })


@module_access_required('returns')
def return_invoice_pdf(request, return_id):
    """Enhanced return invoice PDF matching the sales invoice style. Supports retail and wholesale returns."""
    if not request.user.is_authenticated:
        return HttpResponse("Authentication required", status=403)
    if not has_permission(request.user, 'can_returns'):
        return HttpResponse("Permission denied", status=403)
    ret = get_object_or_404(
        ReturnInvoice.objects.select_related('original_invoice', 'wholesale_original_invoice', 'returned_by'),
        pk=return_id
    )
    items = list(ret.items.all())

    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.lib.utils import ImageReader

    width = 60 * mm
    header_height = 65 * mm
    item_per_height = 12 * mm
    summary_height = 55 * mm
    footer_height = 15 * mm

    height = (
        header_height
        + (len(items) * item_per_height)
        + summary_height
        + footer_height
    )

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(width, height))
    pdf.setTitle(ret.return_number or f"RI-{ret.id:06d}")

    w, h = width, height
    DARK = HexColor("#111827")
    GRAY = HexColor("#6b7280")
    RED = HexColor("#dc2626")
    GREEN = HexColor("#059669")
    MARGIN = 4 * mm
    right = w - MARGIN

    settings_obj = SystemSetting.objects.filter(pk=1).first()
    company_name = settings_obj.company_name if settings_obj else 'POS'
    company_phone = settings_obj.company_phone if settings_obj else ''
    company_address = settings_obj.company_address if settings_obj else ''
    company_email = settings_obj.company_email if settings_obj else ''

    y = h - 8 * mm
    logo_y = y
    if settings_obj and settings_obj.company_logo:
        try:
            logo_path = settings_obj.company_logo.path
            logo = ImageReader(logo_path)
            logo_size = 10 * mm
            pdf.drawImage(logo, w / 2 - logo_size / 2, logo_y - logo_size, width=logo_size, height=logo_size, preserveAspectRatio=True)
            y = logo_y - logo_size - 3 * mm
        except Exception:
            pass
    pdf.setFillColor(DARK)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawCentredString(w / 2, y, company_name)
    y -= 5 * mm
    pdf.setFont("Helvetica", 7)
    if company_address:
        pdf.drawCentredString(w / 2, y, company_address)
    y -= 4 * mm
    if company_email:
        pdf.drawCentredString(w / 2, y, company_email)
        y -= 4 * mm
    pdf.drawCentredString(w / 2, y, company_phone)
    y -= 5 * mm
    pdf.setStrokeColor(DARK)
    pdf.line(MARGIN, y, right, y)

    is_wholesale = ret.invoice_type == 'wholesale'
    y -= 6 * mm
    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(DARK)
    pdf.setFont("Helvetica-Bold", 8)
    if is_wholesale:
        pdf.setFillColor(GREEN)
        pdf.drawCentredString(w / 2, y, "WHOLESALE RETURN")
    else:
        pdf.setFillColor(RED)
        pdf.drawCentredString(w / 2, y, "RETAIL RETURN")
    pdf.setFillColor(DARK)
    y -= 5 * mm

    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawCentredString(w / 2, y, ret.return_number or f"RI-{ret.id:06d}")
    y -= 6 * mm

    invoice_ref = ret.wholesale_original_invoice if is_wholesale else ret.original_invoice
    invoice_number = invoice_ref.invoice_number if invoice_ref else "N/A"

    pdf.setFont("Helvetica", 7)
    info_lines = [
        ("Return Invoice No.", ret.return_number or f"RI-{ret.id:06d}"),
        ("Original Invoice No.", invoice_number),
        ("Date & Time", ret.created_at.strftime("%d-%m-%Y %I:%M:%S %p")),
        ("Cashier/Operator", ret.returned_by.name or ret.returned_by.email),
    ]

    if is_wholesale and ret.wholesale_original_invoice:
        customer_name = ret.customer_name or ret.wholesale_original_invoice.wholesale_customer.company_name or "Wholesale Customer"
        customer_phone = ret.customer_phone or (ret.wholesale_original_invoice.wholesale_customer.phone if ret.wholesale_original_invoice.wholesale_customer else "")
    else:
        customer_name = ret.customer_name or (ret.original_invoice.customer_name if ret.original_invoice else "") or (
            ret.original_invoice.user.name if ret.original_invoice and ret.original_invoice.user else "Walk-in Customer"
        )
        customer_phone = ret.customer_phone or (ret.original_invoice.user.phone if ret.original_invoice and ret.original_invoice.user else "")

    info_lines.append(("Customer Name", customer_name))
    info_lines.append(("Customer Phone", customer_phone or "N/A"))

    for label, val in info_lines:
        pdf.drawString(MARGIN, y, label)
        pdf.drawRightString(right, y, val)
        y -= 4 * mm

    pdf.line(MARGIN, y, right, y)
    y -= 5 * mm

    pdf.setFont("Helvetica-Bold", 7)
    pdf.drawString(MARGIN, y, "ITEM")
    pdf.drawRightString(right, y, "TOTAL")
    y -= 4 * mm

    pdf.setFont("Helvetica", 7)
    for item in items:
        pdf.drawString(MARGIN, y, item.product_name[:22])
        pdf.drawRightString(right, y, f"{_cs()}{float(item.subtotal):.0f}")
        y -= 4 * mm
        pdf.setFont("Helvetica", 6)
        pdf.setFillColor(GRAY)
        qty_line = f"{item.quantity} x {_cs()}{float(item.price):.0f}"
        if item.barcode:
            qty_line += f" | Barcode: {item.barcode}"
        if item.sku:
            qty_line += f" | SKU: {item.sku}"
        pdf.drawString(MARGIN, y, qty_line[:50])
        pdf.setFillColor(DARK)
        pdf.setFont("Helvetica", 7)
        y -= 5 * mm

    pdf.line(MARGIN, y, right, y)
    y -= 6 * mm

    subtotal = float(ret.subtotal_amount) if ret.subtotal_amount else float(ret.total_refund_amount)
    discount = float(ret.discount_amount) if ret.discount_amount else 0
    tax = float(ret.tax_amount) if ret.tax_amount else 0
    tax_pct = float(ret.tax_percentage) if ret.tax_percentage else 0
    net_refund = subtotal - discount + tax
    if net_refund < 0:
        net_refund = 0
    payment_method = ret.payment_method or (ret.original_invoice.payment_method if ret.original_invoice else "") or (ret.wholesale_original_invoice.payment_method if ret.wholesale_original_invoice else "") or "N/A"

    pdf.setFont("Helvetica", 8)
    pdf.drawString(MARGIN, y, "Subtotal")
    pdf.drawRightString(right, y, f"{_cs()}{subtotal:.0f}")
    y -= 5 * mm

    if discount > 0:
        pdf.setFont("Helvetica", 7)
        pdf.setFillColor(HexColor("#f59e0b"))
        pdf.drawString(MARGIN, y, "Discount")
        pdf.drawRightString(right, y, f"{_cs()}{discount:.0f}")
        y -= 5 * mm

    if tax > 0:
        pdf.setFont("Helvetica", 7)
        pdf.setFillColor(GRAY)
        pdf.drawString(MARGIN, y, "Total Tax")
        pdf.drawRightString(right, y, f"{_cs()}{tax:.0f}")
        y -= 5 * mm

    pdf.setFillColor(DARK)
    pdf.line(MARGIN, y, right, y)
    y -= 6 * mm

    pdf.setFont("Helvetica-Bold", 10)
    pdf.setFillColor(RED)
    pdf.drawString(MARGIN, y, "NET REFUND")
    pdf.drawRightString(right, y, f"{_cs()}{net_refund:.0f}")
    pdf.setFillColor(DARK)
    y -= 6 * mm

    pdf.setFont("Helvetica", 7)
    pdf.drawString(MARGIN, y, "Payment Method")
    pdf.drawRightString(right, y, payment_method.upper())
    y -= 5 * mm
    pdf.drawString(MARGIN, y, "Refund Amount")
    pdf.drawRightString(right, y, f"{_cs()}{net_refund:.0f}")
    y -= 5 * mm

    if ret.notes:
        pdf.setFont("Helvetica", 6)
        pdf.setFillColor(GRAY)
        pdf.drawString(MARGIN, y, f"Notes: {ret.notes[:50]}")
        pdf.setFillColor(DARK)
        y -= 5 * mm

    y -= 5 * mm
    pdf.line(MARGIN, y, right, y)
    y -= 6 * mm
    pdf.setFont("Helvetica-Bold", 7)
    pdf.setFillColor(DARK)
    pdf.drawCentredString(w / 2, y, "Thank you for your business")

    pdf.save()
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{ret.return_number or ret.id}.pdf"'
    return response


@login_required
@module_access_required('returns')
def return_invoice_receipt(request, return_id):
    """Alias for return_invoice_pdf for backward compatibility."""
    return return_invoice_pdf(request, return_id)


# =========================
# WHOLESALE MODULE
# =========================

WHOLESALE_CART_KEY = "wholesale_cart"


def _get_ws_cart(request):
    return request.session.get(WHOLESALE_CART_KEY, [])


def _save_ws_cart(request, cart):
    request.session[WHOLESALE_CART_KEY] = cart
    request.session.modified = True


def _compute_ws_cart_totals(cart):
    subtotal = 0
    total_tax = 0
    for item in cart:
        qty = int(item.get("quantity", 1))
        price = float(item.get("price", 0))
        disc = float(item.get("discount", 0))
        tax_rate = float(item.get("tax_rate", 0))
        line_total = qty * price - disc
        if line_total < 0:
            line_total = 0
        item_tax = round(line_total * tax_rate / 100, 2)
        item["line_total"] = round(line_total, 2)
        item["tax_rate"] = tax_rate
        item["tax_amount"] = item_tax
        item["line_total_with_tax"] = round(line_total + item_tax, 2)
        subtotal += line_total
        total_tax += item_tax
    return round(subtotal, 2), round(total_tax, 2)


def _get_wholesale_effective_price(food):
    price = float(food.wholesale_price) if food.wholesale_price else None
    if price is None:
        return None, ""
    disc_label = ""
    if food.wholesale_discount_type == "percentage" and food.wholesale_discount_value > 0:
        price = round(price - (price * float(food.wholesale_discount_value) / 100), 2)
        disc_label = f"WS {float(food.wholesale_discount_value):.0f}% off"
    elif food.wholesale_discount_type == "fixed" and food.wholesale_discount_value > 0:
        price = round(price - float(food.wholesale_discount_value), 2)
        if price < 0:
            price = 0
        disc_label = f"WS {_cs()}{float(food.wholesale_discount_value):.0f} off"
    return price, disc_label


# Wholesale POS Page
@login_required
@module_access_required('wholesale')
def wholesale_pos_page(request):
    if not has_permission(request.user, 'can_access_wholesale'):
        messages.error(request, "You do not have permission to access wholesale.")
        if request.user.is_operator:
            return redirect('users:operator_dashboard')
        return redirect('users:admin_dashboard')
    categories = Category.objects.filter(is_active=True)
    return render(request, "operator/wholesale_dashboard.html", {
        "categories": categories,
        "can_create_invoice": True,
    })


# Wholesale Customer Search (AJAX)
@login_required
@module_access_required('wholesale')
def wholesale_customer_search(request):
    q = request.GET.get("q", "")
    customers = WholesaleCustomer.objects.filter(is_active=True, account_status='active')
    if q:
        customers = customers.filter(
            Q(company_name__icontains=q) |
            Q(email__icontains=q)
        )
    customers = customers[:15]
    return JsonResponse({
        "customers": [{
            "id": c.id,
            "company_name": c.company_name,
            "contact_person": c.contact_person,
            "email": c.email or "",
            "phone": c.phone or "",
            "balance": float(c.balance),
            "credit_limit": float(c.credit_limit),
        } for c in customers]
    })


# Wholesale Cart AJAX
@login_required
@require_POST
@module_access_required('wholesale')
def wholesale_cart_add(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        data = json.loads(request.body)
        food_id = data.get("food_id")
        qty = int(data.get("quantity", 1))
        tax_rate = float(data.get("tax_rate", 0))
        food = get_object_or_404(Food, pk=food_id)
        if not food.wholesale_price:
            return JsonResponse({"error": f"{food.name} has no wholesale price set"}, status=400)
        if food.stock <= 0:
            return JsonResponse({"error": "Out of stock"}, status=400)
        if qty > food.stock:
            return JsonResponse({"error": f"Insufficient stock. Only {food.stock} available."}, status=400)
        effective_price, disc_label = _get_wholesale_effective_price(food)
        if effective_price is None:
            return JsonResponse({"error": f"{food.name} has no wholesale price"}, status=400)
        cart = _get_ws_cart(request)
        existing = None
        for item in cart:
            if item["food_id"] == food_id:
                existing = item
                break
        if existing:
            new_qty = int(existing["quantity"]) + qty
            if new_qty > food.stock:
                return JsonResponse({"error": f"Insufficient stock. Only {food.stock} available."}, status=400)
            existing["quantity"] = new_qty
        else:
            cart.append({
                "food_id": food_id,
                "name": food.name,
                "price": effective_price,
                "quantity": qty,
                "discount": 0,
                "stock": food.stock,
                "image": food.image.url if food.image else "",
                "discount_label": disc_label,
                "tax_rate": tax_rate,
            })
        _save_ws_cart(request, cart)
        subtotal, _ = _compute_ws_cart_totals(cart)
        return JsonResponse({"success": True, "cart": cart, "subtotal": subtotal, "total_tax": _compute_ws_cart_totals(cart)[1]})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@login_required
@require_POST
@module_access_required('wholesale')
def wholesale_cart_update(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        data = json.loads(request.body)
        food_id = data.get("food_id")
        qty = int(data.get("quantity", 0))
        tax_rate = data.get("tax_rate")
        food = get_object_or_404(Food, pk=food_id)
        cart = _get_ws_cart(request)
        for item in cart:
            if item["food_id"] == food_id:
                if qty <= 0:
                    cart.remove(item)
                else:
                    if qty > food.stock:
                        return JsonResponse({"error": f"Insufficient stock. Only {food.stock} available."}, status=400)
                    item["quantity"] = qty
                    if tax_rate is not None:
                        item["tax_rate"] = float(tax_rate)
                break
        _save_ws_cart(request, cart)
        subtotal, total_tax = _compute_ws_cart_totals(cart)
        return JsonResponse({"success": True, "cart": cart, "subtotal": subtotal, "total_tax": total_tax})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@login_required
@require_POST
@module_access_required('wholesale')
def wholesale_cart_remove(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        data = json.loads(request.body)
        food_id = data.get("food_id")
        cart = _get_ws_cart(request)
        cart[:] = [item for item in cart if item["food_id"] != food_id]
        _save_ws_cart(request, cart)
        subtotal, _ = _compute_ws_cart_totals(cart)
        return JsonResponse({"success": True, "cart": cart, "subtotal": subtotal})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@login_required
@require_POST
@module_access_required('wholesale')
def wholesale_cart_clear(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    request.session.pop(WHOLESALE_CART_KEY, None)
    request.session.pop("wholesale_invoice_discount", None)
    return JsonResponse({"success": True, "cart": []})


@login_required
@module_access_required('wholesale')
def wholesale_cart_data(request):
    if not request.user.is_authenticated or not (request.user.is_operator or request.user.is_staff):
        return JsonResponse({"error": "Access denied"}, status=403)
    cart = _get_ws_cart(request)
    subtotal, total_tax = _compute_ws_cart_totals(cart)
    inv_disc = float(request.session.get("wholesale_invoice_discount", 0))
    after_inv_disc = round(subtotal - inv_disc, 2)
    if after_inv_disc < 0:
        after_inv_disc = 0
    grand_total = round(after_inv_disc + total_tax, 2)
    return JsonResponse({
        "cart": cart,
        "subtotal": subtotal,
        "invoice_discount": inv_disc,
        "total_tax": total_tax,
        "grand_total": grand_total,
    })


@login_required
@require_POST
@module_access_required('wholesale')
def wholesale_set_discount(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    if not has_permission(request.user, 'can_apply_discount'):
        return JsonResponse({"error": "Permission denied"}, status=403)
    try:
        data = json.loads(request.body)
        inv_disc = float(data.get("discount", 0))
        if inv_disc < 0:
            inv_disc = 0
        request.session["wholesale_invoice_discount"] = inv_disc
        request.session.modified = True
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# Wholesale Checkout
@login_required
@require_POST
@module_access_required('wholesale')
def wholesale_checkout(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        from django.db import transaction
        data = json.loads(request.body)
        cart = _get_ws_cart(request)
        if not cart:
            return JsonResponse({"error": "Cart is empty"}, status=400)

        customer_id = data.get("wholesale_customer_id")
        if not customer_id:
            return JsonResponse({"error": "Please select a wholesale customer"}, status=400)

        payment_method = data.get("payment_method", "cash")
        cash_received = float(data.get("cash_received", 0))
        deposit_amount = float(data.get("deposit_amount", 0))
        redeem_amount = float(data.get("redeem_amount", 0))
        credit_used = float(data.get("credit_used", 0))

        with transaction.atomic():
            wholesale_customer = WholesaleCustomer.objects.select_for_update().get(pk=customer_id, is_active=True)
            if wholesale_customer.account_status != 'active':
                return JsonResponse({"error": f"Account is {wholesale_customer.account_status}. Cannot process checkout."}, status=400)
            balance_before = float(wholesale_customer.balance)

            # Build checkout cart and validate wholesale prices
            for item in cart:
                food = Food.objects.get(pk=item["food_id"])
                qty = int(item["quantity"])
                if qty > food.stock:
                    return JsonResponse({"error": f"Insufficient stock for {food.name}: only {food.stock} available"}, status=400)
                if not food.wholesale_price:
                    return JsonResponse({"error": f"{food.name} has no wholesale price set"}, status=400)

            subtotal, total_tax = _compute_ws_cart_totals(cart)
            inv_disc = float(request.session.get("wholesale_invoice_discount", 0))
            after_disc = round(subtotal - inv_disc, 2)
            if after_disc < 0:
                after_disc = 0
            grand_total = round(after_disc + total_tax, 2)

            # Handle deposit before checkout
            if deposit_amount > 0:
                wholesale_customer.balance = F("balance") + deposit_amount
                wholesale_customer.total_deposits = F("total_deposits") + deposit_amount
                wholesale_customer.save()
                wholesale_customer.refresh_from_db()
                WholesaleDeposit.objects.create(
                    customer=wholesale_customer,
                    amount=deposit_amount,
                    payment_method=payment_method,
                    notes="Deposit during checkout",
                    deposit_date=timezone.now(),
                    balance_before=balance_before,
                    balance_after=float(wholesale_customer.balance),
                    created_by=request.user,
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                )
                balance_before = float(wholesale_customer.balance)

            # Handle redeem from account during checkout
            if redeem_amount > 0:
                if redeem_amount > float(wholesale_customer.balance):
                    return JsonResponse({"error": f"Insufficient account balance. Available: {_cs()}{float(wholesale_customer.balance):.0f}, Requested: {_cs()}{redeem_amount:.0f}"}, status=400)
                wholesale_customer.balance = F("balance") - redeem_amount
                wholesale_customer.total_redeemed = F("total_redeemed") + redeem_amount
                wholesale_customer.save()
                wholesale_customer.refresh_from_db()
                redeem_balance_before = float(wholesale_customer.balance) + redeem_amount
                WholesaleAccountTransaction.objects.create(
                    customer=wholesale_customer,
                    transaction_type='redeem',
                    amount=redeem_amount,
                    balance_before=redeem_balance_before,
                    balance_after=float(wholesale_customer.balance),
                    performed_by=request.user,
                    notes="Redeemed during checkout",
                )

            # Calculate payable after redemption
            payable = max(0, grand_total - redeem_amount)

            # Handle credit usage
            if credit_used > 0:
                available_credit = float(wholesale_customer.credit_limit) - float(wholesale_customer.used_credit)
                if credit_used > available_credit:
                    return JsonResponse({"error": f"Insufficient credit. Available: {_cs()}{available_credit:.0f}, Requested: {_cs()}{credit_used:.0f}"}, status=400)
                if credit_used > payable:
                    return JsonResponse({"error": f"Credit amount (Rs {credit_used:.0f}) exceeds payable amount (Rs {payable:.0f})"}, status=400)
                wholesale_customer.used_credit = F("used_credit") + credit_used
                wholesale_customer.save()
                wholesale_customer.refresh_from_db()
                WholesaleAccountTransaction.objects.create(
                    customer=wholesale_customer,
                    invoice=None,
                    transaction_type='credit',
                    amount=credit_used,
                    balance_before=float(wholesale_customer.balance),
                    balance_after=float(wholesale_customer.balance),
                    performed_by=request.user,
                    notes=f"Credit used for invoice",
                )

            # Calculate actual payment and change
            cash_needed = max(0, payable - credit_used)
            if cash_received < cash_needed:
                return JsonResponse({"error": f"Insufficient payment. Need {_cs()}{cash_needed:.0f}, received {_cs()}{cash_received:.0f}"}, status=400)
            change_due = round(cash_received - cash_needed, 2)

            # Determine payment status
            paid_amount = round(min(cash_received, cash_needed) + redeem_amount, 2)
            if credit_used > 0:
                remaining_due = round(credit_used, 2)
            else:
                remaining_due = round(max(0, grand_total - paid_amount), 2)
            if credit_used > 0:
                payment_status = 'credit_due'
            elif remaining_due > 0:
                payment_status = 'partial'
            else:
                payment_status = 'paid'

            # Deduct stock and record batch consumption
            stock_movements = []
            svc = InventoryValuationService()
            for item in cart:
                food = Food.objects.get(pk=item["food_id"])
                qty = int(item["quantity"])
                if qty > food.stock:
                    return JsonResponse({"error": f"Insufficient stock for {food.name}: only {food.stock} available"}, status=400)
                stock_before = food.stock
                food.stock -= qty
                if food.stock < 0:
                    food.stock = 0
                food.save()
                stock_movements.append({
                    "food": food,
                    "qty": qty,
                    "stock_before": stock_before,
                    "stock_after": food.stock,
                })
                item['_batch_consumptions'] = svc.consume(food, qty)

            # Create wholesale invoice
            highest_tax_pct = max([float(item.get("tax_rate", 0)) for item in cart] or [0])
            invoice = WholesaleInvoice.objects.create(
                wholesale_customer=wholesale_customer,
                payment_method=payment_method,
                tax_percentage=highest_tax_pct,
                tax_amount=total_tax,
                subtotal_amount=subtotal,
                total_amount=grand_total,
                discount_amount=inv_disc,
                deposit_amount=deposit_amount,
                redeemed_amount=redeem_amount,
                balance_before=balance_before - deposit_amount if deposit_amount > 0 else balance_before,
                cash_received=cash_received,
                change_due=change_due,
                credit_used=credit_used,
                paid_amount=paid_amount,
                remaining_due=remaining_due,
                payment_status=payment_status,
            )

            for item in cart:
                food = Food.objects.filter(pk=item["food_id"]).first()
                ws_item = WholesaleInvoiceItem.objects.create(
                    wholesale_invoice=invoice,
                    product_name=item["name"],
                    wholesale_price=float(item["price"]),
                    quantity=int(item["quantity"]),
                    subtotal=float(item["line_total"]),
                    tax_percentage=float(item.get("tax_rate", 0)),
                    tax_amount=float(item.get("tax_amount", 0)),
                )
                total_cogs = Decimal('0')
                for batch, take, unit_cost in item.get('_batch_consumptions', []):
                    SaleItemCost.objects.create(
                        inventory_batch=batch,
                        wholesale_invoice_item=ws_item,
                        quantity=take,
                        unit_cost=unit_cost,
                    )
                    total_cogs += Decimal(str(take)) * unit_cost
                if item.get('_batch_consumptions'):
                    qty = int(item['quantity'])
                    ws_item.unit_cost_at_sale = float(total_cogs / Decimal(str(qty)))
                    ws_item.save(update_fields=['unit_cost_at_sale'])

            invoice.generate_qr_code(request)
            invoice.save()

            # Record payment transaction
            remaining_balance = float(wholesale_customer.balance)
            WholesaleAccountTransaction.objects.create(
                customer=wholesale_customer,
                invoice=invoice,
                transaction_type='payment',
                amount=paid_amount,
                balance_before=remaining_balance,
                balance_after=remaining_balance,
                performed_by=request.user,
                notes=f"Payment for invoice {invoice.invoice_number}",
            )

        # Record stock movements with invoice reference
        for sm in stock_movements:
            _record_stock_movement(
                food=sm["food"],
                transaction_type='wholesale_sale',
                quantity_change=-sm["qty"],
                stock_before=sm["stock_before"],
                stock_after=sm["stock_after"],
                reference_number=invoice.invoice_number,
                created_by=request.user if request.user.is_authenticated else None,
            )

        # Clear session
        request.session.pop(WHOLESALE_CART_KEY, None)
        request.session.pop("wholesale_invoice_discount", None)

        return JsonResponse({
            "success": True,
            "invoice_no": invoice.invoice_number,
            "uuid_token": invoice.uuid_token,
            "payment_status": payment_status,
            "credit_used": credit_used,
            "remaining_due": remaining_due,
        })
    except Exception as e:
        traceback.print_exc()
        return JsonResponse({"error": str(e)}, status=400)


# Wholesale Invoice PDF
@module_access_required('wholesale')
def wholesale_invoice_pdf(request, uuid_token):
    invoice = WholesaleInvoice.objects.filter(uuid_token=uuid_token).first()
    if not invoice:
        return HttpResponse("Invoice not found", status=404)

    items = list(invoice.items.all())
    cust = invoice.wholesale_customer

    width = 60 * mm
    header_height = 65 * mm
    item_per_height = 9 * mm
    summary_height = 40 * mm

    deposit_amt = float(invoice.deposit_amount) if invoice.deposit_amount else 0
    redeem_amt = float(invoice.redeemed_amount) if invoice.redeemed_amount else 0
    acct_section_lines = 1  # prev balance + current balance = min 2
    if deposit_amt > 0:
        acct_section_lines += 1
    if redeem_amt > 0:
        acct_section_lines += 1
    acct_height = 10 * mm + acct_section_lines * 4 * mm

    payment_height = 10 * mm + 4 * 4 * mm  # header + 4 rows
    qr_height = 40 * mm
    footer_height = 15 * mm

    height = (
        header_height
        + (len(items) * item_per_height)
        + summary_height
        + acct_height
        + payment_height
        + qr_height
        + footer_height
    )

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(width, height))
    pdf.setTitle(invoice.invoice_number)

    w, h = width, height
    DARK = HexColor("#111827")
    GRAY = HexColor("#6b7280")
    GREEN = HexColor("#059669")
    MARGIN = 4 * mm
    right = w - MARGIN
    label_x = MARGIN
    val_x = 32 * mm

    settings_obj = SystemSetting.objects.filter(pk=1).first()
    company_name = settings_obj.company_name if settings_obj else 'POS'
    company_phone = settings_obj.company_phone if settings_obj else ''
    company_address = settings_obj.company_address if settings_obj else ''
    company_email = settings_obj.company_email if settings_obj else ''

    # HEADER - Wholesale
    y = h - 8 * mm
    logo_y = y
    if settings_obj and settings_obj.company_logo:
        try:
            logo_path = settings_obj.company_logo.path
            logo = ImageReader(logo_path)
            logo_size = 10 * mm
            pdf.drawImage(logo, w / 2 - logo_size / 2, logo_y - logo_size, width=logo_size, height=logo_size, preserveAspectRatio=True)
            y = logo_y - logo_size - 3 * mm
        except Exception:
            pass
    pdf.setFillColor(GREEN)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawCentredString(w / 2, y, f"{company_name} - WHOLESALE")

    y -= 5 * mm
    pdf.setFillColor(DARK)
    pdf.setFont("Helvetica", 7)
    if company_address:
        pdf.drawCentredString(w / 2, y, company_address)

    y -= 4 * mm
    if company_email:
        pdf.drawCentredString(w / 2, y, company_email)
        y -= 4 * mm
    pdf.drawCentredString(w / 2, y, company_phone)

    y -= 5 * mm
    pdf.setStrokeColor(GREEN)
    pdf.line(MARGIN, y, right, y)

    # CUSTOMER INFO
    y -= 6 * mm
    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(DARK)
    customer_lines = [
        ("Company", cust.company_name),
        ("Contact", cust.contact_person),
        ("Phone", cust.phone or "--------"),
        ("Customer ID", f"#{cust.id}"),
    ]
    for label, val in customer_lines:
        pdf.drawString(MARGIN, y, label)
        pdf.drawRightString(right, y, val)
        y -= 4 * mm

    pdf.line(MARGIN, y, right, y)
    y -= 5 * mm

    # INVOICE INFO
    pdf.setFont("Helvetica", 7)
    info_lines = [
        ("Invoice #", invoice.invoice_number),
        ("Date", timezone.localtime(invoice.created_at).strftime("%d-%m-%Y %I:%M:%S %p")),
        ("Payment Method", (invoice.payment_method or "N/A").upper()),
    ]
    for label, val in info_lines:
        pdf.drawString(MARGIN, y, label)
        pdf.drawRightString(right, y, val)
        y -= 4 * mm

    pdf.line(MARGIN, y, right, y)

    # ITEMS
    y -= 5 * mm
    subtotal = 0
    pdf.setFont("Helvetica-Bold", 7)
    pdf.drawString(MARGIN, y, "ITEM")
    pdf.drawRightString(right, y, "TOTAL")
    y -= 4 * mm

    pdf.setFont("Helvetica", 7)
    for item in items:
        item_tax = float(item.tax_amount) if item.tax_amount else 0
        item_tax_pct = float(item.tax_percentage) if item.tax_percentage else 0
        subtotal += float(item.subtotal)
        name = item.product_name[:22]
        pdf.drawString(MARGIN, y, name)
        pdf.drawRightString(right, y, f"{_cs()}{float(item.subtotal):.0f}")
        y -= 4 * mm
        pdf.setFont("Helvetica", 6)
        pdf.setFillColor(GRAY)
        qty_line = f"{item.quantity} x {_cs()}{float(item.wholesale_price):.0f}"
        if item_tax > 0:
            qty_line += f"  |  Tax ({item_tax_pct:.0f}%): {_cs()}{item_tax:.0f}"
        pdf.drawString(MARGIN, y, qty_line)
        pdf.setFillColor(DARK)
        pdf.setFont("Helvetica", 7)
        y -= 5 * mm

    pdf.line(MARGIN, y, right, y)

    # SUMMARY
    y -= 6 * mm
    tax_amount = float(invoice.tax_amount) if invoice.tax_amount else 0
    tax_pct = float(invoice.tax_percentage) if invoice.tax_percentage else 0
    grand_total = float(invoice.total_amount)
    sub_amt = float(invoice.subtotal_amount) if invoice.subtotal_amount else subtotal
    disc_amt = float(invoice.discount_amount) if invoice.discount_amount else 0
    inv_total = sub_amt - disc_amt + tax_amount
    if inv_total < 0:
        inv_total = 0

    pdf.setFont("Helvetica", 7)
    pdf.drawString(label_x, y, "Subtotal")
    pdf.drawRightString(right, y, f"{_cs()}{sub_amt:.0f}")
    y -= 4 * mm

    if disc_amt > 0:
        pdf.setFillColor(GREEN)
        pdf.drawString(label_x, y, "Discount")
        pdf.drawRightString(right, y, f"-{_cs()}{disc_amt:.0f}")
        pdf.setFillColor(DARK)
        y -= 4 * mm

    if tax_amount > 0:
        pdf.setFillColor(GRAY)
        pdf.drawString(label_x, y, "Total Tax")
        pdf.drawRightString(right, y, f"{_cs()}{tax_amount:.0f}")
        pdf.setFillColor(DARK)
        y -= 4 * mm

    pdf.line(MARGIN, y, right, y)
    y -= 5 * mm

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(label_x, y, "INVOICE TOTAL")
    pdf.drawRightString(right, y, f"{_cs()}{inv_total:.0f}")
    y -= 7 * mm

    # ==========================
    # CUSTOMER ACCOUNT SUMMARY
    # ==========================
    prev_balance = float(invoice.balance_before)
    remaining_balance = float(cust.balance)

    pdf.setStrokeColor(GREEN)
    pdf.setDash(1, 2)
    pdf.line(MARGIN, y, right, y)
    pdf.setDash()
    y -= 4 * mm

    pdf.setFont("Helvetica-Bold", 7)
    pdf.setFillColor(GREEN)
    pdf.drawString(label_x, y, "CUSTOMER ACCOUNT SUMMARY")
    y -= 4 * mm

    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(DARK)
    pdf.drawString(label_x, y, "Previous Balance")
    pdf.drawRightString(right, y, f"{_cs()}{prev_balance:.0f}")
    y -= 4 * mm

    if deposit_amt > 0:
        pdf.setFillColor(GREEN)
        pdf.drawString(label_x, y, "Deposit")
        pdf.drawRightString(right, y, f"+{_cs()}{deposit_amt:.0f}")
        pdf.setFillColor(DARK)
        y -= 4 * mm

    if redeem_amt > 0:
        pdf.setFillColor(HexColor("#f59e0b"))
        pdf.drawString(label_x, y, "Redeemed")
        pdf.drawRightString(right, y, f"-{_cs()}{redeem_amt:.0f}")
        pdf.setFillColor(DARK)
        y -= 4 * mm

    pdf.setFont("Helvetica-Bold", 8)
    pdf.setFillColor(GREEN)
    pdf.drawString(label_x, y, "Current Balance")
    pdf.drawRightString(right, y, f"{_cs()}{remaining_balance:.0f}")
    y -= 5 * mm

    pdf.setStrokeColor(GREEN)
    pdf.setDash(1, 2)
    pdf.line(MARGIN, y, right, y)
    pdf.setDash()
    y -= 5 * mm

    # ==========================
    # PAYMENT SUMMARY
    # ==========================
    pdf.setFont("Helvetica-Bold", 7)
    pdf.setFillColor(GREEN)
    pdf.drawString(label_x, y, "PAYMENT SUMMARY")
    y -= 4 * mm

    amount_paid = max(0, inv_total - redeem_amt)
    cash_val = float(invoice.cash_received) if invoice.cash_received else amount_paid
    change_val = float(invoice.change_due) if invoice.change_due else max(0, cash_val - amount_paid)
    remaining_due = max(0, amount_paid - cash_val)

    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(DARK)

    pdf.drawString(label_x, y, "Invoice Total")
    pdf.drawRightString(right, y, f"{_cs()}{inv_total:.0f}")
    y -= 4 * mm

    if redeem_amt > 0:
        pdf.setFillColor(HexColor("#f59e0b"))
        pdf.drawString(label_x, y, "Redeemed Amount")
        pdf.drawRightString(right, y, f"-{_cs()}{redeem_amt:.0f}")
        pdf.setFillColor(DARK)
        y -= 4 * mm

    pdf.setFont("Helvetica", 7)
    pdf.drawString(label_x, y, "Amount Paid")
    pdf.drawRightString(right, y, f"{_cs()}{amount_paid:.0f}")
    y -= 4 * mm

    remaining_due = max(0, amount_paid - cash_val)
    if remaining_due > 0:
        pdf.setFillColor(HexColor("#ef4444"))
        pdf.drawString(label_x, y, "Remaining Due")
        pdf.drawRightString(right, y, f"{_cs()}{remaining_due:.0f}")

    if remaining_due > 0:
        pdf.setFillColor(DARK)
    else:
        pdf.setFillColor(GREEN)
        pdf.drawString(label_x, y, "Remaining Due")
        pdf.drawRightString(right, y, _cs() + "0")
        pdf.setFillColor(DARK)
    y -= 5 * mm

    # Cash / Change line
    pdf.line(MARGIN, y, right, y)
    y -= 4 * mm
    pdf.setFont("Helvetica", 6)
    pdf.setFillColor(GRAY)
    pdf.drawString(label_x, y, f"Cash Received: {_cs()}{cash_val:.0f}  |  Change Due: {_cs()}{change_val:.0f}")
    pdf.setFillColor(DARK)
    y -= 6 * mm

    # QR CODE
    y -= 8 * mm
    qr = _generate_invoice_qr_image(invoice, request, prefix="wholesale")
    qr_size = 25 * mm
    qr_x = (w - qr_size) / 2
    pdf.drawImage(qr, qr_x, y - qr_size, width=qr_size, height=qr_size)
    y -= qr_size + 4 * mm

    pdf.setFont("Helvetica", 6)
    pdf.setFillColor(GRAY)
    pdf.drawCentredString(w / 2, y, "Scan To Verify Wholesale Invoice")
    pdf.setFillColor(DARK)

    # FOOTER
    y -= 6 * mm
    pdf.setFont("Helvetica-Bold", 7)
    pdf.setFillColor(DARK)
    pdf.drawCentredString(w / 2, y, "Thank you for your wholesale order")

    pdf.save()
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{invoice.invoice_number}.pdf"'
    return response


# Wholesale Invoice Verify
@login_required
@module_access_required('wholesale')
def wholesale_invoice_verify(request, uuid_token):
    invoice = WholesaleInvoice.objects.filter(uuid_token=uuid_token).first()
    if not invoice:
        return render(request, "food-delivery/invoice_verify.html", {
            "valid": False,
            "error": "Wholesale invoice not found.",
        })
    items = invoice.items.all()
    subtotal = float(invoice.subtotal_amount) if invoice.subtotal_amount else sum(float(item.subtotal) for item in items)
    tax_pct = float(invoice.tax_percentage) if invoice.tax_percentage else 0
    tax = float(invoice.tax_amount) if invoice.tax_amount else 0
    customer = invoice.wholesale_customer
    invoice_date = timezone.localtime(invoice.created_at).strftime("%d-%m-%Y %I:%M:%S %p")
    context = {
        "valid": True,
        "invoice": invoice,
        "items": items,
        "customer_name": customer.company_name,
        "customer_email": customer.email or "N/A",
        "subtotal": subtotal,
        "tax": tax,
        "tax_pct": tax_pct,
        "payment_method": invoice.payment_method or "N/A",
        "invoice_date": invoice_date,
        "is_wholesale": True,
    }
    return render(request, "food-delivery/invoice_verify.html", context)


# =========================
# WHOLESALE CUSTOMER MANAGEMENT
# =========================

@staff_member_required
@module_access_required('wholesale')
def wholesale_customer_list(request):
    if not has_permission(request.user, 'can_access_wholesale'):
        messages.error(request, "You do not have permission to access wholesale.")
        return redirect('users:admin_dashboard')
    customers = WholesaleCustomer.objects.all()
    return render(request, "admin/wholesale_customers.html", {
        "customers": customers,
        "active_page": "wholesale_customers",
    })


@staff_member_required
@module_access_required('wholesale')
def wholesale_customer_add(request):
    if request.method == "POST":
        WholesaleCustomer.objects.create(
            company_name=request.POST.get("company_name"),
            contact_person=request.POST.get("contact_person"),
            email=request.POST.get("email", ""),
            phone=request.POST.get("phone", ""),
            address=request.POST.get("address", ""),
            is_active=request.POST.get("is_active") == "on",
            credit_limit=request.POST.get("credit_limit", 0),
        )
        messages.success(request, "Wholesale customer added successfully")
        return redirect("users:wholesale_customer_list")
    return render(request, "admin/wholesale_customer_form.html", {
        "customer": None,
        "active_page": "wholesale_customers",
    })


@staff_member_required
@module_access_required('wholesale')
def wholesale_customer_edit(request, pk):
    if not has_permission(request.user, 'can_access_wholesale'):
        messages.error(request, "You do not have permission to access wholesale.")
        return redirect('users:admin_dashboard')
    customer = get_object_or_404(WholesaleCustomer, pk=pk)
    if request.method == "POST":
        customer.company_name = request.POST.get("company_name")
        customer.contact_person = request.POST.get("contact_person")
        customer.email = request.POST.get("email", "")
        customer.phone = request.POST.get("phone", "")
        customer.address = request.POST.get("address", "")
        customer.is_active = request.POST.get("is_active") == "on"
        customer.credit_limit = request.POST.get("credit_limit", 0)
        customer.save()
        messages.success(request, "Wholesale customer updated successfully")
        return redirect("users:wholesale_customer_list")
    return render(request, "admin/wholesale_customer_form.html", {
        "customer": customer,
        "active_page": "wholesale_customers",
    })


@staff_member_required
@module_access_required('wholesale')
def wholesale_customer_toggle(request, pk):
    if not has_permission(request.user, 'can_access_wholesale'):
        messages.error(request, "You do not have permission to access wholesale.")
        return redirect('users:admin_dashboard')
    customer = get_object_or_404(WholesaleCustomer, pk=pk)
    customer.is_active = not customer.is_active
    customer.save()
    status = "activated" if customer.is_active else "deactivated"
    messages.success(request, f"Wholesale customer {status}")
    return redirect("users:wholesale_customer_list")


# =========================
# WHOLESALE REPORTS
# =========================

@login_required
@module_access_required('wholesale')
def wholesale_sales_report(request):
    if not has_permission(request.user, 'can_view_reports') and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    invoices = WholesaleInvoice.objects.all()
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            start_utc = timezone.make_aware(start_dt, dt_timezone.utc)
            end_utc = timezone.make_aware(end_dt, dt_timezone.utc)
            invoices = invoices.filter(created_at__range=[start_utc, end_utc])
        except (ValueError, TypeError):
            pass
    total_sales = invoices.aggregate(total=Sum("total_amount"))["total"] or 0
    total_tax = invoices.aggregate(total=Sum("tax_amount"))["total"] or 0
    invoice_count = invoices.count()
    return render(request, "admin/wholesale_sales_report.html", {
        "invoices": invoices[:100],
        "total_sales": float(total_sales),
        "total_tax": float(total_tax),
        "invoice_count": invoice_count,
        "start_date": start_date,
        "end_date": end_date,
        "active_page": "wholesale_reports",
    })


@login_required
@module_access_required('wholesale')
def wholesale_customer_report(request):
    if not has_permission(request.user, 'can_view_reports') and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    customers = WholesaleCustomer.objects.all()
    data = []
    for c in customers:
        invoice_count = WholesaleInvoice.objects.filter(wholesale_customer=c).count()
        total_purchases = WholesaleInvoice.objects.filter(wholesale_customer=c).aggregate(total=Sum("total_amount"))["total"] or 0
        data.append({
            "customer": c,
            "invoice_count": invoice_count,
            "total_purchases": float(total_purchases),
        })
    return render(request, "admin/wholesale_customer_report.html", {
        "customer_data": data,
        "active_page": "wholesale_reports",
    })


@login_required
@module_access_required('wholesale')
def wholesale_invoice_report(request):
    if not has_permission(request.user, 'can_view_reports') and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    invoices = WholesaleInvoice.objects.all().select_related('wholesale_customer')
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            start_utc = timezone.make_aware(start_dt, dt_timezone.utc)
            end_utc = timezone.make_aware(end_dt, dt_timezone.utc)
            invoices = invoices.filter(created_at__range=[start_utc, end_utc])
        except (ValueError, TypeError):
            pass
    total_revenue = invoices.aggregate(total=Sum("total_amount"))["total"] or 0
    return render(request, "admin/wholesale_invoice_report.html", {
        "invoices": invoices[:100],
        "total_revenue": float(total_revenue),
        "start_date": start_date,
        "end_date": end_date,
        "active_page": "wholesale_reports",
    })


@login_required
@module_access_required('wholesale')
def wholesale_revenue_summary(request):
    if not has_permission(request.user, 'can_view_reports') and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = today_start.replace(day=1)
    year_start = today_start.replace(month=1, day=1)

    today_revenue = WholesaleInvoice.objects.filter(created_at__gte=today_start).aggregate(total=Sum("total_amount"))["total"] or 0
    monthly_revenue = WholesaleInvoice.objects.filter(created_at__gte=month_start).aggregate(total=Sum("total_amount"))["total"] or 0
    yearly_revenue = WholesaleInvoice.objects.filter(created_at__gte=year_start).aggregate(total=Sum("total_amount"))["total"] or 0
    total_revenue = WholesaleInvoice.objects.aggregate(total=Sum("total_amount"))["total"] or 0
    invoice_count = WholesaleInvoice.objects.count()
    customer_count = WholesaleCustomer.objects.count()
    active_customer_count = WholesaleCustomer.objects.filter(is_active=True).count()

    return render(request, "admin/wholesale_revenue_summary.html", {
        "today_revenue": float(today_revenue),
        "monthly_revenue": float(monthly_revenue),
        "yearly_revenue": float(yearly_revenue),
        "total_revenue": float(total_revenue),
        "invoice_count": invoice_count,
        "customer_count": customer_count,
        "active_customer_count": active_customer_count,
        "active_page": "wholesale_reports",
    })


# =========================
# WHOLESALE ACCOUNT / WALLET
# =========================

@login_required
@module_access_required('wholesale')
def wholesale_account_info(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    customer_id = request.GET.get("customer_id")
    if not customer_id:
        return JsonResponse({"error": "Customer ID required"}, status=400)
    try:
        customer = WholesaleCustomer.objects.get(pk=customer_id, is_active=True)
        transactions = WholesaleAccountTransaction.objects.filter(customer=customer)[:20]
        return JsonResponse({
            "success": True,
            "balance": float(customer.balance),
            "total_deposits": float(customer.total_deposits),
            "total_redeemed": float(customer.total_redeemed),
            "credit_limit": float(customer.credit_limit),
            "used_credit": float(customer.used_credit),
            "available_credit": float(customer.available_credit),
            "account_status": customer.account_status,
            "transactions": [{
                "id": t.id,
                "type": t.transaction_type,
                "amount": float(t.amount),
                "balance_before": float(t.balance_before),
                "balance_after": float(t.balance_after),
                "notes": t.notes,
                "created_at": t.created_at.strftime("%d-%m-%Y %I:%M %p"),
                "invoice": t.invoice.invoice_number if t.invoice else None,
            } for t in transactions],
        })
    except WholesaleCustomer.DoesNotExist:
        return JsonResponse({"error": "Customer not found"}, status=404)


@login_required
@module_access_required('wholesale')
@require_POST
def wholesale_account_deposit(request):
    if not has_permission(request.user, 'can_create_wholesale_deposit') and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    try:
        from django.db import transaction
        from decimal import Decimal
        data = json.loads(request.body)
        customer_id = data.get("customer_id")
        amount = float(data.get("amount", 0))
        payment_method = data.get("payment_method", "cash")
        notes = data.get("notes", "")

        if not customer_id:
            return JsonResponse({"error": "Customer ID required"}, status=400)
        if amount <= 0:
            return JsonResponse({"error": "Deposit amount must be greater than zero"}, status=400)

        with transaction.atomic():
            customer = WholesaleCustomer.objects.select_for_update().get(pk=customer_id, is_active=True)
            if customer.account_status != 'active':
                return JsonResponse({"error": f"Account is {customer.account_status}. Cannot process deposit."}, status=400)
            balance_before = customer.balance
            customer.balance = F("balance") + Decimal(str(amount))
            customer.total_deposits = F("total_deposits") + Decimal(str(amount))
            customer.save()
            customer.refresh_from_db()

            deposit = WholesaleDeposit.objects.create(
                customer=customer,
                amount=amount,
                payment_method=payment_method,
                notes=notes or "Deposit via POS",
                deposit_date=timezone.now(),
                balance_before=balance_before,
                balance_after=customer.balance,
                created_by=request.user,
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
            )

        return JsonResponse({
            "success": True,
            "balance": float(customer.balance),
            "total_deposits": float(customer.total_deposits),
            "deposit_id": deposit.id,
            "message": f"Deposit of {_cs()}{amount:.0f} successful",
        })
    except WholesaleCustomer.DoesNotExist:
        return JsonResponse({"error": "Customer not found"}, status=404)
    except Exception as e:
        traceback.print_exc()
        return JsonResponse({"error": str(e)}, status=400)


# =========================
# BUSINESS EXPENSE MODULE
# =========================

@login_required
@module_access_required('expenses')
def expense_category_list(request):
    if not has_permission(request.user, 'can_manage_expense_categories') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:admin_dashboard')
    categories = ExpenseCategory.objects.all()
    return render(request, "admin/expense_category_list.html", {
        "categories": categories,
        "active_page": "expenses",
    })


@login_required
@module_access_required('expenses')
def expense_category_add(request):
    if not has_permission(request.user, 'can_manage_expense_categories') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:expense_category_list')
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if name:
            ExpenseCategory.objects.create(name=name)
            messages.success(request, f"Category '{name}' created.")
            return redirect('users:expense_category_list')
        messages.error(request, "Category name is required.")
    return render(request, "admin/expense_category_form.html", {
        "category": None,
        "active_page": "expenses",
    })


@login_required
@module_access_required('expenses')
def expense_category_edit(request, pk):
    if not has_permission(request.user, 'can_manage_expense_categories') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:expense_category_list')
    category = get_object_or_404(ExpenseCategory, pk=pk)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if name:
            category.name = name
            category.save()
            messages.success(request, "Category updated.")
            return redirect('users:expense_category_list')
        messages.error(request, "Category name is required.")
    return render(request, "admin/expense_category_form.html", {
        "category": category,
        "active_page": "expenses",
    })


@login_required
@module_access_required('expenses')
def expense_category_toggle(request, pk):
    if not has_permission(request.user, 'can_manage_expense_categories') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:expense_category_list')
    category = get_object_or_404(ExpenseCategory, pk=pk)
    category.is_active = not category.is_active
    category.save()
    status = "enabled" if category.is_active else "disabled"
    messages.success(request, f"Category '{category.name}' {status}.")
    return redirect('users:expense_category_list')


@login_required
@module_access_required('expenses')
def expense_list(request):
    if not has_permission(request.user, 'can_view_expenses') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:admin_dashboard')
    expenses = BusinessExpense.objects.filter(is_deleted=False).select_related('category', 'created_by')

    category_id = request.GET.get("category", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    payment_method = request.GET.get("payment_method", "")
    created_by_id = request.GET.get("created_by", "")
    amount_min = request.GET.get("amount_min", "")
    amount_max = request.GET.get("amount_max", "")
    search = request.GET.get("search", "")

    if category_id:
        expenses = expenses.filter(category_id=category_id)
    if date_from:
        expenses = expenses.filter(expense_date__gte=date_from)
    if date_to:
        expenses = expenses.filter(expense_date__lte=date_to + " 23:59:59")
    if payment_method:
        expenses = expenses.filter(payment_method=payment_method)
    if created_by_id:
        expenses = expenses.filter(created_by_id=created_by_id)
    if amount_min:
        expenses = expenses.filter(amount__gte=amount_min)
    if amount_max:
        expenses = expenses.filter(amount__lte=amount_max)
    if search:
        expenses = expenses.filter(Q(title__icontains=search) | Q(description__icontains=search))

    categories = ExpenseCategory.objects.filter(is_active=True)
    users = User.objects.filter(is_staff=True)

    return render(request, "admin/expense_list.html", {
        "expenses": expenses,
        "categories": categories,
        "users": users,
        "active_page": "expenses",
        "search": search,
        "filters": {
            "category": category_id, "date_from": date_from, "date_to": date_to,
            "payment_method": payment_method, "created_by": created_by_id,
            "amount_min": amount_min, "amount_max": amount_max,
        },
    })


@login_required
@module_access_required('expenses')
def expense_add(request):
    if not has_permission(request.user, 'can_create_expense') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:expense_list')
    categories = ExpenseCategory.objects.filter(is_active=True)

    if request.method == "POST":
        category_id = request.POST.get("category")
        title = request.POST.get("title", "").strip()
        amount = request.POST.get("amount", "0")
        payment_method = request.POST.get("payment_method", "cash")
        expense_date = request.POST.get("expense_date")
        description = request.POST.get("description", "")

        if not title:
            messages.error(request, "Expense title is required.")
        elif not category_id:
            messages.error(request, "Category is required.")
        else:
            try:
                from django.db import transaction
                with transaction.atomic():
                    category = ExpenseCategory.objects.get(pk=category_id, is_active=True)
                    expense = BusinessExpense.objects.create(
                        category=category,
                        title=title,
                        amount=amount,
                        payment_method=payment_method,
                        expense_date=expense_date,
                        description=description,
                        created_by=request.user,
                    )
                    AccountingEntry.objects.create(
                        entry_type='expense',
                        expense=expense,
                        description=f"Business Expense: {title}",
                        debit_account='Business Expenses',
                        credit_account='Cash/Bank',
                        amount=amount,
                    )
                messages.success(request, f"Expense '{title}' recorded.")
                return redirect('users:expense_list')
            except Exception as e:
                traceback.print_exc()
                messages.error(request, f"Error: {str(e)}")

    return render(request, "admin/expense_form.html", {
        "categories": categories,
        "expense": None,
        "active_page": "expenses",
    })


@login_required
@module_access_required('expenses')
def expense_edit(request, pk):
    if not has_permission(request.user, 'can_edit_expenses') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:expense_list')
    expense = get_object_or_404(BusinessExpense, pk=pk, is_deleted=False)
    categories = ExpenseCategory.objects.filter(is_active=True)

    if request.method == "POST":
        category_id = request.POST.get("category")
        title = request.POST.get("title", "").strip()
        amount = request.POST.get("amount", "0")
        payment_method = request.POST.get("payment_method", "cash")
        expense_date = request.POST.get("expense_date")
        description = request.POST.get("description", "")

        if not title:
            messages.error(request, "Expense title is required.")
        else:
            try:
                expense.category_id = category_id
                expense.title = title
                expense.amount = amount
                expense.payment_method = payment_method
                expense.expense_date = expense_date
                expense.description = description
                expense.save()
                messages.success(request, "Expense updated.")
                return redirect('users:expense_list')
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")

    return render(request, "admin/expense_form.html", {
        "categories": categories,
        "expense": expense,
        "active_page": "expenses",
    })


@login_required
@module_access_required('expenses')
@require_POST
def expense_delete(request, pk):
    if not has_permission(request.user, 'can_delete_expenses') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:expense_list')
    expense = get_object_or_404(BusinessExpense, pk=pk, is_deleted=False)
    expense.is_deleted = True
    expense.deleted_at = timezone.now()
    expense.deleted_by = request.user
    expense.save()
    messages.success(request, f"Expense '{expense.title}' deleted.")
    return redirect('users:expense_list')


@login_required
@module_access_required('expenses')
def expense_detail(request, pk):
    if not has_permission(request.user, 'can_view_expenses') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:expense_list')
    expense = get_object_or_404(
        BusinessExpense.objects.select_related('category', 'created_by', 'deleted_by'),
        pk=pk
    )
    return render(request, "admin/expense_detail.html", {
        "expense": expense,
        "active_page": "expenses",
    })


@login_required
@module_access_required('expenses')
def expense_report(request):
    if not has_permission(request.user, 'can_view_expense_reports') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:admin_dashboard')
    expenses = BusinessExpense.objects.filter(is_deleted=False).select_related('category', 'created_by')

    date_filter = request.GET.get("date_filter", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    today = timezone.localdate()

    if date_filter == "today":
        expenses = expenses.filter(expense_date__date=today)
    elif date_filter == "yesterday":
        expenses = expenses.filter(expense_date__date=today - timedelta(days=1))
    elif date_filter == "week":
        week_start = today - timedelta(days=today.weekday())
        expenses = expenses.filter(expense_date__date__gte=week_start)
    elif date_filter == "month":
        expenses = expenses.filter(expense_date__date__gte=today.replace(day=1))
    elif start_date and end_date:
        expenses = expenses.filter(expense_date__date__gte=start_date, expense_date__date__lte=end_date)

    total_amount = expenses.aggregate(total=Sum("amount"))["total"] or 0
    total_count = expenses.count()
    by_category = expenses.values('category__name').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('-total')
    by_user = expenses.values('created_by__email').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('-total')
    by_payment = expenses.values('payment_method').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('-total')

    return render(request, "admin/expense_report.html", {
        "total_amount": float(total_amount),
        "total_count": total_count,
        "by_category": by_category,
        "by_user": by_user,
        "by_payment": by_payment,
        "date_filter": date_filter,
        "start_date": start_date,
        "end_date": end_date,
        "active_page": "expenses",
    })


@login_required
@module_access_required('expenses')
def expense_export(request):
    if not has_permission(request.user, 'can_view_expense_reports') and not request.user.is_staff:
        return JsonResponse({"error": "Permission denied"}, status=403)
    export_format = request.GET.get("format", "csv")
    expenses = BusinessExpense.objects.filter(is_deleted=False).select_related('category', 'created_by').order_by('-created_at')

    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    if start_date and end_date:
        expenses = expenses.filter(expense_date__date__gte=start_date, expense_date__date__lte=end_date)

    if export_format == "csv":
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="business_expenses.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Title', 'Category', 'Amount', 'Payment Method', 'Expense Date', 'Description', 'Created By', 'Created At'])
        for e in expenses:
            writer.writerow([
                e.id, e.title, e.category.name, float(e.amount), e.payment_method,
                e.expense_date.strftime('%Y-%m-%d %H:%M'), e.description,
                e.created_by.email if e.created_by else '', e.created_at.strftime('%Y-%m-%d %H:%M'),
            ])
        return response

    elif export_format == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []
        title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=18, spaceAfter=6, textColor=colors.HexColor("#1e293b"))
        elements.append(Paragraph("Business Expense Report", title_style))
        if start_date and end_date:
            elements.append(Paragraph(f"Period: {start_date} to {end_date}", ParagraphStyle("Sub", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#64748b"), spaceAfter=20)))
        elements.append(Spacer(1, 10))

        data = [["ID", "Title", "Category", "Amount", "Date", "Created By"]]
        for e in expenses:
            data.append([
                str(e.id), e.title, e.category.name, f"{_cs()}{float(e.amount):,.2f}",
                e.expense_date.strftime("%d-%m-%Y"), e.created_by.email if e.created_by else "",
            ])
        table = Table(data, colWidths=[40, 140, 100, 80, 80, 120])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (2, 0), (3, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ]))
        elements.append(table)
        total = expenses.aggregate(total=Sum("amount"))["total"] or 0
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"<b>Total Expenses: {_cs()}{float(total):,.2f}</b>", styles["Normal"]))
        doc.build(elements)
        buf.seek(0)
        return FileResponse(buf, as_attachment=True, filename="business_expenses.pdf")

    return JsonResponse({"error": "Unsupported format"}, status=400)


# =========================
# WHOLESALE DEPOSIT MODULE
# =========================

@login_required
@module_access_required('wholesale')
def wholesale_deposit_list(request):
    if not has_permission(request.user, 'can_view_wholesale_deposits') and not request.user.is_staff:
        messages.error(request, "You do not have permission to view deposits.")
        return redirect('users:admin_dashboard')
    deposits = WholesaleDeposit.objects.all().select_related('customer', 'created_by')

    customer_id = request.GET.get("customer", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    payment_method = request.GET.get("payment_method", "")
    created_by_id = request.GET.get("created_by", "")
    amount_min = request.GET.get("amount_min", "")
    amount_max = request.GET.get("amount_max", "")

    if customer_id:
        deposits = deposits.filter(customer_id=customer_id)
    if date_from:
        deposits = deposits.filter(deposit_date__gte=date_from)
    if date_to:
        deposits = deposits.filter(deposit_date__lte=date_to + " 23:59:59")
    if payment_method:
        deposits = deposits.filter(payment_method=payment_method)
    if created_by_id:
        deposits = deposits.filter(created_by_id=created_by_id)
    if amount_min:
        deposits = deposits.filter(amount__gte=amount_min)
    if amount_max:
        deposits = deposits.filter(amount__lte=amount_max)

    customers = WholesaleCustomer.objects.filter(is_active=True)
    users = User.objects.filter(is_staff=True)

    return render(request, "admin/wholesale_deposit_list.html", {
        "deposits": deposits,
        "customers": customers,
        "users": users,
        "active_page": "wholesale_deposits",
        "filters": {
            "customer": customer_id,
            "date_from": date_from,
            "date_to": date_to,
            "payment_method": payment_method,
            "created_by": created_by_id,
            "amount_min": amount_min,
            "amount_max": amount_max,
        },
    })


@login_required
@module_access_required('wholesale')
def wholesale_deposit_create(request):
    if not has_permission(request.user, 'can_create_wholesale_deposit') and not request.user.is_staff:
        messages.error(request, "You do not have permission to create deposits.")
        return redirect('users:wholesale_deposit_list')
    customers = WholesaleCustomer.objects.filter(is_active=True, account_status='active')
    from .forms import WholesaleDepositForm

    if request.method == "POST":
        form = WholesaleDepositForm(request.POST)
        if form.is_valid():
            customer_id = form.cleaned_data['customer_id']
            amount = form.cleaned_data['amount']
            payment_method = form.cleaned_data['payment_method']
            reference_number = form.cleaned_data.get('reference_number', '') or ''
            notes = form.cleaned_data.get('notes', '') or ''
            deposit_date = form.cleaned_data['deposit_date']

            try:
                from decimal import Decimal
                from django.db import transaction
                with transaction.atomic():
                    customer = WholesaleCustomer.objects.select_for_update().get(pk=customer_id, is_active=True)
                    if customer.account_status != 'active':
                        messages.error(request, f"Account is {customer.account_status}. Cannot process deposit.")
                        return redirect('users:wholesale_deposit_create')
                    balance_before = customer.balance
                    customer.balance = F("balance") + Decimal(str(amount))
                    customer.total_deposits = F("total_deposits") + Decimal(str(amount))
                    customer.save()
                    customer.refresh_from_db()

                    deposit = WholesaleDeposit.objects.create(
                        customer=customer,
                        amount=amount,
                        payment_method=payment_method,
                        reference_number=reference_number or None,
                        notes=notes,
                        deposit_date=deposit_date,
                        balance_before=balance_before,
                        balance_after=customer.balance,
                        created_by=request.user,
                        ip_address=request.META.get('REMOTE_ADDR'),
                        user_agent=request.META.get('HTTP_USER_AGENT', ''),
                    )
                messages.success(request, f"Deposit of {_cs()}{amount} successful for {customer.company_name}")
                return redirect('users:wholesale_deposit_list')
            except WholesaleCustomer.DoesNotExist:
                messages.error(request, "Customer not found or inactive.")
            except Exception as e:
                traceback.print_exc()
                messages.error(request, f"Error processing deposit: {str(e)}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = WholesaleDepositForm(initial={
            'deposit_date': timezone.now().strftime('%Y-%m-%d %H:%M'),
        })

    return render(request, "admin/wholesale_deposit_form.html", {
        "form": form,
        "customers": customers,
        "active_page": "wholesale_deposits",
    })


@login_required
@module_access_required('wholesale')
def wholesale_deposit_detail(request, pk):
    if not has_permission(request.user, 'can_view_wholesale_deposits') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:wholesale_deposit_list')
    deposit = get_object_or_404(
        WholesaleDeposit.objects.select_related('customer', 'created_by', 'reversed_by'),
        pk=pk
    )
    return render(request, "admin/wholesale_deposit_detail.html", {
        "deposit": deposit,
        "active_page": "wholesale_deposits",
    })


@login_required
@module_access_required('wholesale')
@require_POST
def wholesale_deposit_reverse(request, pk):
    if not has_permission(request.user, 'can_delete_wholesale_deposits') and not request.user.is_staff:
        messages.error(request, "You do not have permission to reverse deposits.")
        return redirect('users:wholesale_deposit_list')
    try:
        reason = request.POST.get("reason", "").strip()
        from django.db import transaction
        with transaction.atomic():
            deposit = WholesaleDeposit.objects.select_for_update().get(pk=pk)
            if deposit.is_reversed:
                messages.error(request, "Deposit has already been reversed.")
                return redirect('users:wholesale_deposit_detail', pk=pk)

            customer = WholesaleCustomer.objects.select_for_update().get(pk=deposit.customer_id)
            balance_before = customer.balance
            customer.balance = F("balance") - deposit.amount
            customer.total_deposits = F("total_deposits") - deposit.amount
            customer.save()
            customer.refresh_from_db()

            deposit.is_reversed = True
            deposit.reversed_at = timezone.now()
            deposit.reversed_by = request.user
            deposit.reverse_reason = reason
            deposit.save()

            WholesaleAccountTransaction.objects.create(
                customer=customer,
                transaction_type='adjustment',
                amount=-deposit.amount,
                balance_before=balance_before,
                balance_after=customer.balance,
                performed_by=request.user,
                notes=f"Reversal of Deposit #{deposit.id}: {reason}" if reason else f"Reversal of Deposit #{deposit.id}",
            )
        messages.success(request, f"Deposit #{deposit.id} reversed successfully.")
    except WholesaleDeposit.DoesNotExist:
        messages.error(request, "Deposit not found.")
    except Exception as e:
        traceback.print_exc()
        messages.error(request, f"Error reversing deposit: {str(e)}")
    return redirect('users:wholesale_deposit_list')


@login_required
@module_access_required('wholesale')
def wholesale_deposit_report(request):
    if not has_permission(request.user, 'can_view_wholesale_deposits') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:admin_dashboard')
    deposits = WholesaleDeposit.objects.filter(is_reversed=False).select_related('customer', 'created_by')

    date_filter = request.GET.get("date_filter", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    today = timezone.localdate()

    if date_filter == "today":
        deposits = deposits.filter(deposit_date__date=today)
    elif date_filter == "yesterday":
        deposits = deposits.filter(deposit_date__date=today - timedelta(days=1))
    elif date_filter == "week":
        week_start = today - timedelta(days=today.weekday())
        deposits = deposits.filter(deposit_date__date__gte=week_start)
    elif date_filter == "month":
        deposits = deposits.filter(deposit_date__date__gte=today.replace(day=1))
    elif start_date and end_date:
        deposits = deposits.filter(deposit_date__date__gte=start_date, deposit_date__date__lte=end_date)

    total_amount = deposits.aggregate(total=Sum("amount"))["total"] or 0
    total_count = deposits.count()
    by_customer = deposits.values('customer__company_name').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('-total')[:20]
    by_user = deposits.values('created_by__email').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('-total')
    by_payment = deposits.values('payment_method').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('-total')

    return render(request, "admin/wholesale_deposit_report.html", {
        "total_amount": float(total_amount),
        "total_count": total_count,
        "by_customer": by_customer,
        "by_user": by_user,
        "by_payment": by_payment,
        "date_filter": date_filter,
        "start_date": start_date,
        "end_date": end_date,
        "active_page": "wholesale_deposits",
    })


@login_required
@module_access_required('wholesale')
def wholesale_deposit_export(request):
    if not has_permission(request.user, 'can_export_wholesale_deposits') and not request.user.is_staff:
        return JsonResponse({"error": "Permission denied"}, status=403)

    export_format = request.GET.get("format", "csv")
    deposits = WholesaleDeposit.objects.all().select_related('customer', 'created_by').order_by('-created_at')

    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")
    if start_date and end_date:
        deposits = deposits.filter(deposit_date__date__gte=start_date, deposit_date__date__lte=end_date)

    if export_format == "csv":
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="wholesale_deposits.csv"'
        writer = csv.writer(response)
        writer.writerow(['Deposit ID', 'Customer', 'Amount', 'Payment Method', 'Reference', 'Deposit Date', 'Balance Before', 'Balance After', 'Created By', 'Notes', 'Reversed', 'Created At'])
        for d in deposits:
            writer.writerow([
                d.id, d.customer.company_name, float(d.amount), d.payment_method,
                d.reference_number or '', d.deposit_date.strftime('%Y-%m-%d %H:%M'),
                float(d.balance_before), float(d.balance_after),
                d.created_by.email if d.created_by else '', d.notes,
                'Yes' if d.is_reversed else 'No',
                d.created_at.strftime('%Y-%m-%d %H:%M'),
            ])
        return response

    return JsonResponse({"error": "Unsupported format"}, status=400)


# =========================
# WHOLESALE CREDIT SETTLEMENTS
# =========================

@login_required
@module_access_required('wholesale')
def wholesale_credit_settlement_list(request):
    if not has_permission(request.user, 'can_manage_wholesale_credit') and not request.user.is_staff:
        messages.error(request, "You do not have permission to manage credit settlements.")
        return redirect('users:admin_dashboard')
    settlements = WholesaleCreditSettlement.objects.all().select_related('customer', 'created_by', 'invoice').order_by('-created_at')
    customers = WholesaleCustomer.objects.filter(is_active=True).order_by('company_name')
    return render(request, "admin/wholesale_credit_settlement_list.html", {
        "settlements": settlements,
        "customers": customers,
        "active_page": "wholesale_credit",
    })


@login_required
@module_access_required('wholesale')
def wholesale_credit_settlement_create(request):
    if not has_permission(request.user, 'can_manage_wholesale_credit') and not request.user.is_staff:
        messages.error(request, "You do not have permission.")
        return redirect('users:admin_dashboard')
    customers = WholesaleCustomer.objects.filter(is_active=True).order_by('company_name')

    if request.method == "POST":
        try:
            from django.db import transaction
            from decimal import Decimal
            customer_id = request.POST.get("customer_id")
            invoice_id = request.POST.get("invoice_id", "")
            amount = Decimal(request.POST.get("amount", "0"))
            settlement_date_str = request.POST.get("settlement_date", "")
            notes = request.POST.get("notes", "")

            if not customer_id:
                messages.error(request, "Customer is required.")
                return redirect('users:wholesale_credit_settlement_create')
            if amount <= 0:
                messages.error(request, "Amount must be greater than zero.")
                return redirect('users:wholesale_credit_settlement_create')

            settlement_date = timezone.now()
            if settlement_date_str:
                try:
                    settlement_date = datetime.strptime(settlement_date_str, '%Y-%m-%d %H:%M')
                    settlement_date = timezone.make_aware(settlement_date, timezone.get_current_timezone())
                except (ValueError, TypeError):
                    pass

            with transaction.atomic():
                customer = WholesaleCustomer.objects.select_for_update().get(pk=customer_id, is_active=True)

                settlement = WholesaleCreditSettlement.objects.create(
                    customer=customer,
                    invoice_id=invoice_id if invoice_id else None,
                    amount=amount,
                    settlement_date=settlement_date,
                    notes=notes,
                    created_by=request.user,
                )

                # Reduce customer's used_credit
                customer.used_credit = F("used_credit") - amount
                customer.save()
                customer.refresh_from_db()

                WholesaleAccountTransaction.objects.create(
                    customer=customer,
                    invoice_id=invoice_id if invoice_id else None,
                    transaction_type='credit_settlement',
                    amount=amount,
                    balance_before=float(customer.balance),
                    balance_after=float(customer.balance),
                    performed_by=request.user,
                    notes=f"Credit settlement #{settlement.id}: {notes or 'Payment received'}",
                )

                # Update invoice if specified
                if invoice_id:
                    inv = WholesaleInvoice.objects.get(pk=invoice_id)
                    inv.remaining_due = F("remaining_due") - amount
                    inv.save()
                    inv.refresh_from_db()
                    if inv.remaining_due <= 0:
                        inv.payment_status = 'paid'
                        inv.remaining_due = 0
                        inv.save()

            messages.success(request, f"Credit settlement of {_cs()}{amount} received from {customer.company_name}.")
            return redirect('users:wholesale_credit_settlement_list')
        except Exception as e:
            traceback.print_exc()
            messages.error(request, f"Error: {str(e)}")

    return render(request, "admin/wholesale_credit_settlement_form.html", {
        "customers": customers,
        "active_page": "wholesale_credit",
    })


@login_required
@module_access_required('wholesale')
def wholesale_credit_settlement_detail(request, pk):
    if not has_permission(request.user, 'can_manage_wholesale_credit') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:wholesale_credit_settlement_list')
    settlement = get_object_or_404(
        WholesaleCreditSettlement.objects.select_related('customer', 'created_by', 'invoice'),
        pk=pk
    )
    return render(request, "admin/wholesale_credit_settlement_detail.html", {
        "settlement": settlement,
        "active_page": "wholesale_credit",
    })


@login_required
@module_access_required('wholesale')
def wholesale_customer_outstanding_invoices(request):
    if not has_permission(request.user, 'can_manage_wholesale_credit') and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    customer_id = request.GET.get("customer_id")
    if not customer_id:
        return JsonResponse({"error": "Customer ID required"}, status=400)
    try:
        customer = WholesaleCustomer.objects.get(pk=customer_id)
        invoices = WholesaleInvoice.objects.filter(
            wholesale_customer=customer,
            payment_status__in=['credit_due', 'partial', 'overdue'],
            remaining_due__gt=0,
        ).order_by('-created_at')
        return JsonResponse({
            "success": True,
            "customer_name": customer.company_name,
            "used_credit": float(customer.used_credit),
            "available_credit": float(customer.available_credit),
            "invoices": [{
                "id": inv.id,
                "invoice_number": inv.invoice_number,
                "total_amount": float(inv.total_amount),
                "remaining_due": float(inv.remaining_due),
                "payment_status": inv.payment_status,
                "created_at": inv.created_at.strftime("%d-%m-%Y"),
            } for inv in invoices],
        })
    except WholesaleCustomer.DoesNotExist:
        return JsonResponse({"error": "Customer not found"}, status=404)


@login_required
@module_access_required('wholesale')
def wholesale_credit_report(request):
    if not has_permission(request.user, 'can_view_wholesale_credit_reports') and not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('users:admin_dashboard')
    customers = WholesaleCustomer.objects.filter(is_active=True).order_by('company_name')
    total_credit_given = WholesaleInvoice.objects.filter(payment_status__in=['credit_due', 'partial', 'overdue']).aggregate(total=Sum("credit_used"))["total"] or 0
    total_credit_settled = WholesaleCreditSettlement.objects.aggregate(total=Sum("amount"))["total"] or 0
    outstanding_count = WholesaleInvoice.objects.filter(payment_status__in=['credit_due', 'partial', 'overdue']).count()

    customer_data = []
    for c in customers:
        if c.used_credit > 0 or c.credit_limit > 0:
            invs = WholesaleInvoice.objects.filter(wholesale_customer=c, payment_status__in=['credit_due', 'partial', 'overdue'])
            total_outstanding = invs.aggregate(total=Sum("remaining_due"))["total"] or 0
            settlements = WholesaleCreditSettlement.objects.filter(customer=c).aggregate(total=Sum("amount"))["total"] or 0
            customer_data.append({
                "customer": c,
                "used_credit": c.used_credit,
                "available_credit": c.available_credit,
                "credit_limit": c.credit_limit,
                "total_settled": settlements,
                "outstanding": total_outstanding,
                "invoice_count": invs.count(),
            })

    return render(request, "admin/wholesale_credit_report.html", {
        "customers": customers,
        "customer_data": customer_data,
        "total_credit_given": float(total_credit_given),
        "total_credit_settled": float(total_credit_settled),
        "outstanding_count": outstanding_count,
        "active_page": "wholesale_credit",
    })


# =========================
# WHOLESALE CUSTOMER DETAIL (with full account history)
# =========================

@staff_member_required
@module_access_required('wholesale')
def wholesale_customer_detail(request, pk):
    customer = get_object_or_404(WholesaleCustomer, pk=pk)
    invoices = WholesaleInvoice.objects.filter(wholesale_customer=customer).order_by('-created_at')[:50]
    transactions = WholesaleAccountTransaction.objects.filter(customer=customer).select_related('performed_by', 'invoice').order_by('-created_at')[:100]
    total_purchases = invoices.aggregate(total=Sum("total_amount"))["total"] or 0
    credit_settlements = WholesaleCreditSettlement.objects.filter(customer=customer).aggregate(total=Sum("amount"))["total"] or 0
    active_credit_invoices = WholesaleInvoice.objects.filter(wholesale_customer=customer, payment_status__in=['credit_due', 'partial', 'overdue'])
    return render(request, "admin/wholesale_customer_detail.html", {
        "customer": customer,
        "invoices": invoices,
        "transactions": transactions,
        "total_purchases": float(total_purchases),
        "credit_settlements": float(credit_settlements),
        "active_credit_invoices": active_credit_invoices,
        "active_page": "wholesale_customers",
    })


# =========================
# WHOLESALE ACCOUNT TRANSACTIONS API (paginated)
# =========================

@login_required
@module_access_required('wholesale')
def wholesale_account_transactions(request):
    if not request.user.is_operator and not request.user.is_staff:
        return JsonResponse({"error": "Access denied"}, status=403)
    customer_id = request.GET.get("customer_id")
    if not customer_id:
        return JsonResponse({"error": "Customer ID required"}, status=400)
    try:
        customer = WholesaleCustomer.objects.get(pk=customer_id)
        page = int(request.GET.get("page", 1))
        per_page = 20
        offset = (page - 1) * per_page
        qs = WholesaleAccountTransaction.objects.filter(customer=customer).select_related('performed_by', 'invoice').order_by('-created_at')
        total = qs.count()
        txns = qs[offset:offset + per_page]
        return JsonResponse({
            "success": True,
            "transactions": [{
                "id": t.id,
                "type": t.transaction_type,
                "amount": float(t.amount),
                "balance_before": float(t.balance_before),
                "balance_after": float(t.balance_after),
                "performed_by": t.performed_by.name if t.performed_by else None,
                "notes": t.notes,
                "invoice": t.invoice.invoice_number if t.invoice else None,
                "created_at": t.created_at.strftime("%d-%m-%Y %I:%M %p"),
            } for t in txns],
            "total": total,
            "page": page,
            "pages": (total + per_page - 1) // per_page,
        })
    except WholesaleCustomer.DoesNotExist:
        return JsonResponse({"error": "Customer not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# =========================
# WHOLESALE ACCOUNT ADJUSTMENT (admin only)
# =========================

@staff_member_required
@module_access_required('wholesale')
@require_POST
def wholesale_account_adjust(request):
    try:
        from django.db import transaction
        data = json.loads(request.body)
        customer_id = data.get("customer_id")
        amount = float(data.get("amount", 0))
        reason = data.get("reason", "").strip()

        if not customer_id:
            return JsonResponse({"error": "Customer ID required"}, status=400)
        if amount == 0:
            return JsonResponse({"error": "Adjustment amount cannot be zero"}, status=400)
        if not reason:
            return JsonResponse({"error": "Reason is required for adjustments"}, status=400)

        with transaction.atomic():
            customer = WholesaleCustomer.objects.select_for_update().get(pk=customer_id)
            balance_before = float(customer.balance)
            new_balance = balance_before + amount
            if new_balance < 0:
                return JsonResponse({"error": f"Adjustment would result in negative balance (current: {_cs()}{balance_before:.0f})"}, status=400)
            customer.balance = F("balance") + amount
            if amount > 0:
                customer.total_deposits = F("total_deposits") + amount
            customer.save()
            customer.refresh_from_db()

            WholesaleAccountTransaction.objects.create(
                customer=customer,
                transaction_type='adjustment',
                amount=amount,
                balance_before=balance_before,
                balance_after=float(customer.balance),
                performed_by=request.user,
                notes=f"Manual adjustment: {reason}",
            )

        return JsonResponse({
            "success": True,
            "balance": float(customer.balance),
            "message": f"Account adjusted by {_cs()}{amount:.0f}. New balance: {_cs()}{float(customer.balance):.0f}",
        })
    except WholesaleCustomer.DoesNotExist:
        return JsonResponse({"error": "Customer not found"}, status=404)
    except Exception as e:
        traceback.print_exc()
        return JsonResponse({"error": str(e)}, status=400)


# =========================
# SOFTWARE OWNER
# =========================
def _get_or_create_settings():
    obj, _ = SystemSetting.objects.get_or_create(pk=1, defaults={'company_name': 'POS'})
    return obj


@login_required
@software_owner_required
def software_owner_dashboard(request):
    return redirect('users:system_settings')


SYSTEM_MODULES = [
    ('pos', 'POS'),
    ('wholesale', 'Wholesale'),
    ('products', 'Products'),
    ('customers', 'Customers'),
    ('inventory', 'Inventory'),
    ('purchases', 'Purchases'),
    ('returns', 'Returns'),
    ('expenses', 'Expenses'),
    ('reports', 'Reports'),
    ('profit_loss', 'Profit & Loss'),
    ('credit_management', 'Credit Management'),
    ('deposits', 'Deposits'),
    ('user_management', 'User Management'),
    ('cash_handling', 'Cash Handling'),
]


def _has_superuser():
    return User.objects.filter(is_staff=True, is_software_owner=False).exists()


@login_required
@software_owner_required
def system_settings(request):
    settings_obj = _get_or_create_settings()

    if request.method == 'POST':
        company_name = request.POST.get('company_name', '').strip()
        company_address = request.POST.get('company_address', '').strip()
        company_phone = request.POST.get('company_phone', '').strip()
        company_email = request.POST.get('company_email', '').strip()
        company_website = request.POST.get('company_website', '').strip()
        tax_number = request.POST.get('tax_number', '').strip()
        tax_label = request.POST.get('tax_label', 'GST').strip()
        tax_rate = request.POST.get('tax_rate', '0')
        currency_symbol = request.POST.get('currency_symbol', '\u20b9').strip()
        currency_code = request.POST.get('currency_code', 'INR').strip()
        invoice_prefix = request.POST.get('invoice_prefix', 'INV-').strip()
        invoice_footer_text = request.POST.get('invoice_footer_text', '').strip()
        receipt_prefix = request.POST.get('receipt_prefix', 'RCP-').strip()
        receipt_footer_text = request.POST.get('receipt_footer_text', '').strip()
        default_payment_terms = request.POST.get('default_payment_terms', '').strip()
        tz_value = request.POST.get('timezone', 'Asia/Kolkata').strip()
        enable_notifications = request.POST.get('enable_notifications') == 'on'
        low_stock_threshold = request.POST.get('low_stock_threshold', '10')
        enabled_modules = request.POST.getlist('enabled_modules')

        default_theme = request.POST.get('default_theme', 'user_choice')
        allow_theme_selection = request.POST.get('allow_theme_selection') == 'on'
        dark_mode_enabled = request.POST.get('dark_mode_enabled') == 'on'

        superuser_name = request.POST.get('superuser_name', '').strip()
        superuser_email = request.POST.get('superuser_email', '').strip().lower()
        superuser_password = request.POST.get('superuser_password', '')
        superuser_confirm = request.POST.get('superuser_confirm', '')

        has_superuser = _has_superuser()
        creating_superuser = not has_superuser and superuser_name and superuser_email

        if not company_name:
            messages.error(request, "Company name is required.")
            return render(request, 'admin/system_settings.html', {
                'settings': settings_obj, 'modules': SYSTEM_MODULES, 'title': 'System Settings'
            })

        if creating_superuser:
            if not superuser_password:
                messages.error(request, "Superuser password is required.")
                return render(request, 'admin/system_settings.html', {
                    'settings': settings_obj, 'modules': SYSTEM_MODULES, 'title': 'System Settings'
                })
            if superuser_password != superuser_confirm:
                messages.error(request, "Superuser passwords do not match.")
                return render(request, 'admin/system_settings.html', {
                    'settings': settings_obj, 'modules': SYSTEM_MODULES, 'title': 'System Settings'
                })
            if User.objects.filter(email=superuser_email).exists():
                messages.error(request, "A user with this email already exists.")
                return render(request, 'admin/system_settings.html', {
                    'settings': settings_obj, 'modules': SYSTEM_MODULES, 'title': 'System Settings'
                })

        try:
            tax_rate = Decimal(str(tax_rate))
            low_stock_threshold = int(low_stock_threshold)
        except (ValueError, decimal.InvalidOperation):
            messages.error(request, "Invalid numeric values.")
            return render(request, 'admin/system_settings.html', {
                'settings': settings_obj, 'modules': SYSTEM_MODULES, 'title': 'System Settings'
            })

        with transaction.atomic():
            logo_changed = False
            if 'company_logo' in request.FILES:
                if settings_obj.company_logo:
                    try:
                        settings_obj.company_logo.delete(save=False)
                    except Exception:
                        pass
                settings_obj.company_logo = request.FILES['company_logo']
                logo_changed = True

            changed_fields = []
            if settings_obj.company_name != company_name:
                changed_fields.append(f"company_name changed")
            if settings_obj.company_address != company_address:
                changed_fields.append('company_address changed')
            if settings_obj.company_phone != company_phone:
                changed_fields.append('company_phone changed')
            if settings_obj.company_email != company_email:
                changed_fields.append('company_email changed')

            settings_obj.company_name = company_name
            settings_obj.company_address = company_address
            settings_obj.company_phone = company_phone
            settings_obj.company_email = company_email
            settings_obj.company_website = company_website
            settings_obj.tax_number = tax_number
            settings_obj.tax_label = tax_label
            settings_obj.tax_rate = tax_rate
            settings_obj.currency_symbol = currency_symbol
            settings_obj.currency_code = currency_code
            settings_obj.invoice_prefix = invoice_prefix
            settings_obj.invoice_footer_text = invoice_footer_text
            settings_obj.receipt_prefix = receipt_prefix
            settings_obj.receipt_footer_text = receipt_footer_text
            settings_obj.default_payment_terms = default_payment_terms
            settings_obj.timezone = tz_value
            settings_obj.enable_notifications = enable_notifications
            settings_obj.low_stock_threshold = low_stock_threshold
            settings_obj.enabled_modules = enabled_modules
            settings_obj.default_theme = default_theme
            settings_obj.allow_theme_selection = allow_theme_selection
            settings_obj.dark_mode_enabled = dark_mode_enabled
            settings_obj.updated_by = request.user

            if creating_superuser:
                superuser = User.objects.create_user(
                    email=superuser_email,
                    name=superuser_name,
                    password=superuser_password,
                    is_staff=True,
                    is_active=True,
                    timezone=tz_value,
                )
                settings_obj.superuser_created = True

            settings_obj.save()

            if creating_superuser:
                AuditLog.objects.create(
                    user=request.user, action='user_creation',
                    description=f"Superuser '{superuser.name}' ({superuser.email}) created by {request.user.email}",
                    ip_address=request.META.get('REMOTE_ADDR'),
                )

            if logo_changed:
                AuditLog.objects.create(
                    user=request.user, action='logo_change',
                    description=f"Company logo changed by {request.user.email}",
                    ip_address=request.META.get('REMOTE_ADDR'),
                )

            if changed_fields:
                AuditLog.objects.create(
                    user=request.user, action='settings_change',
                    description=f"System settings updated by {request.user.email}: {', '.join(changed_fields)}",
                    ip_address=request.META.get('REMOTE_ADDR'),
                )

        msg = "System settings saved successfully."
        if creating_superuser:
            msg += f" Superuser '{superuser.name}' created."
        messages.success(request, msg)
        return redirect('users:system_settings')

    context = {
        'settings': settings_obj,
        'modules': SYSTEM_MODULES,
        'has_superuser': _has_superuser(),
        'title': 'System Settings',
    }
    return render(request, 'admin/system_settings.html', context)


@login_required
def save_theme(request):
    if request.method == 'POST':
        import json
        try:
            data = json.loads(request.body)
        except Exception:
            data = request.POST
        theme = data.get('theme', '')
        if theme not in ('light', 'dark'):
            return JsonResponse({'success': False, 'error': 'Invalid theme'}, status=400)
        request.user.theme_preference = theme
        request.user.save(update_fields=['theme_preference'])
        return JsonResponse({'success': True, 'theme': theme})
    return JsonResponse({'success': False}, status=405)


@login_required
@software_owner_required
def audit_logs(request):
    logs = AuditLog.objects.select_related('user').all()[:200]
    context = {
        'logs': logs,
        'title': 'Audit Logs',
    }
    return render(request, 'admin/audit_logs.html', context)
