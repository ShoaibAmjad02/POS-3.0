import os
import re
import csv
import io
import zipfile
import traceback
import logging
from collections import OrderedDict

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {
    '.xlsx': 'excel',
    '.xls': 'excel',
    '.csv': 'csv',
    '.sql': 'sql',
    '.db': 'sqlite',
    '.sqlite': 'sqlite',
    '.sqlite3': 'sqlite',
    '.zip': 'zip',
}

DELIMITERS = [',', ';', '\t', '|']


def detect_file_type(file_path):
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    return SUPPORTED_EXTENSIONS.get(ext, 'unknown')


def extract_zip(uploaded_file):
    files = []
    with zipfile.ZipFile(uploaded_file) as z:
        for name in z.namelist():
            _, ext = os.path.splitext(name)
            if ext.lower() in SUPPORTED_EXTENSIONS:
                files.append({'name': name, 'content': z.read(name)})
    return files


def _detect_delimiter(sample: str) -> str:
    lines = sample.split('\n')
    if not lines:
        return ','
    header = lines[0].strip()
    counts = {d: header.count(d) for d in DELIMITERS}
    best = max(counts, key=counts.get)
    if counts[best] < 1:
        return ','
    return best


def parse_csv(content):
    try:
        decoded = content.decode('utf-8-sig')
    except UnicodeDecodeError:
        try:
            decoded = content.decode('utf-16')
        except UnicodeDecodeError:
            decoded = content.decode('latin-1')

    delimiter = _detect_delimiter(decoded)
    if delimiter != ',':
        decoded = decoded.replace(delimiter, ',')

    reader = csv.DictReader(io.StringIO(decoded))
    headers = reader.fieldnames or []
    if not headers:
        logger.warning('CSV parse: no headers found in first row')
        return [], [], 0

    rows = []
    for i, row in enumerate(reader):
        if i >= 100:
            break
        rows.append(row)

    non_empty_lines = sum(1 for line in decoded.split('\n') if line.strip())
    logger.info('CSV parse: headers=%s rows=%d lines=%d', headers[:10], len(rows), non_empty_lines)
    return headers, rows, non_empty_lines


def detect_tables_from_csv(headers, sample_rows):
    tables = []
    table = {'name': 'main', 'columns': [], 'row_count': 0, 'sample_rows': sample_rows}
    for h in headers:
        table['columns'].append({'name': h, 'type': _infer_column_type(h, sample_rows)})
    table['row_count'] = len(sample_rows) if sample_rows else 0
    tables.append(table)
    return tables


def _infer_column_type(col_name, sample_rows):
    name_lower = col_name.lower()
    if any(k in name_lower for k in ['price', 'amount', 'cost', 'total', 'salary', 'rate']):
        return 'decimal'
    if any(k in name_lower for k in ['qty', 'quantity', 'stock', 'count', 'points', 'age']):
        return 'integer'
    if any(k in name_lower for k in ['date', 'time', 'created', 'updated', 'expiry']):
        return 'date'
    if any(k in name_lower for k in ['email', 'phone', 'mobile']):
        return 'string'
    if any(k in name_lower for k in ['is_', 'active', 'enabled', 'available']):
        return 'boolean'
    return 'string'


MODULE_PATTERNS = {
    'Products': ['product', 'item', 'food', 'menu', 'sku', 'barcode', 'price', 'cost', 'stock', 'category'],
    'Categories': ['category', 'department', 'group', 'section', 'type', 'class', 'active', 'name'],
    'Inventory': ['stock', 'inventory', 'quantity', 'qty', 'batch', 'warehouse', 'product', 'sku'],
    'Customers': ['customer', 'client', 'buyer', 'patron', 'guest', 'email', 'phone', 'address'],
    'Loyalty Cards': ['loyalty', 'cardnumber', 'membership', 'reward', 'points', 'card', 'member'],
    'Suppliers': ['supplier', 'vendor', 'distributor', 'wholesaler', 'company', 'contact'],
    'Purchases': ['purchase', 'buy', 'procurement', 'po', 'order', 'received', 'supplier', 'product'],
    'Sales': ['sale', 'invoice', 'receipt', 'transaction', 'order', 'checkout', 'customer', 'total'],
    'Expenses': ['expense', 'spend', 'payment', 'cost', 'bill', 'overhead', 'amount', 'date'],
    'Expense Categories': ['expense', 'category', 'type', 'costcategory', 'expense category'],
    'Employees': ['employee', 'staff', 'user', 'personnel', 'worker', 'operator', 'salary', 'department'],
}


def detect_modules(column_names, all_data=None):
    detected = OrderedDict()
    col_text = ' '.join(c.lower() for c in column_names)
    col_text = re.sub(r'[^a-zA-Z0-9 ]', '', col_text)

    if all_data:
        for row in all_data[:50]:
            vals = ' '.join(str(v).lower() for v in row.values() if v)
            col_text += ' ' + re.sub(r'[^a-zA-Z0-9 ]', '', vals)

    logger.info('detect_modules scanning text (first 200 chars): %s', col_text[:200])

    for module, keywords in MODULE_PATTERNS.items():
        score = 0
        for kw in keywords:
            if kw in col_text:
                score += 1
        if score >= 2:
            confidence = min(int((score / len(keywords)) * 100), 95)
            detected[module] = {
                'confidence': confidence,
                'matched_keywords': score,
                'total_keywords': len(keywords),
                'estimated_count': 0,
                'warnings': [],
            }
            logger.info('detect_modules DETECTED %s (score=%d/%d confidence=%d%%)', module, score, len(keywords), confidence)

    if not detected:
        logger.warning('detect_modules found NO modules in columns: %s', column_names)

    return detected


def analyze_file(uploaded_file):
    file_type = detect_file_type(uploaded_file.name)
    result = {
        'file_name': uploaded_file.name,
        'file_size': uploaded_file.size,
        'file_type': file_type,
        'tables': [],
        'modules': OrderedDict(),
        'warnings': [],
        'total_records': 0,
        'status': 'analyzed',
    }

    logger.info('analyze_file START file=%s type=%s size=%d', uploaded_file.name, file_type, uploaded_file.size)

    try:
        content = uploaded_file.read()
        logger.info('analyze_file read %d bytes', len(content))

        if file_type == 'zip':
            extracted = extract_zip(io.BytesIO(content))
            if not extracted:
                result['warnings'].append('No supported files found inside ZIP')
                logger.warning('analyze_file ZIP empty: %s', uploaded_file.name)
                return result
            all_headers = []
            all_rows = []
            for f in extracted:
                ftype = detect_file_type(f['name'])
                logger.info('analyze_file ZIP entry: %s type=%s size=%d', f['name'], ftype, len(f['content']))
                if ftype == 'csv':
                    h, rows, _ = parse_csv(f['content'])
                    all_headers.extend(h)
                    all_rows.extend(rows)
                    tables = detect_tables_from_csv(h, rows)
                    result['tables'].extend(tables)
            if not all_headers and not result['tables']:
                result['warnings'].append('Could not parse any files inside ZIP')
                logger.warning('analyze_file ZIP unparseable: %s', uploaded_file.name)
                return result
            result['modules'] = detect_modules(all_headers, all_rows)
            result['total_records'] = len(all_rows)
            logger.info('analyze_file ZIP done: %d modules %d records', len(result['modules']), result['total_records'])
            return result

        if file_type == 'csv':
            headers, rows, line_count = parse_csv(content)
            logger.info('analyze_file CSV parsed: headers=%s rows=%d', headers[:10], len(rows))
            tables = detect_tables_from_csv(headers, rows)
            result['tables'] = tables
            result['modules'] = detect_modules(headers, rows)
            result['total_records'] = len(rows)
            logger.info('analyze_file CSV done: %d modules %d records', len(result['modules']), result['total_records'])
            return result

        if file_type == 'excel':
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                logger.info('analyze_file Excel sheets: %s', wb.sheetnames)
                all_headers = []
                all_rows = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    headers = []
                    rows = []
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i == 0:
                            headers = [str(c) if c is not None else '' for c in row]
                        else:
                            if any(c is not None for c in row):
                                rows.append({headers[j]: str(row[j]) if j < len(row) and row[j] is not None else '' for j in range(len(headers))})
                                if len(rows) >= 100:
                                    break
                    if headers:
                        logger.info('  sheet=%s headers=%s rows=%d', sheet_name, headers[:10], len(rows))
                        all_headers.extend(headers)
                        all_rows.extend(rows)
                        tables = detect_tables_from_csv(headers, rows)
                        for t in tables:
                            t['name'] = sheet_name
                        result['tables'].extend(tables)
                wb.close()
                result['modules'] = detect_modules(all_headers, all_rows)
                result['total_records'] = len(all_rows)
                logger.info('analyze_file Excel done: %d modules %d records', len(result['modules']), result['total_records'])
                return result
            except ImportError:
                msg = 'openpyxl not installed. Cannot parse Excel files.'
                result['warnings'].append(msg)
                logger.error(msg)
                return result

        if file_type == 'sql':
            try:
                sql_text = content.decode('utf-8', errors='replace')
            except Exception:
                sql_text = content.decode('latin-1')
            tables = _detect_tables_from_sql(sql_text)
            logger.info('analyze_file SQL tables: %d', len(tables))
            result['tables'] = tables
            result['total_records'] = sum(t.get('row_count', 0) for t in tables)
            all_headers = []
            for t in tables:
                all_headers.extend(c['name'] for c in t.get('columns', []))
            result['modules'] = detect_modules(all_headers)
            logger.info('analyze_file SQL done: %d modules', len(result['modules']))
            return result

        if file_type == 'sqlite':
            try:
                import sqlite3
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix='.db') as tmp:
                    tmp.write(content)
                    tmp_path = tmp.name
                conn = sqlite3.connect(tmp_path)
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                db_tables = [r[0] for r in cursor.fetchall()]
                logger.info('analyze_file SQLite tables: %s', db_tables)
                all_headers = []
                for tbl in db_tables:
                    cursor.execute(f'PRAGMA table_info("{tbl}")')
                    cols = [{'name': r[1], 'type': r[2] or 'text'} for r in cursor.fetchall()]
                    cursor.execute(f'SELECT COUNT(*) FROM "{tbl}"')
                    count = cursor.fetchone()[0]
                    sample_rows = []
                    cursor.execute(f'SELECT * FROM "{tbl}" LIMIT 100')
                    col_names = [d[0] for d in cursor.description]
                    for row in cursor.fetchall():
                        sample_rows.append({col_names[i]: str(row[i]) if row[i] is not None else '' for i in range(len(col_names))})
                    all_headers.extend(col_names)
                    result['tables'].append({
                        'name': tbl,
                        'columns': cols,
                        'row_count': count,
                        'sample_rows': sample_rows,
                    })
                conn.close()
                os.unlink(tmp_path)
                result['modules'] = detect_modules(all_headers)
                result['total_records'] = sum(t['row_count'] for t in result['tables'])
                logger.info('analyze_file SQLite done: %d modules %d records', len(result['modules']), result['total_records'])
                return result
            except ImportError:
                msg = 'sqlite3 not available. Cannot parse SQLite files.'
                result['warnings'].append(msg)
                logger.error(msg)
                return result

        result['warnings'].append(f'Unsupported or unparseable file type: {file_type}')
        logger.warning('analyze_file unsupported type: %s', file_type)
        return result

    except Exception as e:
        result['status'] = 'error'
        result['warnings'].append(f'Analysis error: {str(e)}')
        result['traceback'] = traceback.format_exc()
        logger.error('analyze_file ERROR: %s', traceback.format_exc())
        return result


def _detect_tables_from_sql(sql_text):
    tables = []
    create_pattern = re.compile(r'CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?[`"\']?(\w+)[`"\']?\s*\((.*?)\);', re.IGNORECASE | re.DOTALL)
    insert_pattern = re.compile(r'INSERT\s+INTO\s+[`"\']?(\w+)[`"\']?\s*(?:\(.*?\))?\s*VALUES', re.IGNORECASE)

    insert_counts = {}
    for m in insert_pattern.finditer(sql_text):
        tbl = m.group(1)
        insert_counts[tbl] = insert_counts.get(tbl, 0) + 1

    for m in create_pattern.finditer(sql_text):
        tbl_name = m.group(1)
        cols_text = m.group(2)
        col_pattern = re.compile(r'[`"\']?(\w+)[`"\']?\s+(\w+)')
        columns = []
        for cm in col_pattern.finditer(cols_text):
            col_name = cm.group(1)
            col_type = cm.group(2).lower()
            display_type = 'string'
            if any(t in col_type for t in ['int', 'decimal', 'float', 'double', 'numeric', 'real']):
                display_type = 'decimal' if any(t in col_type for t in ['decimal', 'float', 'double', 'numeric']) else 'integer'
            elif any(t in col_type for t in ['date', 'time', 'timestamp']):
                display_type = 'date'
            elif 'bool' in col_type:
                display_type = 'boolean'
            columns.append({'name': col_name, 'type': display_type})
        tables.append({
            'name': tbl_name,
            'columns': columns,
            'row_count': insert_counts.get(tbl_name, 0),
            'sample_rows': [],
        })
    return tables
